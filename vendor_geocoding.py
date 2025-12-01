"""
Utility script to geocode vendor addresses in an Excel sheet while avoiding
repeated API calls when coordinates already exist.
"""
from __future__ import annotations

import os
import time
from typing import Iterable, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"
API_KEY = os.getenv("GOOGLE_API_KEY", "")

VENDOR_FILE = "Vender list.xlsx"
VENDOR_SHEET_NAME = "工作表1"
ADDRESS_COLUMNS = ["address"]


def build_address(row: pd.Series, columns: Iterable[str]) -> str:
    parts = []
    for column in columns:
        value = row.get(column)
        if pd.notna(value):
            parts.append(str(value))
    return ", ".join(parts).strip()


def geocode_address(address: str, api_key: str, retry: int = 3, sleep_sec: float = 0.2) -> Tuple[float | None, float | None]:
    if not address:
        return None, None

    params = {"address": address, "key": api_key}

    for _ in range(retry):
        try:
            response = requests.get(GEOCODE_URL, params=params, timeout=10)
            data = response.json()
            status = data.get("status")
            if status == "OK" and data.get("results"):
                location = data["results"][0]["geometry"]["location"]
                return location["lat"], location["lng"]
            if status in {"OVER_QUERY_LIMIT", "RESOURCE_EXHAUSTED"}:
                time.sleep(2)
                continue
            return None, None
        except Exception:
            time.sleep(sleep_sec)

    return None, None


def ensure_vendor_file(file_path: str, sheet_name: str, address_columns: Iterable[str]) -> pd.DataFrame:
    if not os.path.exists(file_path):
        template_df = pd.DataFrame(columns=list(address_columns) + ["Vendor_Lat", "Vendor_Lon"])
        with pd.ExcelWriter(file_path) as writer:
            template_df.to_excel(writer, sheet_name=sheet_name, index=False)
        raise FileNotFoundError(
            f"找不到 '{file_path}'。已建立範本檔案，請填入資料後重新執行程式。"
        )

    return pd.read_excel(file_path, sheet_name=sheet_name)


def update_coordinates(df: pd.DataFrame, api_key: str) -> pd.DataFrame:
    if "Vendor_Lat" not in df.columns:
        df["Vendor_Lat"] = None
    if "Vendor_Lon" not in df.columns:
        df["Vendor_Lon"] = None

    for idx, row in df.iterrows():
        if pd.notna(row["Vendor_Lat"]) and pd.notna(row["Vendor_Lon"]):
            print(f"第 {idx + 2} 列：已有經緯度，跳過 API 呼叫。")
            continue

        address = build_address(row, ADDRESS_COLUMNS)
        print(f"第 {idx + 2} 列：Geocoding '{address}' ...")
        lat, lon = geocode_address(address, api_key=api_key)
        df.at[idx, "Vendor_Lat"] = lat
        df.at[idx, "Vendor_Lon"] = lon
        time.sleep(0.2)

    return df


def write_back_to_excel(df: pd.DataFrame, file_path: str, sheet_name: str) -> None:
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    header_row = 1
    header_map = {cell.value: cell.column for cell in worksheet[header_row]}

    col_lat = header_map.get("Vendor_Lat")
    if not col_lat:
        col_lat = worksheet.max_column + 1
        worksheet.cell(row=header_row, column=col_lat, value="Vendor_Lat")

    col_lon = header_map.get("Vendor_Lon")
    if not col_lon:
        col_lon = worksheet.max_column + 1
        worksheet.cell(row=header_row, column=col_lon, value="Vendor_Lon")

    for idx, row in df.iterrows():
        excel_row = idx + 2
        worksheet.cell(row=excel_row, column=col_lat, value=row["Vendor_Lat"])
        worksheet.cell(row=excel_row, column=col_lon, value=row["Vendor_Lon"])

    workbook.save(file_path)


if __name__ == "__main__":
    try:
        vendor_df = ensure_vendor_file(VENDOR_FILE, VENDOR_SHEET_NAME, ADDRESS_COLUMNS)
    except FileNotFoundError as error:
        print(error)
    else:
        updated_df = update_coordinates(vendor_df, api_key=API_KEY)
        write_back_to_excel(updated_df, VENDOR_FILE, VENDOR_SHEET_NAME)
        print(f"已將 Vendor_Lat / Vendor_Lon 寫回 {VENDOR_FILE}")
