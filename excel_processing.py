from transport_distance import compute_transport_distance
from datetime import date, datetime
from transport_distance import RouteResult
from transport_distance import compute_transport_distance_from_queries
from transport_distance import geocode_place, transport_type_to_mode
from contextlib import suppress
from dataclasses import dataclass, field
from docx import Document
from docx.shared import Inches
from docxtpl import DocxTemplate, InlineImage
from functools import reduce
from tkinter import filedialog, messagebox
import logging
import json
import math
# import matplotlib
# matplotlib.use('Agg')  # 強制使用不會開視窗的 Agg 後端
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
import pandas as pd
import pythoncom
import re
import shutil
import time
import tkinter as tk
import win32com.client as win32
import xlsxwriter
from openpyxl.styles.colors import Color
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, to_excel
import traceback
import unicodedata
import uuid
from typing import Any, Dict, List

FACTORY_OVERVIEW_INFO = {
    "竹南": {"name": "竹南廠", "address": "苗栗縣竹南鎮公義里科義街1號1、5樓"},
    "竹北": {"name": "竹北廠", "address": "新竹縣竹北市北興里智慧一路1號"},
}
FACTORY_SITE_ALIASES = {
    "": "",
    "竹南": "竹南",
    "竹南廠": "竹南",
    "竹北": "竹北",
    "竹北廠": "竹北",
}

COM_OPEN_RETRY_COUNT = 5
COM_OPEN_RETRY_DELAY_SEC = 0.5
COM_OPEN_TIMEOUT_SEC = 30.0
COM_SAVE_RETRY_COUNT = 10
COM_SAVE_RETRY_DELAY_SEC = 2.0
COM_SAVE_TIMEOUT_SEC = 120.0
COM_REFRESH_SETTLE_SEC = 30.0
COM_REFRESH_TIMEOUT_SEC = 120.0
COM_REFRESH_POLL_SEC = 1.0
COM_REFRESH_RETRY_COUNT = 2
COM_REFRESH_RETRY_DELAY_SEC = 1.0
XL_CONNECTION_TYPE_OLEDB = 1
XL_CONNECTION_TYPE_ODBC = 2
INVALID_EXTERNAL_FORMULA_RE = re.compile(r"\[\d+\][^!]*!", re.IGNORECASE)
INVALID_ERROR_FORMULA_VALUES = {
    "=#NULL!": "#NULL!",
    "=#DIV/0!": "#DIV/0!",
    "=#VALUE!": "#VALUE!",
    "=#REF!": "#REF!",
    "=#NAME?": "#NAME?",
    "=#NUM!": "#NUM!",
    "=#N/A": "#N/A",
}
TRANSPORT_LOCATION_MAPPING_FILENAME = "airport_port_land_location_mapping.xlsx"
TRANSPORT_DISTANCE_CACHE_FILENAME = "transport_distance_cache.json"
TRANSPORT_NEGATIVE_CACHE_TTL_SEC = 24 * 60 * 60
TRANSPORT_GEOCODE_TIMEOUT_SEC = 3.0
TRANSPORT_ROUTE_TIMEOUT_SEC = 5.0
TRANSPORT_ROUTE_RETRY_COUNT = 1
TRANSPORT_ROUTE_RETRY_DELAY_SEC = 0.2
RESOURCES_DIRNAME = "resources"
REPORT_WORKBOOK_TEMPLATE_FILENAME = "report_temp.xlsx"
PLCI_TABLE_FORMAT_FILENAME = "PLCI_table_format.xlsx"
REPORT_TEMPLATE_FILENAMES = {
    "竹南": "智邦-產品碳足跡盤查總報告書_竹南_temp.docx",
    "竹北": "智邦-產品碳足跡盤查總報告書_竹北_temp.docx",
    "越南": "智邦-產品碳足跡盤查總報告書_越南_temp.docx",
}
CARBON_STAGE_OPTIONS = (
    ("Raw Material", "原料取得"),
    ("Manufacturing", "製造"),
    ("Distribution", "配送"),
    ("Usage", "使用"),
    ("Recycling", "廢棄回收"),
)
UNSPECIFIED_EMISSION_COLUMN = "unspecified(kg CO2-eq)"
EMISSION_RESULT_COLUMNS = (
    UNSPECIFIED_EMISSION_COLUMN,
    "fossil(kg CO2-eq)",
    "biogenic(kg CO2-eq)",
    "land transformation (kg CO2-eq)",
)
EMISSION_CONTEXT_SUFFIXES = {
    UNSPECIFIED_EMISSION_COLUMN: "unspecified",
    "fossil(kg CO2-eq)": "fossil",
    "biogenic(kg CO2-eq)": "biogenic",
    "land transformation (kg CO2-eq)": "land",
}
REPORT_GENERAL_STAGE_COLUMNS = {
    "Raw Material": "B",
    "Manufacturing": "C",
    "Distribution": "D",
    "Usage": "E",
    "Recycling": "F",
}
AIR_STAGE_CONTEXT_KEYS = {
    "Raw Material": "Raw_Material_Air",
    "Manufacturing": "Manufacturing_Air",
    "Distribution": "Distribution_Air",
    "Usage": "Usage_Air",
    "Recycling": "Recycling_Air",
}
AIR_TRANSPORT_TYPES = {"air", "空運"}
REPORT_GENERAL_ROW_LABELS = (
    "GWP100 - unspecified",
    "GWP100 - fossil",
    "GWP100 - biogenic",
    "GWP100 - land transformation",
)
CARBON_BOUNDARY_STAGE_MAP = {
    "cradle_to_gate": ("Raw Material", "Manufacturing"),
    "cradle_to_grave": tuple(stage for stage, _ in CARBON_STAGE_OPTIONS),
}
CARBON_BOUNDARY_LABELS = {
    "cradle_to_gate": "搖籃到大門",
    "cradle_to_grave": "搖籃到墳墓",
}
CARBON_BOUNDARY_KEYS_BY_LABEL = {
    label: key for key, label in CARBON_BOUNDARY_LABELS.items()
}
DEFAULT_CARBON_BOUNDARY = "cradle_to_grave"
FILE_PERMISSION_DENIED_ERROR_CODE = "FILE_PERMISSION_DENIED"
FILE_PERMISSION_DENIED_USER_MESSAGE = (
    "無法讀取匯入的 Excel 檔案，可能是檔案正在 Excel、OneDrive 或其他程式中開啟或鎖定。"
    "請關閉該檔案、確認 OneDrive 同步完成後再重新執行。"
)
ROAD_TRANSPORT_TYPES = {
    "road",
    "rord",
    "road transport",
    "local land transport",
    "land",
    "truck",
    "express",
}


@dataclass
class TaskResult:
    ok: bool
    error_code: str = ""
    message: str = ""
    artifacts: Dict[str, Any] = field(default_factory=dict)
    elapsed_ms: int = 0
    warnings: List[str] = field(default_factory=list)


class ExcelComSession:
    def __init__(
        self,
        visible: bool = False,
        display_alerts: bool = False,
        enable_events: bool = False,
        screen_updating: bool = False,
        logger=None,
    ):
        self.visible = visible
        self.display_alerts = display_alerts
        self.enable_events = enable_events
        self.screen_updating = screen_updating
        self.logger = logger
        self.excel = None
        self._opened_workbooks = []

    def __enter__(self):
        pythoncom.CoInitialize()
        self.excel = win32.DispatchEx("Excel.Application")
        self.excel.Visible = self.visible
        self.excel.DisplayAlerts = self.display_alerts
        with suppress(Exception):
            self.excel.EnableEvents = self.enable_events
        with suppress(Exception):
            self.excel.ScreenUpdating = self.screen_updating
        return self

    def __exit__(self, exc_type, exc, tb):
        for workbook in reversed(self._opened_workbooks):
            try:
                workbook.Close(SaveChanges=False)
            except Exception as close_exc:
                self._log_cleanup_warning("workbook.Close", close_exc, workbook)
        self._opened_workbooks.clear()
        if self.excel is not None:
            quit_error = None
            for attempt in range(3):
                try:
                    self.excel.Quit()
                    quit_error = None
                    break
                except Exception as exc:
                    quit_error = exc
                    if (
                        hasattr(exc, "args")
                        and exc.args
                        and exc.args[0] == -2147418111
                        and attempt < 2
                    ):
                        time.sleep(0.5)
                        continue
                    break
            if quit_error is not None:
                self._log_cleanup_warning("excel.Quit", quit_error)
        with suppress(Exception):
            pythoncom.CoUninitialize()
        self.excel = None
        return False

    def _describe_dispatch(self, dispatch_obj) -> str:
        if dispatch_obj is None:
            return "<None>"
        parts = [f"type={type(dispatch_obj).__name__}"]
        username = getattr(dispatch_obj, "_username_", "")
        if username:
            parts.append(f"username={username}")
        with suppress(Exception):
            name = getattr(dispatch_obj, "Name", "")
            if name:
                parts.append(f"name={name}")
        with suppress(Exception):
            full_name = getattr(dispatch_obj, "FullName", "")
            if full_name:
                parts.append(f"full_name={full_name}")
        parts.append(f"repr={dispatch_obj!r}")
        return ", ".join(parts)

    def _log_cleanup_warning(self, action: str, exc: Exception, dispatch_obj=None):
        if self.logger is None:
            return
        detail = self._describe_dispatch(dispatch_obj) if dispatch_obj is not None else ""
        if detail:
            self.logger.warning("ExcelComSession cleanup failed during %s: %s | %s", action, exc, detail)
        else:
            self.logger.warning("ExcelComSession cleanup failed during %s: %s", action, exc)

    def open_workbook(
        self,
        path: str,
        *,
        retry_count: int = COM_OPEN_RETRY_COUNT,
        retry_delay_sec: float = COM_OPEN_RETRY_DELAY_SEC,
        timeout_sec: float = COM_OPEN_TIMEOUT_SEC,
        **kwargs,
    ):
        last_error = None
        start = time.time()
        attempt = 0
        while attempt < max(1, retry_count) and (time.time() - start) < max(0.1, timeout_sec):
            attempt += 1
            try:
                workbook = self.excel.Workbooks.Open(path, **kwargs)
                self._opened_workbooks.append(workbook)
                return workbook
            except Exception as exc:
                last_error = exc
                time.sleep(retry_delay_sec)
        if last_error is not None:
            raise TimeoutError(
                f"Workbook 開啟失敗（path={path}, retries={attempt}, timeout_sec={timeout_sec}）: {last_error}"
            ) from last_error
        raise RuntimeError(f"無法開啟工作簿：{path}")

    def close_workbook(self, workbook, save_changes: bool = False):
        with suppress(Exception):
            workbook.Close(SaveChanges=save_changes)
        with suppress(ValueError):
            self._opened_workbooks.remove(workbook)

    def save_with_retry(
        self,
        workbook,
        *,
        retry_count: int = COM_SAVE_RETRY_COUNT,
        retry_delay_sec: float = COM_SAVE_RETRY_DELAY_SEC,
        timeout_sec: float = COM_SAVE_TIMEOUT_SEC,
    ) -> bool:
        if workbook is None:
            raise ValueError("Workbook 儲存失敗：workbook 為 None")
        start = time.time()
        attempt = 0
        while attempt < max(1, retry_count) and (time.time() - start) < max(0.1, timeout_sec):
            attempt += 1
            try:
                save_member = getattr(workbook, "Save", None)
                if callable(save_member):
                    save_member()
                else:
                    dispid = workbook._oleobj_.GetIDsOfNames("Save")
                    workbook._oleobj_.Invoke(dispid, 0, pythoncom.DISPATCH_METHOD, True)
                return True
            except Exception as exc:
                if isinstance(exc, AttributeError):
                    try:
                        dispid = workbook._oleobj_.GetIDsOfNames("Save")
                        workbook._oleobj_.Invoke(dispid, 0, pythoncom.DISPATCH_METHOD, True)
                        return True
                    except Exception as fallback_exc:
                        exc = AttributeError(
                            f"{exc} | direct_save={fallback_exc} | workbook_proxy={self._describe_dispatch(workbook)}"
                        )
                if hasattr(exc, "args") and exc.args and exc.args[0] == -2147418111:
                    time.sleep(retry_delay_sec)
                    continue
                raise
        raise TimeoutError(
            f"Workbook 儲存逾時（retries={attempt}, timeout_sec={timeout_sec}）"
        )

    def refresh_all_and_wait(
        self,
        workbook,
        *,
        retry_count: int = COM_REFRESH_RETRY_COUNT,
        retry_delay_sec: float = COM_REFRESH_RETRY_DELAY_SEC,
        settle_sec: float = COM_REFRESH_SETTLE_SEC,
        timeout_sec: float = COM_REFRESH_TIMEOUT_SEC,
        poll_sec: float = COM_REFRESH_POLL_SEC,
        cancel_callback=None,
        progress_callback=None,
    ) -> bool:
        last_error = None
        last_progress = 0.0

        def _report_progress(fraction: float):
            nonlocal last_progress
            if not callable(progress_callback):
                return
            try:
                fraction = float(fraction)
            except Exception:
                fraction = 0.0
            fraction = max(0.0, min(1.0, fraction))
            if fraction < last_progress:
                fraction = last_progress
            last_progress = fraction
            progress_callback(fraction)

        total_budget = max(0.1, float(settle_sec) + float(timeout_sec))
        _report_progress(0.0)
        for _ in range(max(1, retry_count)):
            try:
                workbook.RefreshAll()
                waited = 0.0
                while waited < settle_sec:
                    if callable(cancel_callback):
                        cancel_callback()
                    time.sleep(0.5)
                    waited += 0.5
                    _report_progress(waited / total_budget)
                start = time.time()
                while time.time() - start < timeout_sec:
                    if callable(cancel_callback):
                        cancel_callback()
                    all_done = True
                    for sheet in workbook.Worksheets:
                        for qt in sheet.QueryTables:
                            if hasattr(qt, "Refreshing") and qt.Refreshing:
                                all_done = False
                                break
                        if not all_done:
                            break
                    if all_done:
                        _report_progress(1.0)
                        return True
                    elapsed = settle_sec + (time.time() - start)
                    _report_progress(elapsed / total_budget)
                    time.sleep(poll_sec)
                last_error = TimeoutError(
                    f"RefreshAll 逾時（timeout_sec={timeout_sec}, poll_sec={poll_sec}）"
                )
            except Exception as exc:
                last_error = exc
            time.sleep(retry_delay_sec)
        if last_error is not None:
            raise last_error
        return False

class UserCancelledError(Exception):
    pass


class ExcelApp:
    def __init__(self, status_callback=None, progress_callback=None):
        self._no_op_status_callback = lambda *_, **__: None
        self._original_status_callback = status_callback
        if callable(status_callback):
            self.status_callback = status_callback
            self._has_status_callback = True
        else:
            self.status_callback = self._no_op_status_callback
            self._has_status_callback = False
        self.progress_callback = progress_callback
        self.file_path = None
        self.last_error = None
        self.last_run_id = ""
        self.last_technical_summary = ""
        self.error_callback = None
        self._format_cache = {}
        self.factory_site = ""
        self.cancel_event = None
        self.was_cancelled = False
        self.base_dir = self.get_base_dir()
        self.resources_dir = os.path.join(self.base_dir, RESOURCES_DIRNAME)
        self.output_root = os.path.join(self.base_dir, "output")
        self.result_dir = os.path.join(self.output_root, "result")
        self.report_dir = os.path.join(self.output_root, "report")
        self.charts_dir = os.path.join(self.output_root, "charts")
        self.tmp_dir = os.path.join(self.output_root, "tmp")
        self.spreadsheet_dir = os.path.join(self.output_root, "spreadsheet")
        self.logs_dir = os.path.join(self.base_dir, "logs")
        self.cache_dir = os.path.join(self.logs_dir, "cache")
        for folder in (
            self.resources_dir,
            self.output_root,
            self.result_dir,
            self.report_dir,
            self.charts_dir,
            self.tmp_dir,
            self.spreadsheet_dir,
            self.logs_dir,
            self.cache_dir,
        ):
            os.makedirs(folder, exist_ok=True)

        # Backward compatibility for legacy methods.
        self.output_dir = self.result_dir

        self.runtime_config = {
            "open_retry_count": COM_OPEN_RETRY_COUNT,
            "open_retry_delay_sec": COM_OPEN_RETRY_DELAY_SEC,
            "open_timeout_sec": COM_OPEN_TIMEOUT_SEC,
            "save_retry_count": COM_SAVE_RETRY_COUNT,
            "save_retry_delay_sec": COM_SAVE_RETRY_DELAY_SEC,
            "save_timeout_sec": COM_SAVE_TIMEOUT_SEC,
            "refresh_retry_count": COM_REFRESH_RETRY_COUNT,
            "refresh_retry_delay_sec": COM_REFRESH_RETRY_DELAY_SEC,
            "refresh_settle_sec": COM_REFRESH_SETTLE_SEC,
            "refresh_timeout_sec": COM_REFRESH_TIMEOUT_SEC,
            "refresh_poll_sec": COM_REFRESH_POLL_SEC,
            "transport_geocode_timeout_sec": TRANSPORT_GEOCODE_TIMEOUT_SEC,
            "transport_route_timeout_sec": TRANSPORT_ROUTE_TIMEOUT_SEC,
            "transport_route_retry_count": TRANSPORT_ROUTE_RETRY_COUNT,
            "transport_route_retry_delay_sec": TRANSPORT_ROUTE_RETRY_DELAY_SEC,
        }
        self.logger = self._build_logger()
        self._warnings = []
        self.current_run_id = ""
        self.transport_distance_cache_path = os.path.join(self.cache_dir, TRANSPORT_DISTANCE_CACHE_FILENAME)
        self._transport_distance_cache = None
        self._transport_distance_cache_dirty = False
        self._transport_distance_cache_save_warned = False
        self._transport_place_cache = {}
        self._transport_route_cache = {}
        self._transport_endpoint_place_cache = {}
        self._transport_location_mapping = None
        self._transport_mapping_missing_warned = False
        self.carbon_boundary = DEFAULT_CARBON_BOUNDARY
        self._calculation_audit_rows = []
        self._validation_findings = []

    def _build_logger(self):
        logger = logging.getLogger("excel_processing")
        if logger.handlers:
            return logger
        logger.setLevel(logging.INFO)
        log_path = os.path.join(self.logs_dir, "excel_processing.log")
        formatter = logging.Formatter(
            "%(asctime)s %(levelname)s %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        file_handler = logging.FileHandler(log_path, encoding="utf-8")
        file_handler.setFormatter(formatter)
        stream_handler = logging.StreamHandler()
        stream_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.addHandler(stream_handler)
        logger.propagate = False
        return logger

    def _new_run_id(self) -> str:
        return f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"

    def _start_task(self, task_name: str):
        run_id = self._new_run_id()
        self.current_run_id = run_id
        self._warnings = []
        self.last_error = None
        self.last_run_id = run_id
        self.last_technical_summary = ""
        self.was_cancelled = False
        self._transport_endpoint_place_cache = {}
        self.logger.info("[run_id=%s] start %s", run_id, task_name)
        return run_id, time.time()

    def _finish_task_log(self, run_id: str, task_name: str, result: TaskResult):
        level = logging.INFO if result.ok else logging.ERROR
        self.logger.log(
            level,
            "[run_id=%s] finish %s ok=%s code=%s elapsed_ms=%s message=%s",
            run_id,
            task_name,
            result.ok,
            result.error_code,
            result.elapsed_ms,
            result.message,
        )

    def _elapsed_ms(self, started_at: float) -> int:
        return int((time.time() - started_at) * 1000)

    @staticmethod
    def _is_permission_denied_error(exc: Exception) -> bool:
        return isinstance(exc, PermissionError) or getattr(exc, "errno", None) == 13

    def _file_permission_denied_message(self, file_path=None) -> str:
        target_path = file_path or self.file_path
        if target_path:
            return f"{FILE_PERMISSION_DENIED_USER_MESSAGE}\n檔案：{os.path.abspath(target_path)}"
        return FILE_PERMISSION_DENIED_USER_MESSAGE

    def _emit_progress(self, value):
        if not callable(self.progress_callback):
            return
        try:
            value = float(value)
        except Exception:
            return
        value = max(0.0, min(100.0, value))
        self.progress_callback(int(round(value)))

    def _make_stage_progress_callback(self, start, end):
        try:
            start = float(start)
        except Exception:
            start = 0.0
        try:
            end = float(end)
        except Exception:
            end = start
        start = max(0.0, min(100.0, start))
        end = max(start, min(100.0, end))
        span = end - start

        def _callback(fraction):
            try:
                fraction = float(fraction)
            except Exception:
                fraction = 0.0
            fraction = max(0.0, min(1.0, fraction))
            self._emit_progress(start + span * fraction)

        return _callback

    def _warn(self, message: str):
        self._warnings.append(message)
        self.logger.warning("[run_id=%s] %s", self.current_run_id or "-", message)

    def _result_ok(self, message: str, artifacts: Dict[str, Any], started_at: float) -> TaskResult:
        return TaskResult(
            ok=True,
            message=message,
            artifacts=artifacts,
            elapsed_ms=self._elapsed_ms(started_at),
            warnings=list(self._warnings),
        )

    def _result_fail(
        self,
        *,
        error_code: str,
        user_message: str,
        started_at: float,
        exc: Exception = None,
    ) -> TaskResult:
        technical = traceback.format_exc() if exc is not None else ""
        if exc is not None:
            self.logger.exception(
                "[run_id=%s] %s: %s", self.current_run_id or "-", error_code, user_message
            )
        self.last_error = user_message
        self.last_technical_summary = self.summarize_technical_details(technical)
        return TaskResult(
            ok=False,
            error_code=error_code,
            message=user_message,
            artifacts={"technical_details": technical},
            elapsed_ms=self._elapsed_ms(started_at),
            warnings=list(self._warnings),
        )

    @staticmethod
    def summarize_technical_details(technical: str) -> str:
        text = str(technical or "").strip()
        if not text:
            return ""
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        for line in reversed(lines):
            if line.startswith("Traceback "):
                continue
            if line.startswith('File "'):
                continue
            return line
        return lines[-1] if lines else ""

    def _coerce_task_result(
        self,
        *,
        task_name: str,
        started_at: float,
        value,
        success_message: str,
        success_artifacts_fn=None,
    ) -> TaskResult:
        if isinstance(value, TaskResult):
            value.elapsed_ms = value.elapsed_ms or self._elapsed_ms(started_at)
            if not value.warnings:
                value.warnings = list(self._warnings)
            return value
        if isinstance(value, dict):
            ok = bool(value.get("ok"))
            if ok:
                artifacts = dict(value)
                return self._result_ok(success_message, artifacts=artifacts, started_at=started_at)
            cancelled = bool(value.get("cancelled"))
            code = str(value.get("error_code") or ("USER_CANCELLED" if cancelled else "TASK_FAILED"))
            msg = str(value.get("message") or value.get("error") or f"{task_name} 失敗")
            return TaskResult(
                ok=False,
                error_code=code,
                message=msg,
                artifacts=dict(value),
                elapsed_ms=self._elapsed_ms(started_at),
                warnings=list(self._warnings),
            )
        if isinstance(value, bool):
            if value:
                artifacts = success_artifacts_fn() if callable(success_artifacts_fn) else {}
                return self._result_ok(success_message, artifacts=artifacts, started_at=started_at)
            code = "USER_CANCELLED" if self.was_cancelled else "TASK_FAILED"
            msg = self.last_error or f"{task_name} 失敗"
            return TaskResult(
                ok=False,
                error_code=code,
                message=msg,
                artifacts={},
                elapsed_ms=self._elapsed_ms(started_at),
                warnings=list(self._warnings),
            )
        if isinstance(value, str) and value:
            artifacts = success_artifacts_fn() if callable(success_artifacts_fn) else {}
            artifacts = dict(artifacts)
            artifacts.setdefault("path", value)
            return self._result_ok(success_message, artifacts=artifacts, started_at=started_at)
        code = "USER_CANCELLED" if self.was_cancelled else "TASK_FAILED"
        msg = self.last_error or f"{task_name} 回傳了不支援的結果型別"
        return TaskResult(
            ok=False,
            error_code=code,
            message=msg,
            artifacts={"legacy_result": value},
            elapsed_ms=self._elapsed_ms(started_at),
            warnings=list(self._warnings),
        )

    def _chart_path(self, filename: str) -> str:
        path = os.path.join(self.charts_dir, filename)
        if not hasattr(self, "_chart_artifacts"):
            self._chart_artifacts = []
        self._chart_artifacts.append(path)
        return path

    def _cleanup_chart_artifacts(self):
        paths = list(getattr(self, "_chart_artifacts", []))
        self._chart_artifacts = []
        for path in paths:
            with suppress(Exception):
                if os.path.exists(path):
                    os.remove(path)
        # Backward cleanup for legacy files that may have been written to project root.
        with suppress(Exception):
            for name in os.listdir(self.base_dir):
                lower = name.lower()
                if (lower.startswith("bar_chart_") or lower.startswith("pie_chart_")) and lower.endswith(".png"):
                    os.remove(os.path.join(self.base_dir, name))

    def _notify_status(self, message):
        """
        Safely invoke the status callback if it is available.
        如果可用，則安全地呼叫狀態回呼。
        """
        callback = self.status_callback
        if callable(callback):
            self._has_status_callback = callback is not self._no_op_status_callback
        else:
            callback = self._no_op_status_callback
            self._has_status_callback = False
        callback(message)

    def _sanitize_invalid_external_formulas(self, workbook) -> int:
        """
        Remove workbook artifacts that Excel COM rejects before it can repair them.

        openpyxl can preserve legacy formulas like [1]Raw Material!A1 and
        formula-like error literals such as =#N/A. Excel may refuse to open
        those files through COM before the normal repair dialog is available.
        """
        removed_cells = []
        fixed_error_cells = []
        blank_formula_cells = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if (
                        isinstance(value, str)
                        and value.startswith("=")
                        and INVALID_EXTERNAL_FORMULA_RE.search(value)
                    ):
                        removed_cells.append(f"{sheet.title}!{cell.coordinate}")
                        cell.value = None
                    elif isinstance(value, str) and value.strip() == "=":
                        blank_formula_cells.append(f"{sheet.title}!{cell.coordinate}")
                        cell.value = None
                        self._add_validation_finding(
                            "WARNING",
                            "BLANK_FORMULA_REMOVED",
                            "已移除空白公式，避免公式鏈不可追溯。",
                            sheet=sheet.title,
                            row=cell.coordinate,
                            recommendation="確認模板延伸列是否應留空，或應填入完整公式。",
                        )
                    elif isinstance(value, str) and value in INVALID_ERROR_FORMULA_VALUES:
                        fixed_error_cells.append(f"{sheet.title}!{cell.coordinate}")
                        cell.value = INVALID_ERROR_FORMULA_VALUES[value]

        if removed_cells:
            sample = ", ".join(removed_cells[:8])
            if len(removed_cells) > 8:
                sample += ", ..."
            self._warn(
                f"已移除 {len(removed_cells)} 個 Excel 無效外部參照公式，避免結果檔開啟時修復失敗。"
                f" 位置: {sample}"
            )
        if fixed_error_cells:
            sample = ", ".join(fixed_error_cells[:8])
            if len(fixed_error_cells) > 8:
                sample += ", ..."
            self._warn(
                f"已修正 {len(fixed_error_cells)} 個被誤存為公式的 Excel 錯誤值。"
                f" 位置: {sample}"
            )
        if blank_formula_cells:
            sample = ", ".join(blank_formula_cells[:8])
            if len(blank_formula_cells) > 8:
                sample += ", ..."
            self._warn(
                f"已移除 {len(blank_formula_cells)} 個空白公式，避免公式鏈不可追溯。"
                f" 位置: {sample}"
            )

        rebuilt_sheet = self._rebuild_sheet_without_styles(workbook, "simapro10.2.0.0")
        if rebuilt_sheet:
            self._warn(
                "已重建 simapro10.2.0.0 工作表並保留儲存格內容，"
                "移除會造成 Excel COM 開啟失敗的樣式殘留。"
            )

        return len(removed_cells) + len(fixed_error_cells) + len(blank_formula_cells) + int(rebuilt_sheet)

    def _rebuild_sheet_without_styles(self, workbook, sheet_name: str) -> bool:
        if sheet_name not in workbook.sheetnames:
            return False

        source_sheet = workbook[sheet_name]
        sheet_index = workbook.worksheets.index(source_sheet)
        sheet_state = source_sheet.sheet_state
        temp_title = f"__{sheet_name[:20]}_clean"
        suffix = 1
        while temp_title in workbook.sheetnames:
            temp_title = f"__{sheet_name[:18]}_{suffix}_clean"
            suffix += 1

        clean_sheet = workbook.create_sheet(temp_title, sheet_index)
        clean_sheet.sheet_state = sheet_state
        for row in source_sheet.iter_rows(values_only=False):
            clean_sheet.append([cell.value for cell in row])

        workbook.remove(source_sheet)
        clean_sheet.title = sheet_name
        return True

    def set_cancel_event(self, cancel_event):
        self.cancel_event = cancel_event

    def clear_cancel(self):
        self.was_cancelled = False
        if self.cancel_event is not None:
            self.cancel_event.clear()

    def request_cancel(self):
        if self.cancel_event is not None:
            self.cancel_event.set()

    def _check_cancel(self):
        if self.cancel_event is not None and self.cancel_event.is_set():
            self.was_cancelled = True
            raise UserCancelledError("Operation cancelled by user.\n已取消作業，未寫入任何檔案。")

    def get_base_dir(self):
        """
        如果是 PyInstaller 打包後的 single-file exe，
        sys.argv[0] 會是使用者實際「雙擊執行」的那顆 .exe 的完整路徑。
        所以把它取 dirname 就能得到 exe 所在資料夾。

        如果是開發階段直接跑 .py，
        __file__ 會是目前 .py 的檔案路徑，我們就取 .py 同層資料夾即可。
        """
        import sys
        if getattr(sys, 'frozen', False):
            # 已經被打包成 exe
            return os.path.dirname(sys.argv[0])
        else:
            # 開發環境跑 .py
            return os.path.dirname(os.path.abspath(__file__))

    def browse_file(self):
        # 瀏覽文件並設置文件路徑
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        self.file_entry.delete(0, tk.END)
        self.file_entry.insert(0, self.file_path)

    def transform_sheet(self) -> TaskResult:
        run_id, started_at = self._start_task("transform_sheet")
        try:
            raw = self._transform_sheet_impl()
            result = self._coerce_task_result(
                task_name="transform_sheet",
                started_at=started_at,
                value=raw,
                success_message="Transform 完成",
                success_artifacts_fn=lambda: {"merged_file": getattr(self, "merged_file", "")},
            )
        except UserCancelledError as exc:
            self.was_cancelled = True
            result = self._result_fail(
                error_code="USER_CANCELLED",
                user_message=str(exc),
                started_at=started_at,
                exc=exc,
            )
        except Exception as exc:
            result = self._result_fail(
                error_code="TRANSFORM_EXCEPTION",
                user_message="Transform 發生未預期錯誤",
                started_at=started_at,
                exc=exc,
            )
        result.artifacts.setdefault("run_id", run_id)
        self._finish_task_log(run_id, "transform_sheet", result)
        return result

    def process_file(
        self,
        file_path=None,
        selected_stages=None,
        calculate_distances=True,
        carbon_boundary=None,
        use_transport_cache=True,
        force_recalculate_distances=False,
    ) -> TaskResult:
        run_id, started_at = self._start_task("process_file")
        try:
            raw = self._process_file_impl(
                file_path=file_path,
                selected_stages=selected_stages,
                calculate_distances=calculate_distances,
                carbon_boundary=carbon_boundary,
                use_transport_cache=use_transport_cache,
                force_recalculate_distances=force_recalculate_distances,
            )
            result = self._coerce_task_result(
                task_name="process_file",
                started_at=started_at,
                value=raw,
                success_message="資料處理完成",
                success_artifacts_fn=lambda: {
                    "result_file": getattr(self, "result_file", ""),
                    "report_file": getattr(self, "report_file", ""),
                },
            )
        except ValueError as exc:
            result = self._result_fail(
                error_code="INVALID_STAGE_SELECTION",
                user_message=str(exc),
                started_at=started_at,
                exc=exc,
            )
        except UserCancelledError as exc:
            self.was_cancelled = True
            result = self._result_fail(
                error_code="USER_CANCELLED",
                user_message=str(exc),
                started_at=started_at,
                exc=exc,
            )
        except Exception as exc:
            if self._is_permission_denied_error(exc):
                result = self._result_fail(
                    error_code=FILE_PERMISSION_DENIED_ERROR_CODE,
                    user_message=self._file_permission_denied_message(file_path),
                    started_at=started_at,
                    exc=exc,
                )
            else:
                result = self._result_fail(
                    error_code="PROCESS_EXCEPTION",
                    user_message="資料處理發生未預期錯誤",
                    started_at=started_at,
                    exc=exc,
                )
        result.artifacts.setdefault("run_id", run_id)
        self._finish_task_log(run_id, "process_file", result)
        return result

    def generate_report(self, template_choice, result_file=None) -> TaskResult:
        run_id, started_at = self._start_task("generate_report")
        try:
            raw = self._generate_report_impl(template_choice, result_file=result_file)
            result = self._coerce_task_result(
                task_name="generate_report",
                started_at=started_at,
                value=raw,
                success_message="報告產生完成",
                success_artifacts_fn=lambda: {
                    "report_doc": raw if isinstance(raw, str) else "",
                    "source_file": os.path.abspath(result_file) if result_file else getattr(self, "result_file", ""),
                },
            )
        except UserCancelledError as exc:
            self.was_cancelled = True
            result = self._result_fail(
                error_code="USER_CANCELLED",
                user_message=str(exc),
                started_at=started_at,
                exc=exc,
            )
        except Exception as exc:
            result = self._result_fail(
                error_code="REPORT_EXCEPTION",
                user_message="報告產生發生未預期錯誤",
                started_at=started_at,
                exc=exc,
            )
        self._cleanup_chart_artifacts()
        result.artifacts.setdefault("run_id", run_id)
        self._finish_task_log(run_id, "generate_report", result)
        return result

    def process_all(
        self,
        carbon_boundary=None,
        calculate_distances=True,
        use_transport_cache=True,
        force_recalculate_distances=False,
    ) -> TaskResult:
        run_id, started_at = self._start_task("process_all")
        if not self.file_path:
            result = TaskResult(
                ok=False,
                error_code="MISSING_INPUT",
                message="請選擇 Excel 文件",
                artifacts={"run_id": run_id},
                elapsed_ms=self._elapsed_ms(started_at),
                warnings=list(self._warnings),
            )
            self._finish_task_log(run_id, "process_all", result)
            return result
        transform_result = self.transform_sheet()
        if not transform_result.ok:
            transform_result.elapsed_ms = self._elapsed_ms(started_at)
            transform_result.artifacts["run_id"] = run_id
            self._finish_task_log(run_id, "process_all", transform_result)
            return transform_result

        merged_path = (
            transform_result.artifacts.get("path")
            or transform_result.artifacts.get("merged_file")
            or transform_result.message
        )
        process_result = self.process_file(
            file_path=merged_path if isinstance(merged_path, str) else None,
            carbon_boundary=carbon_boundary or self.carbon_boundary,
            calculate_distances=calculate_distances,
            use_transport_cache=use_transport_cache,
            force_recalculate_distances=force_recalculate_distances,
        )
        process_result.elapsed_ms = self._elapsed_ms(started_at)
        process_result.artifacts["run_id"] = run_id
        if process_result.ok:
            process_result.message = "Transform + Process 完成"
            process_result.warnings = list(self._warnings + process_result.warnings)
        self._finish_task_log(run_id, "process_all", process_result)
        return process_result

    @staticmethod
    def _normalize_column_name(value):
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value).strip().lower())

    def _find_table_column(self, columns, target_name):
        target = self._normalize_column_name(target_name)
        for idx, name in enumerate(columns, start=1):
            if self._normalize_column_name(name) == target:
                return idx, name
        return None, None

    @staticmethod
    def _coerce_numeric(value):
        if value is None or value == "":
            return None
        if isinstance(value, (int, float, np.number)):
            if pd.isna(value):
                return None
            return float(value)
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _to_excel_date_serial(value, epoch=WINDOWS_EPOCH):
        if value is None or value == "":
            return ""
        if pd.isna(value):
            return ""
        if isinstance(value, datetime):
            return int(to_excel(value.date(), epoch=epoch))
        if isinstance(value, date):
            return int(to_excel(value, epoch=epoch))

        text = str(value).strip()
        if not text:
            return ""
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return text
        return int(to_excel(parsed.date(), epoch=epoch))

    @staticmethod
    def _normalize_transport_lookup_key(value):
        if value is None:
            return ""
        text = unicodedata.normalize("NFKC", str(value).strip())
        text = re.sub(r"\s+", " ", text)
        return text.casefold()

    def _transport_lookup_key_candidates(self, value):
        text = unicodedata.normalize("NFKC", str(value or "").strip())
        if not text:
            return []

        candidates = [text]
        match = re.match(r"^([A-Za-z]{2,5}[A-Za-z0-9]{0,2})(?=\s*[-－–—/])", text)
        if match:
            candidates.append(match.group(1))

        normalized = []
        seen = set()
        for candidate in candidates:
            key = self._normalize_transport_lookup_key(candidate)
            if key and key not in seen:
                seen.add(key)
                normalized.append(key)
        return normalized

    def _get_cached_endpoint_place(self, endpoint_text, use_persistent_cache=True):
        key = self._normalize_transport_lookup_key(endpoint_text)
        if not key:
            return
        if not use_persistent_cache:
            return
        cached = self._transport_endpoint_place_cache.get(key)
        if cached is not None:
            return dict(cached)
        cache = self._load_transport_distance_cache()
        clean_place = self._sanitize_cached_place(cache.get("places", {}).get(key))
        if clean_place:
            self._transport_endpoint_place_cache[key] = clean_place
            return dict(clean_place)
        return None

    def _cache_endpoint_place(self, endpoint_text, place, persist=True, save_immediately=True):
        key = self._normalize_transport_lookup_key(endpoint_text)
        if not key:
            return
        clean_place = self._sanitize_cached_place(place)
        if not clean_place:
            return
        if not clean_place["query"]:
            clean_place["query"] = str(endpoint_text or "").strip()
        self._transport_endpoint_place_cache[key] = clean_place
        if persist:
            cache = self._load_transport_distance_cache()
            cache.setdefault("places", {})[key] = clean_place
            self._transport_distance_cache_dirty = True
            if save_immediately:
                self._save_transport_distance_cache()

    def _is_road_transport_type(self, transport_type):
        normalized = self._normalize_column_name(transport_type)
        normalized = re.sub(r"[\s\-_/,;:]+", " ", normalized).strip()
        return normalized in ROAD_TRANSPORT_TYPES

    def _transport_database_alignment_warning(self, transport_type, database_name):
        transport_text = self._normalize_column_name(transport_type)
        database_text = self._normalize_column_name(database_name)
        if not transport_text or not database_text:
            return ""

        if "air" in transport_text or "aircraft" in transport_text:
            if "aircraft" not in database_text and "air" not in database_text:
                return "Air transport row does not appear to use an aircraft/air database."
        elif "sea" in transport_text or "ship" in transport_text or "ocean" in transport_text:
            sea_terms = ("sea", "ship", "bulk carrier", "ocean", "vessel")
            if not any(term in database_text for term in sea_terms):
                return "Sea transport row does not appear to use a sea/shipping database."
        elif self._is_road_transport_type(transport_type):
            blocked_terms = ("aircraft", "sea", "bulk carrier", "ship", "ocean", "vessel")
            if any(term in database_text for term in blocked_terms):
                return "Road transport row appears to use an air/sea database."
            road_terms = ("lorry", "truck", "road", "freight")
            if not any(term in database_text for term in road_terms):
                return "Road transport row database does not clearly identify road freight."
        return ""

    def _resource_candidates(self, filename, *legacy_paths):
        candidates = [os.path.join(self.resources_dir, filename), *legacy_paths]
        unique_candidates = []
        seen = set()
        for path in candidates:
            normalized = os.path.normcase(os.path.normpath(path))
            if normalized in seen:
                continue
            seen.add(normalized)
            unique_candidates.append(path)
        return unique_candidates

    def _find_existing_resource(self, filename, *legacy_paths):
        return next(
            (path for path in self._resource_candidates(filename, *legacy_paths) if os.path.exists(path)),
            None,
        )

    def _get_required_resource_path(self, filename, label, *legacy_paths):
        resource_path = self._find_existing_resource(filename, *legacy_paths)
        if resource_path:
            return resource_path

        searched_paths = "\n".join(self._resource_candidates(filename, *legacy_paths))
        raise FileNotFoundError(f"找不到{label}：{filename}\n已檢查路徑：\n{searched_paths}")

    def _transport_mapping_candidates(self):
        return self._resource_candidates(
            TRANSPORT_LOCATION_MAPPING_FILENAME,
            os.path.join(self.spreadsheet_dir, TRANSPORT_LOCATION_MAPPING_FILENAME),
            os.path.join(self.base_dir, "output", "spreadsheet", TRANSPORT_LOCATION_MAPPING_FILENAME),
            os.path.join(self.base_dir, TRANSPORT_LOCATION_MAPPING_FILENAME),
        )

    def _load_transport_location_mapping(self):
        if self._transport_location_mapping is not None:
            return self._transport_location_mapping

        mapping_path = next(
            (path for path in self._transport_mapping_candidates() if os.path.exists(path)),
            None,
        )
        if not mapping_path:
            if not self._transport_mapping_missing_warned:
                self._warn(
                    f"找不到運輸端點對照表 {TRANSPORT_LOCATION_MAPPING_FILENAME}，運輸端點將直接使用原始文字查詢。"
                )
                self._transport_mapping_missing_warned = True
            self._transport_location_mapping = {}
            return self._transport_location_mapping

        mapping = {}
        copied_mapping_path = None
        try:
            try:
                workbook = openpyxl.load_workbook(mapping_path, read_only=True, data_only=True)
            except PermissionError:
                copied_mapping_path = os.path.join(
                    self.tmp_dir,
                    f"{os.path.splitext(TRANSPORT_LOCATION_MAPPING_FILENAME)[0]}_readcopy_{os.getpid()}.xlsx",
                )
                shutil.copy2(mapping_path, copied_mapping_path)
                workbook = openpyxl.load_workbook(copied_mapping_path, read_only=True, data_only=True)

            def add_lookup(raw_key, record, *, overwrite=False):
                key = self._normalize_transport_lookup_key(raw_key)
                if not key:
                    return
                if overwrite or key not in mapping:
                    mapping[key] = record

            def load_mapping_sheet(worksheet, *, overwrite):
                rows = worksheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if not headers:
                    return

                header_index = {
                    self._normalize_transport_lookup_key(name): idx
                    for idx, name in enumerate(headers)
                    if name is not None
                }

                def col(name):
                    return header_index.get(self._normalize_transport_lookup_key(name))

                lookup_idx = col("lookup_key")
                road_location_idx = col("road_location_for_geocode")
                aliases_idx = col("aliases")
                if lookup_idx is None or road_location_idx is None:
                    return

                for row in rows:
                    if not row:
                        continue
                    lookup_key = row[lookup_idx] if lookup_idx < len(row) else None
                    road_location = row[road_location_idx] if road_location_idx < len(row) else None
                    if not lookup_key or not road_location:
                        continue
                    record = {
                        str(headers[idx]): row[idx] if idx < len(row) else None
                        for idx in range(len(headers))
                        if headers[idx] is not None
                    }
                    record["mapping_path"] = mapping_path
                    record["mapping_sheet"] = worksheet.title
                    add_lookup(lookup_key, record, overwrite=overwrite)

                    aliases = row[aliases_idx] if aliases_idx is not None and aliases_idx < len(row) else None
                    for alias in str(aliases or "").split(";"):
                        add_lookup(alias.strip(), record)

            if "Source" in workbook.sheetnames:
                load_mapping_sheet(workbook["Source"], overwrite=False)
            if "mapping" in workbook.sheetnames:
                load_mapping_sheet(workbook["mapping"], overwrite=True)
            if not mapping and workbook.sheetnames:
                load_mapping_sheet(workbook.active, overwrite=True)

            workbook.close()
            if copied_mapping_path:
                with suppress(Exception):
                    os.remove(copied_mapping_path)
        except Exception as exc:
            self._warn(
                f"讀取運輸端點對照表失敗，運輸端點將直接使用原始文字查詢: {mapping_path} | {exc}"
            )
            mapping = {}
            if copied_mapping_path:
                with suppress(Exception):
                    os.remove(copied_mapping_path)

        self._transport_location_mapping = mapping
        return self._transport_location_mapping

    @staticmethod
    def normalize_factory_site(factory_site):
        text = str(factory_site or "").strip()
        return FACTORY_SITE_ALIASES.get(text, text)

    def _apply_factory_overview_info(self, workbook):
        if "overview" not in workbook.sheetnames:
            return

        factory_site = self.normalize_factory_site(self.factory_site)
        overview = workbook["overview"]
        if not factory_site:
            overview["C3"] = ""
            overview["C4"] = ""
            return

        factory_info = FACTORY_OVERVIEW_INFO.get(factory_site)
        if factory_info:
            overview["C3"] = factory_info["name"]
            overview["C4"] = factory_info["address"]

    def _disable_refresh_on_file_open(self, workbook):
        connection_attr_by_type = {
            XL_CONNECTION_TYPE_OLEDB: "OLEDBConnection",
            XL_CONNECTION_TYPE_ODBC: "ODBCConnection",
        }
        try:
            connections = workbook.Connections
        except Exception as exc:
            self._warn(f"Workbook connection inspection skipped: {exc}")
            return

        for conn in connections:
            conn_name = ""
            with suppress(Exception):
                conn_name = str(conn.Name)
            try:
                conn_type = int(conn.Type)
            except Exception as exc:
                self._warn(f"Connection refresh-on-open skipped: {conn_name or '<unknown>'}: {exc}")
                continue
            connection_attr = connection_attr_by_type.get(conn_type)
            if not connection_attr:
                continue
            try:
                connection_obj = getattr(conn, connection_attr)
                connection_obj.RefreshOnFileOpen = False
            except Exception as exc:
                self._warn(f"Connection refresh-on-open skipped: {conn_name or '<unknown>'}: {exc}")

    @staticmethod
    def _format_report_number(value) -> str:
        """Format generated report numbers with exactly four decimals."""
        if value is None or value == "":
            return ""
        try:
            if pd.isna(value):
                return ""
        except (TypeError, ValueError):
            pass
        try:
            return f"{float(value):.4f}"
        except (TypeError, ValueError):
            return str(value)

    @staticmethod
    def _new_transport_distance_cache():
        return {"version": 1, "places": {}, "routes": {}, "route_errors": {}}

    @staticmethod
    def _json_safe(value):
        try:
            return json.loads(json.dumps(value, ensure_ascii=False, default=str))
        except Exception:
            return {}

    def _sanitize_cached_place(self, place):
        if not isinstance(place, dict):
            return None
        try:
            lat = float(place.get("lat"))
            lon = float(place.get("lon"))
        except (TypeError, ValueError):
            return None
        if not (math.isfinite(lat) and math.isfinite(lon)):
            return None
        if lat < -90 or lat > 90 or lon < -180 or lon > 180:
            return None
        return {
            "query": str(place.get("query") or "").strip(),
            "lat": lat,
            "lon": lon,
            "label": str(place.get("label") or place.get("query") or "").strip(),
            "provider": str(place.get("provider") or "cache").strip(),
        }

    def _load_transport_distance_cache(self):
        if self._transport_distance_cache is not None:
            return self._transport_distance_cache

        cache = self._new_transport_distance_cache()
        path = self.transport_distance_cache_path
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as fh:
                    loaded = json.load(fh)
                if isinstance(loaded, dict):
                    for section in ("places", "routes", "route_errors"):
                        if isinstance(loaded.get(section), dict):
                            cache[section] = loaded[section]
            except Exception as exc:
                self._warn(f"讀取運輸距離快取失敗，將使用空白快取: {path} | {exc}")

        self._transport_distance_cache = cache
        places = cache.get("places", {})
        if isinstance(places, dict):
            for key, place in places.items():
                clean_place = self._sanitize_cached_place(place)
                if clean_place:
                    self._transport_endpoint_place_cache[str(key)] = clean_place
        return self._transport_distance_cache

    def _save_transport_distance_cache(self):
        cache = self._transport_distance_cache
        if not isinstance(cache, dict):
            return
        try:
            os.makedirs(os.path.dirname(self.transport_distance_cache_path), exist_ok=True)
            tmp_path = f"{self.transport_distance_cache_path}.tmp"
            with open(tmp_path, "w", encoding="utf-8") as fh:
                json.dump(cache, fh, ensure_ascii=False, indent=2)
            os.replace(tmp_path, self.transport_distance_cache_path)
            self._transport_distance_cache_dirty = False
        except Exception as exc:
            if not self._transport_distance_cache_save_warned:
                self._warn(f"寫入運輸距離快取失敗，這次結果仍會繼續處理: {exc}")
                self._transport_distance_cache_save_warned = True

    def _route_cache_key_to_string(self, route_cache_key):
        parts = [str(part or "") for part in route_cache_key]
        return json.dumps(parts, ensure_ascii=False, separators=(",", ":"))

    def _route_result_to_cache_payload(self, result):
        return {
            "mode": str(result.mode or "cached"),
            "code": str(result.code or "Cached"),
            "distance_m": float(result.distance_m),
            "duration_s": float(result.duration_s or 0.0),
            "metadata": self._json_safe(result.metadata if isinstance(result.metadata, dict) else {}),
            "cached_at": time.time(),
        }

    def _cached_payload_to_route_result(self, payload):
        if not isinstance(payload, dict):
            return None
        try:
            distance_m = float(payload.get("distance_m", payload.get("distance")))
            duration_s = float(payload.get("duration_s", payload.get("duration") or 0.0))
        except (TypeError, ValueError):
            return None
        if not math.isfinite(distance_m) or distance_m < 0:
            return None
        metadata = payload.get("metadata")
        if not isinstance(metadata, dict):
            metadata = {}
        return RouteResult(
            mode=str(payload.get("mode") or "cached"),
            code=str(payload.get("code") or "Cached"),
            distance_m=distance_m,
            duration_s=duration_s,
            geometry=[],
            segments=[],
            metadata=dict(metadata),
        )

    def _get_persistent_route_result(self, route_cache_key):
        cache = self._load_transport_distance_cache()
        routes = cache.setdefault("routes", {})
        cache_key = self._route_cache_key_to_string(route_cache_key)
        result = self._cached_payload_to_route_result(routes.get(cache_key))
        if result is None and cache_key in routes:
            routes.pop(cache_key, None)
            self._transport_distance_cache_dirty = True
        return result

    def _store_persistent_route_result(self, route_cache_key, result):
        if result is None:
            return
        cache = self._load_transport_distance_cache()
        cache_key = self._route_cache_key_to_string(route_cache_key)
        cache.setdefault("routes", {})[cache_key] = self._route_result_to_cache_payload(result)
        cache.setdefault("route_errors", {}).pop(cache_key, None)
        self._transport_distance_cache_dirty = True
        self._save_transport_distance_cache()

    def _get_recent_persistent_route_error(self, route_cache_key):
        cache = self._load_transport_distance_cache()
        route_errors = cache.setdefault("route_errors", {})
        cache_key = self._route_cache_key_to_string(route_cache_key)
        payload = route_errors.get(cache_key)
        if not isinstance(payload, dict):
            return None
        try:
            cached_at = float(payload.get("cached_at") or 0.0)
        except (TypeError, ValueError):
            cached_at = 0.0
        if time.time() - cached_at <= TRANSPORT_NEGATIVE_CACHE_TTL_SEC:
            return str(payload.get("message") or "cached route lookup failure")
        route_errors.pop(cache_key, None)
        self._transport_distance_cache_dirty = True
        self._save_transport_distance_cache()
        return None

    def _store_persistent_route_error(self, route_cache_key, exc):
        cache = self._load_transport_distance_cache()
        cache_key = self._route_cache_key_to_string(route_cache_key)
        cache.setdefault("route_errors", {})[cache_key] = {
            "message": str(exc),
            "cached_at": time.time(),
            "ttl_sec": TRANSPORT_NEGATIVE_CACHE_TTL_SEC,
        }
        self._transport_distance_cache_dirty = True
        self._save_transport_distance_cache()

    def _resolve_road_transport_endpoint(self, transport_type, endpoint):
        return self._resolve_road_transport_endpoint_details(transport_type, endpoint)["query"]

    @staticmethod
    def _record_lat_lon(record):
        if not record:
            return None
        try:
            lat = float(record.get("latitude"))
            lon = float(record.get("longitude"))
        except (TypeError, ValueError):
            return None
        if not (math.isfinite(lat) and math.isfinite(lon)):
            return None
        if lat < -90 or lat > 90 or lon < -180 or lon > 180:
            return None
        return lat, lon

    def _resolve_road_transport_endpoint_details(
        self,
        transport_type,
        endpoint,
        prefer_local_mapping=True,
        use_endpoint_cache=True,
    ):
        endpoint_text = str(endpoint or "").strip()
        if not endpoint_text:
            return {"query": endpoint_text, "record": None, "lat_lon": None, "place": None}

        if prefer_local_mapping:
            mapping = self._load_transport_location_mapping()
            record = next(
                (
                    mapping.get(candidate_key)
                    for candidate_key in self._transport_lookup_key_candidates(endpoint_text)
                    if mapping.get(candidate_key)
                ),
                None,
            )
            if record:
                return {
                    "query": str(record.get("road_location_for_geocode") or endpoint_text).strip(),
                    "record": record,
                    "lat_lon": self._record_lat_lon(record),
                    "place": None,
                }

        cached_place = self._get_cached_endpoint_place(
            endpoint_text,
            use_persistent_cache=use_endpoint_cache,
        )
        if cached_place is not None:
            return {
                "query": endpoint_text,
                "record": None,
                "lat_lon": (float(cached_place["lat"]), float(cached_place["lon"])),
                "place": dict(cached_place),
            }

        return {"query": endpoint_text, "record": None, "lat_lon": None, "place": None}

    def _transport_place_from_endpoint_details(self, endpoint_details, transport_type, use_cache=True):
        cached_place = endpoint_details.get("place")
        if isinstance(cached_place, dict):
            return dict(cached_place)
        lat_lon = endpoint_details.get("lat_lon")
        query = endpoint_details.get("query") or ""
        if lat_lon is not None:
            lat, lon = lat_lon
            record = endpoint_details.get("record") or {}
            return {
                "query": query,
                "lat": lat,
                "lon": lon,
                "label": str(record.get("name_en") or record.get("lookup_key") or query),
                "provider": str(record.get("mapping_sheet") or "mapping"),
            }
        place = geocode_place(
            query,
            transport_type=transport_type,
            timeout_sec=self.runtime_config["transport_geocode_timeout_sec"],
            cache=self._transport_place_cache if use_cache else None,
        )
        self._cache_endpoint_place(query, place, persist=use_cache)
        return place

    def calculate_transport_distances(
        self,
        workbook,
        transport_tables,
        sheet_name="Distribution",
        table_indexes=None,
        use_transport_cache=True,
        force_recalculate_distances=False,
    ):
        if sheet_name not in workbook.sheetnames:
            self._warn(f"找不到 {sheet_name} 工作表，略過距離計算。")
            return transport_tables

        sheet = workbook[sheet_name]
        is_raw_material = self._normalize_column_name(sheet_name) == "raw material"
        is_distribution = self._normalize_column_name(sheet_name) == "distribution"
        pending_jobs = []
        route_jobs = {}
        route_keys = []
        updated_count = 0
        skipped_existing_count = 0
        stats = {
            "cache_hits": 0,
            "external_queries": 0,
            "failed": 0,
            "local_mapping_routes": 0,
            "local_calculations": 0,
            "negative_cache_hits": 0,
        }
        table_index_set = set(table_indexes) if table_indexes is not None else None

        def column_name_at(columns, column_idx):
            if column_idx is None or column_idx < 1 or column_idx > len(columns):
                return None
            return columns[column_idx - 1]

        def raw_numeric_factor(row, column_name):
            if column_name is None:
                return None
            value = row.get(column_name)
            if value is None or value == "" or pd.isna(value):
                return 0.0
            return self._coerce_numeric(value)

        def calculate_ton_km_factor(
            row,
            weight_col_name,
            raw_material_col_names=None,
            allocated_proportion_col_name=None,
        ):
            if is_raw_material:
                if raw_material_col_names is None:
                    return None
                manufacturer_alloc = raw_numeric_factor(row, raw_material_col_names["manufacturer_alloc"])
                part_source_alloc = raw_numeric_factor(row, raw_material_col_names["part_source_alloc"])
                weight_parts = raw_numeric_factor(row, raw_material_col_names["weight_parts"])
                bom_quantity = raw_numeric_factor(row, raw_material_col_names["bom_quantity"])
                if None in (manufacturer_alloc, part_source_alloc, weight_parts, bom_quantity):
                    return None
                return weight_parts * bom_quantity * part_source_alloc * manufacturer_alloc / 1000.0

            if weight_col_name is None:
                return None
            weight_kg = self._coerce_numeric(row.get(weight_col_name))
            if weight_kg is None:
                return None
            if is_distribution:
                allocated_proportion = raw_numeric_factor(row, allocated_proportion_col_name)
                if allocated_proportion is None:
                    return None
                return weight_kg / 1000.0 * allocated_proportion
            return weight_kg / 1000.0

        for table_idx, (start_idx, sheet_data) in enumerate(transport_tables, start=1):
            if table_index_set is not None and table_idx not in table_index_set:
                continue

            distance_col_idx, distance_col_name = self._find_table_column(sheet_data.columns, "distance transported (km)")
            _, start_col_name = self._find_table_column(sheet_data.columns, "starting point")
            _, end_col_name = self._find_table_column(sheet_data.columns, "end point")
            _, transport_col_name = self._find_table_column(sheet_data.columns, "type of transport")
            _, database_col_name = self._find_table_column(sheet_data.columns, "name of database")
            ton_km_col_idx, ton_km_col_name = self._find_table_column(sheet_data.columns, "Ton‧Km")
            _, weight_col_name = self._find_table_column(sheet_data.columns, "Weight (product+package)（Kg）")
            _, allocated_proportion_col_name = self._find_table_column(sheet_data.columns, "allocated proportion")
            if distance_col_name is not None:
                sheet_data[distance_col_name] = sheet_data[distance_col_name].astype(object)
            if ton_km_col_name is not None:
                sheet_data[ton_km_col_name] = sheet_data[ton_km_col_name].astype(object)
            raw_material_col_names = None
            if is_raw_material:
                _, manufacturer_alloc_col_name = self._find_table_column(
                    sheet_data.columns,
                    "allocated proportion (Manufacturer)",
                )
                _, part_source_alloc_col_name = self._find_table_column(
                    sheet_data.columns,
                    "allocated proportion (Part source)",
                )
                _, weight_parts_col_name = self._find_table_column(sheet_data.columns, "Weight of Parts")
                _, bom_quantity_col_name = self._find_table_column(sheet_data.columns, "BOM Quantity")
                next_column_idx = distance_col_idx if distance_col_idx is not None else 0
                raw_material_col_names = {
                    "manufacturer_alloc": manufacturer_alloc_col_name
                    or column_name_at(sheet_data.columns, next_column_idx + 1),
                    "part_source_alloc": part_source_alloc_col_name
                    or column_name_at(sheet_data.columns, next_column_idx + 2),
                    "weight_parts": weight_parts_col_name
                    or column_name_at(sheet_data.columns, next_column_idx + 3),
                    "bom_quantity": bom_quantity_col_name
                    or column_name_at(sheet_data.columns, next_column_idx + 4),
                }

            required = {
                "starting point": start_col_name,
                "end point": end_col_name,
                "type of transport": transport_col_name,
                "distance transported (km)": distance_col_name,
            }
            missing = [name for name, actual in required.items() if actual is None]
            if missing:
                raise ValueError(f"【{sheet_name} 第 {table_idx} 個表格】缺少必要欄位: {missing}")

            for row_offset, row in sheet_data.iterrows():
                self._check_cancel()
                excel_row = start_idx + 4 + row_offset
                transport_type_raw = row.get(transport_col_name)
                database_name_raw = row.get(database_col_name) if database_col_name else ""
                if not pd.isna(transport_type_raw):
                    warning = self._transport_database_alignment_warning(
                        transport_type_raw,
                        database_name_raw,
                    )
                    if warning:
                        self._add_validation_finding(
                            "WARNING",
                            "TRANSPORT_DATABASE_TYPE_MISMATCH",
                            warning,
                            stage=sheet_name,
                            sheet=sheet_name,
                            row=str(excel_row),
                            recommendation="確認 type of transport 與 name of database 是否對應；若為多段運輸，請拆分或修正係數。",
                        )
                existing_distance_km = self._coerce_numeric(row.get(distance_col_name))
                if (
                    not force_recalculate_distances
                    and existing_distance_km is not None
                    and not math.isclose(existing_distance_km, 0.0, abs_tol=1e-12)
                ):
                    ton_km_factor = calculate_ton_km_factor(
                        row,
                        weight_col_name,
                        raw_material_col_names,
                        allocated_proportion_col_name,
                    )
                    if ton_km_col_name is not None and ton_km_factor is not None:
                        ton_km_value = ton_km_factor * existing_distance_km
                        sheet_data.at[row_offset, ton_km_col_name] = ton_km_value
                        if ton_km_col_idx is not None:
                            sheet.cell(row=excel_row, column=ton_km_col_idx).value = ton_km_value
                    skipped_existing_count += 1
                    continue

                start_point_raw = row.get(start_col_name)
                end_point_raw = row.get(end_col_name)
                if pd.isna(transport_type_raw) or pd.isna(start_point_raw) or pd.isna(end_point_raw):
                    continue
                transport_type = str(transport_type_raw or "").strip()
                start_point = str(start_point_raw or "").strip()
                end_point = str(end_point_raw or "").strip()
                if not transport_type or not start_point or not end_point:
                    continue

                resolved_start = self._resolve_road_transport_endpoint_details(
                    transport_type,
                    start_point,
                    prefer_local_mapping=use_transport_cache,
                    use_endpoint_cache=use_transport_cache,
                )
                resolved_end = self._resolve_road_transport_endpoint_details(
                    transport_type,
                    end_point,
                    prefer_local_mapping=use_transport_cache,
                    use_endpoint_cache=use_transport_cache,
                )
                resolved_start_point = resolved_start["query"]
                resolved_end_point = resolved_end["query"]
                route_cache_key = (
                    self._normalize_column_name(transport_type),
                    self._normalize_transport_lookup_key(resolved_start_point),
                    self._normalize_transport_lookup_key(resolved_end_point),
                )
                job = {
                    "table_idx": table_idx,
                    "sheet_data": sheet_data,
                    "row_offset": row_offset,
                    "excel_row": excel_row,
                    "distance_col_idx": distance_col_idx,
                    "distance_col_name": distance_col_name,
                    "ton_km_col_idx": ton_km_col_idx,
                    "ton_km_col_name": ton_km_col_name,
                    "transport_type": transport_type,
                    "start_point": start_point,
                    "end_point": end_point,
                    "resolved_start": resolved_start,
                    "resolved_end": resolved_end,
                    "resolved_start_point": resolved_start_point,
                    "resolved_end_point": resolved_end_point,
                    "ton_km_factor": calculate_ton_km_factor(
                        row,
                        weight_col_name,
                        raw_material_col_names,
                        allocated_proportion_col_name,
                    ),
                }
                pending_jobs.append(job)
                if route_cache_key not in route_jobs:
                    route_jobs[route_cache_key] = []
                    route_keys.append(route_cache_key)
                route_jobs[route_cache_key].append(job)

        if pending_jobs:
            existing_message = (
                "重新計算所有距離"
                if force_recalculate_distances
                else f"沿用既有距離 {skipped_existing_count} 筆"
            )
            self._notify_status(
                f"{sheet_name} 運輸距離待計算 {len(pending_jobs)} 筆，"
                f"去重後 {len(route_keys)} 條路線，{existing_message}。"
            )

        def notify_route_status(completed_count, route_message=""):
            message = (
                f"計算 {sheet_name} 運輸距離... {completed_count}/{len(route_keys)} | "
                f"快取 {stats['cache_hits']} | 外部查詢 {stats['external_queries']} | "
                f"失敗 {stats['failed']}"
            )
            if stats["local_mapping_routes"]:
                message += f" | 本地對照 {stats['local_mapping_routes']}"
            if route_message:
                message += f" | {route_message}"
            self._notify_status(message)

        route_results = {}
        route_errors = {}
        for route_idx, route_cache_key in enumerate(route_keys, start=1):
            self._check_cancel()
            representative_job = route_jobs[route_cache_key][0]
            transport_type = representative_job["transport_type"]
            start_point = representative_job["start_point"]
            end_point = representative_job["end_point"]
            resolved_start = representative_job["resolved_start"]
            resolved_end = representative_job["resolved_end"]
            resolved_start_point = representative_job["resolved_start_point"]
            resolved_end_point = representative_job["resolved_end_point"]
            route_message = f"{transport_type}: {start_point} -> {end_point}"
            if resolved_start.get("record") is not None or resolved_end.get("record") is not None:
                stats["local_mapping_routes"] += 1
            notify_route_status(route_idx - 1, route_message)
            try:
                result = None
                if use_transport_cache and not force_recalculate_distances:
                    result = self._transport_route_cache.get(route_cache_key)
                    if result is not None:
                        stats["cache_hits"] += 1
                    if result is None:
                        result = self._get_persistent_route_result(route_cache_key)
                        if result is not None:
                            self._transport_route_cache[route_cache_key] = result
                            stats["cache_hits"] += 1
                    if result is None:
                        cached_error = self._get_recent_persistent_route_error(route_cache_key)
                        if cached_error:
                            stats["cache_hits"] += 1
                            stats["negative_cache_hits"] += 1
                            stats["failed"] += 1
                            route_errors[route_cache_key] = RuntimeError(
                                f"最近失敗路線快取仍有效，略過外部查詢: {cached_error}"
                            )
                            notify_route_status(route_idx, route_message)
                            continue

                if result is None:
                    mode = transport_type_to_mode(transport_type)
                    network_required = (
                        mode in {"driving", "driving-traffic", "walking", "cycling", "osrm"}
                        or resolved_start.get("lat_lon") is None
                        or resolved_end.get("lat_lon") is None
                    )
                    if network_required:
                        stats["external_queries"] += 1
                    else:
                        stats["local_calculations"] += 1

                    if resolved_start.get("lat_lon") is not None or resolved_end.get("lat_lon") is not None:
                        from_place = self._transport_place_from_endpoint_details(
                            resolved_start,
                            transport_type,
                            use_cache=use_transport_cache,
                        )
                        to_place = self._transport_place_from_endpoint_details(
                            resolved_end,
                            transport_type,
                            use_cache=use_transport_cache,
                        )
                        result = compute_transport_distance(
                            mode=mode,
                            from_lat=float(from_place["lat"]),
                            from_lon=float(from_place["lon"]),
                            to_lat=float(to_place["lat"]),
                            to_lon=float(to_place["lon"]),
                            timeout_sec=self.runtime_config["transport_route_timeout_sec"],
                        )
                        result.metadata.update(
                            {
                                "transport_type": transport_type,
                                "from_query": resolved_start_point,
                                "to_query": resolved_end_point,
                                "from_place": from_place,
                                "to_place": to_place,
                            }
                        )
                    else:
                        result = compute_transport_distance_from_queries(
                            transport_type=transport_type,
                            from_query=resolved_start_point,
                            to_query=resolved_end_point,
                            timeout_sec=self.runtime_config["transport_route_timeout_sec"],
                            geocode_timeout_sec=self.runtime_config["transport_geocode_timeout_sec"],
                            geocode_cache=self._transport_place_cache if use_transport_cache else None,
                        )
                    result.metadata.update(
                        {
                            "original_from_query": start_point,
                            "original_to_query": end_point,
                            "resolved_from_query": resolved_start_point,
                            "resolved_to_query": resolved_end_point,
                        }
                    )
                    if use_transport_cache:
                        self._transport_route_cache[route_cache_key] = result
                        self._store_persistent_route_result(route_cache_key, result)
                route_results[route_cache_key] = result
                notify_route_status(route_idx, route_message)
            except Exception as exc:
                stats["failed"] += 1
                route_errors[route_cache_key] = exc
                if use_transport_cache:
                    self._store_persistent_route_error(route_cache_key, exc)
                notify_route_status(route_idx, route_message)

        for route_cache_key in route_keys:
            self._check_cancel()
            result = route_results.get(route_cache_key)
            if result is None:
                exc = route_errors.get(route_cache_key)
                for job in route_jobs[route_cache_key]:
                    self._warn(
                        f"{sheet_name} 第 {job['table_idx']} 個表格第 {job['excel_row']} 列距離計算失敗: "
                        f"{job['transport_type']} | {job['start_point']} -> {job['end_point']} "
                        f"(resolved: {job['resolved_start_point']} -> {job['resolved_end_point']}) | {exc}"
                    )
                continue

            first_job = route_jobs[route_cache_key][0]
            self._cache_endpoint_place(
                first_job["start_point"],
                result.metadata.get("from_place"),
                persist=use_transport_cache,
                save_immediately=False,
            )
            self._cache_endpoint_place(
                first_job["resolved_start_point"],
                result.metadata.get("from_place"),
                persist=use_transport_cache,
                save_immediately=False,
            )
            self._cache_endpoint_place(
                first_job["end_point"],
                result.metadata.get("to_place"),
                persist=use_transport_cache,
                save_immediately=False,
            )
            self._cache_endpoint_place(
                first_job["resolved_end_point"],
                result.metadata.get("to_place"),
                persist=use_transport_cache,
                save_immediately=False,
            )

            distance_km = round(result.distance_m / 1000.0, 4)
            for job in route_jobs[route_cache_key]:
                sheet_data = job["sheet_data"]
                row_offset = job["row_offset"]
                sheet_data.at[row_offset, job["distance_col_name"]] = distance_km
                sheet.cell(row=job["excel_row"], column=job["distance_col_idx"]).value = distance_km

                ton_km_col_name = job["ton_km_col_name"]
                ton_km_col_idx = job["ton_km_col_idx"]
                ton_km_factor = job["ton_km_factor"]
                if ton_km_col_name is not None and ton_km_factor is not None:
                    ton_km_value = ton_km_factor * distance_km
                    sheet_data.at[row_offset, ton_km_col_name] = ton_km_value
                    if ton_km_col_idx is not None:
                        sheet.cell(row=job["excel_row"], column=ton_km_col_idx).value = ton_km_value

                updated_count += 1

        if use_transport_cache and self._transport_distance_cache_dirty:
            self._save_transport_distance_cache()

        self._notify_status(
            f"{sheet_name} 運輸距離更新完成，共更新 {updated_count} 筆，"
            f"沿用既有距離 {skipped_existing_count} 筆，快取命中 {stats['cache_hits']} 筆，"
            f"外部查詢 {stats['external_queries']} 筆，失敗 {stats['failed']} 筆。"
        )
        return transport_tables

    @staticmethod
    def normalize_carbon_boundary(carbon_boundary=None):
        value = str(carbon_boundary or DEFAULT_CARBON_BOUNDARY).strip()
        if value in CARBON_BOUNDARY_STAGE_MAP:
            return value
        if value in CARBON_BOUNDARY_KEYS_BY_LABEL:
            return CARBON_BOUNDARY_KEYS_BY_LABEL[value]
        raise ValueError(
            "未知的產品碳足跡邊界模式，請選擇「搖籃到大門」或「搖籃到墳墓」。"
        )

    @staticmethod
    def stages_for_carbon_boundary(carbon_boundary=None):
        boundary = ExcelApp.normalize_carbon_boundary(carbon_boundary)
        return list(CARBON_BOUNDARY_STAGE_MAP[boundary])

    def _normalize_selected_carbon_stages(self, selected_stages=None, carbon_boundary=None):
        valid_stages = [stage for stage, _ in CARBON_STAGE_OPTIONS]
        if carbon_boundary is not None:
            return self.stages_for_carbon_boundary(carbon_boundary)
        if selected_stages is None:
            return self.stages_for_carbon_boundary(self.carbon_boundary)

        selected = []
        for stage in selected_stages:
            stage_name = str(stage or "").strip()
            if stage_name in valid_stages and stage_name not in selected:
                selected.append(stage_name)

        if not selected:
            raise ValueError("請至少勾選一個要計算碳排的階段。")
        return selected

    def _reset_process_review_artifacts(self, selected_stages, carbon_boundary):
        self._calculation_audit_rows = []
        self._validation_findings = []
        boundary = self.normalize_carbon_boundary(carbon_boundary)
        label = CARBON_BOUNDARY_LABELS[boundary]
        self._add_validation_finding(
            "INFO",
            "CARBON_BOUNDARY_SELECTED",
            f"本次產品碳足跡邊界為 {label}。",
            recommendation="匯出資料須以此邊界判定納入階段與排除階段。",
        )
        excluded = [stage for stage, _ in CARBON_STAGE_OPTIONS if stage not in set(selected_stages)]
        for stage in excluded:
            self._add_validation_finding(
                "INFO",
                "STAGE_EXCLUDED_BY_BOUNDARY",
                f"{stage} 未納入 {label} 邊界計算。",
                stage=stage,
                recommendation="若要供外部查證，請在報告中說明排除原因與邊界定義。",
            )

    def _add_validation_finding(
        self,
        severity,
        code,
        message,
        *,
        stage="",
        sheet="",
        row="",
        recommendation="",
    ):
        self._validation_findings.append(
            {
                "severity": severity,
                "code": code,
                "message": message,
                "stage": stage,
                "sheet": sheet,
                "row": row,
                "recommendation": recommendation,
            }
        )

    def _add_calculation_audit_row(self, **row):
        self._calculation_audit_rows.append(row)

    def _replace_workbook_sheet(self, workbook, sheet_name):
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        return workbook.create_sheet(sheet_name)

    @staticmethod
    def _excel_safe_value(value):
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except (TypeError, ValueError):
            pass
        if isinstance(value, np.generic):
            return value.item()
        if isinstance(value, (list, tuple, set)):
            return ", ".join(str(item) for item in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False, sort_keys=True)
        return value

    def _append_dict_rows(self, sheet, headers, rows):
        sheet.append(headers)
        for row in rows:
            sheet.append([self._excel_safe_value(row.get(header, "")) for header in headers])

    def _write_calculation_audit_sheet(self, workbook):
        headers = [
            "stage",
            "source_sheet",
            "source_row",
            "activity_name",
            "name_of_database",
            "quantity_column",
            "quantity",
            "source_unit",
            "factor_unit",
            "conversion_factor",
            "unspecified_factor",
            "fossil_factor",
            "biogenic_factor",
            "land_transformation_factor",
            "unspecified_kg_co2e",
            "fossil_kg_co2e",
            "biogenic_kg_co2e",
            "land_transformation_kg_co2e",
            "damage_assessment_formula",
            "coefficient_formula",
            "calculation_basis",
            "status",
        ]
        sheet = self._replace_workbook_sheet(workbook, "Calculation_Audit")
        self._append_dict_rows(sheet, headers, self._calculation_audit_rows)

    def _write_validation_findings_sheet(self, workbook):
        headers = [
            "severity",
            "code",
            "message",
            "stage",
            "sheet",
            "row",
            "recommendation",
        ]
        sheet = self._replace_workbook_sheet(workbook, "Validation_Findings")
        self._append_dict_rows(sheet, headers, self._validation_findings)

    def _write_iso14067_checklist_sheet(self, workbook, selected_stages, carbon_boundary):
        boundary = self.normalize_carbon_boundary(carbon_boundary)
        boundary_label = CARBON_BOUNDARY_LABELS[boundary]
        selected_stage_set = set(selected_stages)
        excluded = [stage for stage, _ in CARBON_STAGE_OPTIONS if stage not in selected_stage_set]
        rows = [
            {
                "item": "ISO 14067 聲明",
                "status": "需人工確認",
                "evidence": "本表為合理性檢核，不代表第三方查證或 ISO 14067 認證。",
                "recommendation": "正式對外聲明前需由授權標準文件與查證單位確認。",
            },
            {
                "item": "產品碳足跡邊界",
                "status": "已設定",
                "evidence": boundary_label,
                "recommendation": "搖籃到大門包含 Raw Material、Manufacturing；搖籃到墳墓包含五階段。",
            },
            {
                "item": "納入階段",
                "status": "已設定",
                "evidence": ", ".join(selected_stages),
                "recommendation": "確認所有納入階段皆有活動數據、係數與計算紀錄。",
            },
            {
                "item": "排除階段",
                "status": "需人工確認" if excluded else "不適用",
                "evidence": ", ".join(excluded) if excluded else "無",
                "recommendation": "若為搖籃到大門，排除 Distribution、Usage、Recycling 應於報告揭露。",
            },
            {
                "item": "功能單位 / 宣告單位",
                "status": "需人工確認",
                "evidence": "需對照 overview 與報告模板內容。",
                "recommendation": "確認產品數量基準、宣告單位與客戶交付格式一致。",
            },
            {
                "item": "資料期間與廠區",
                "status": "需人工確認",
                "evidence": "overview / INPUT 來源欄位已寫入處理流程。",
                "recommendation": "確認 start_date、end_date、factory_site 與活動數據期間一致。",
            },
            {
                "item": "排放係數來源",
                "status": "需人工確認",
                "evidence": "simapro10.2.0.0 / name of database 對應。",
                "recommendation": "確認 database 名稱、版本、地區、單位與係數來源可追溯。",
            },
            {
                "item": "allocation 方法",
                "status": "需人工確認",
                "evidence": "需對照活動數據中的 allocated proportion / allocation 欄位與公司分攤方法。",
                "recommendation": "確認 allocation 方法、分攤基準、比例範圍與每筆分攤計算可追溯；不適用時也應明確標註。",
            },
            {
                "item": "cut-off 準則",
                "status": "需人工確認",
                "evidence": "目前無法僅由輸出 workbook 判斷 cut-off 政策。",
                "recommendation": "確認排除流、排除比例、門檻與理由符合公司/查證要求；未使用 cut-off 時應明確標註。",
            },
            {
                "item": "DQR / 資料品質評分",
                "status": "需人工確認",
                "evidence": "需由公司資料品質方法、DQR 分數或 reviewer note 支持。",
                "recommendation": "確認時間、地域、技術代表性、完整性與可靠性評估方法，並保存人工審查紀錄。",
            },
            {
                "item": "資料缺口與排除流",
                "status": "需人工確認",
                "evidence": "Validation_Findings 工作表與人工盤查紀錄。",
                "recommendation": "所有缺資料、缺係數、單位未定義、排除階段與估算假設都需留存處置理由。",
            },
            {
                "item": "計算過程",
                "status": "已輸出",
                "evidence": "Calculation_Audit 工作表。",
                "recommendation": "逐列檢查 quantity、factor、conversion_factor、component result。",
            },
            {
                "item": "資料品質與異常",
                "status": "已輸出",
                "evidence": "Validation_Findings 工作表。",
                "recommendation": "所有 WARNING / ERROR 應於交付前處理或留存人工確認理由。",
            },
        ]
        sheet = self._replace_workbook_sheet(workbook, "ISO14067_Checklist")
        self._append_dict_rows(sheet, ["item", "status", "evidence", "recommendation"], rows)

    def _apply_boundary_to_workbook(self, workbook, selected_stages, carbon_boundary):
        selected_stage_set = set(selected_stages)
        boundary_label = CARBON_BOUNDARY_LABELS[self.normalize_carbon_boundary(carbon_boundary)]
        for stage_name, _ in CARBON_STAGE_OPTIONS:
            if stage_name in selected_stage_set or stage_name not in workbook.sheetnames:
                continue
            sheet = workbook[stage_name]
            sheet["AB2"] = f"Excluded by boundary: {boundary_label}"
            for cell in ("AC2", "AD2", "AE2", "AF2", "AG2"):
                sheet[cell] = 0

        if "overview" in workbook.sheetnames:
            overview = workbook["overview"]
            terms = [f"'{stage}'!AG2" for stage in selected_stages]
            overview["H2"] = f"={'+'.join(terms)}" if terms else 0

    def _write_process_review_sheets(self, workbook, selected_stages, carbon_boundary):
        self._write_calculation_audit_sheet(workbook)
        self._write_validation_findings_sheet(workbook)
        self._write_iso14067_checklist_sheet(workbook, selected_stages, carbon_boundary)

    def _process_file_impl(
        self,
        file_path=None,
        selected_stages=None,
        calculate_distances=True,
        carbon_boundary=None,
        use_transport_cache=True,
        force_recalculate_distances=False,
    ):
        """數據處理"""
        if file_path is not None:
            self.file_path = file_path
        active_boundary = self.normalize_carbon_boundary(carbon_boundary or self.carbon_boundary)
        self.carbon_boundary = active_boundary
        boundary_was_explicit = carbon_boundary is not None
        selected_stage_list = self._normalize_selected_carbon_stages(
            None if boundary_was_explicit else selected_stages,
            carbon_boundary=active_boundary if boundary_was_explicit or selected_stages is None else None,
        )
        selected_stage_set = set(selected_stage_list)
        self._reset_process_review_artifacts(selected_stage_list, active_boundary)
        self.last_error = None
        self.was_cancelled = False
        self._check_cancel()
        if not self.file_path:
            err_msg = "請選擇 Excel 文件"
            self.last_error = err_msg
            return {"ok": False, "error": err_msg}

        try:
            self._check_cancel()
            self.update_progress_smooth(0, 10, step=1, delay=0.5) # 階段1：讀取 Excel 檔案與資料準備，模擬從 0% 到 10%
            # 使用 openpyxl 讀取原始的 Excel 文件，保留原始格式和樣式
            self._notify_status("讀取 Excel 文件...")
            print("讀取 Excel 文件...")
            self._check_cancel()
            result_workbook = openpyxl.load_workbook(self.file_path, keep_vba=False, keep_links=False)
            stage_tables = {}
            for stage_name in selected_stage_list:
                stage_tables[stage_name] = self.read_multiple_tables(stage_name, self.file_path)

            if calculate_distances and {"Raw Material", "Distribution"} & selected_stage_set:
                self._notify_status("計算運輸距離：準備路線資料與快取...")
            if calculate_distances and "Raw Material" in selected_stage_set:
                self._notify_status("計算 Raw Material 運輸距離...")
                stage_tables["Raw Material"] = self.calculate_transport_distances(
                    result_workbook,
                    stage_tables["Raw Material"],
                    sheet_name="Raw Material",
                    table_indexes=[3, 4],
                    use_transport_cache=use_transport_cache,
                    force_recalculate_distances=force_recalculate_distances,
                )
            if calculate_distances and "Distribution" in selected_stage_set:
                self._notify_status("計算 Distribution 運輸距離...")
                stage_tables["Distribution"] = self.calculate_transport_distances(
                    result_workbook,
                    stage_tables["Distribution"],
                    sheet_name="Distribution",
                    use_transport_cache=use_transport_cache,
                    force_recalculate_distances=force_recalculate_distances,
                )
            
            self.update_progress_smooth(10, 40, step=1, delay=0.05) # 階段2：讀取工作表B，處理工作表並計算數值，模擬進度從 10% 到 40%
            # 以 pandas 讀入另一張關鍵對照表（sheet_B，如 simapro10.2.0.0）
            self._check_cancel()
            sheet_B = pd.read_excel(
                self.file_path,
                sheet_name='simapro10.2.0.0',
                usecols=['單位對照', *EMISSION_RESULT_COLUMNS, 'unit'],
            ).dropna(subset=['單位對照'])
            self._notify_status("處理工作表並獲取總值...")
            print("處理工作表並獲取總值...")
            stage_calc_columns = {
                "Raw Material": "W",
                "Manufacturing": "W",
                "Distribution": "U",
                "Usage": "Q",
                "Recycling": "Q",
            }
            stage_totals = {
                stage_name: tuple(0 for _ in EMISSION_RESULT_COLUMNS)
                for stage_name, _ in CARBON_STAGE_OPTIONS
            }
            for stage_name in selected_stage_list:
                stage_totals[stage_name] = self.process_tables(
                    stage_tables[stage_name],
                    stage_name,
                    stage_calc_columns[stage_name],
                    result_workbook,
                    sheet_B,
                )


            self.update_progress_smooth(40, 70, step=1, delay=0.02) # 階段3：更新報告模板，模擬進度從 40% 到 70%
            self._notify_status("讀取報告模板並寫入計算的數值...")
            print("讀取報告模板並寫入計算的數值...")
            report_path = self._get_required_resource_path(
                REPORT_WORKBOOK_TEMPLATE_FILENAME,
                "Excel 報告模板",
                os.path.join(self.base_dir, REPORT_WORKBOOK_TEMPLATE_FILENAME),
            )
            self._check_cancel()
            report_workbook = openpyxl.load_workbook(report_path)
            # 確保選擇報告中的 'general' 工作表
            if 'general' in report_workbook.sheetnames:
                report_sheet = report_workbook['general']
            else:
                raise ValueError("報告模板中未找到名為 'general' 的工作表。")

            self._notify_status("每個工作表的加總值寫入指定的單元格...")
            print("每個工作表的加總值寫入指定的單元格...")
            # 將每個工作表的四項 GWP 加總值寫入指定的單元格
            self._write_report_stage_totals(report_sheet, stage_totals)
            self.update_progress_smooth(70, 95, step=1, delay=0.05) # 階段4：儲存結果，模擬進度從 70% 到 99%
            # 獲取當前的日期和時間，用於生成檔案名稱
            current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
            self._apply_factory_overview_info(result_workbook)
            self._apply_boundary_to_workbook(
                result_workbook,
                selected_stage_list,
                active_boundary,
            )
            product_name_suffix = ""
            if "overview" in result_workbook.sheetnames:
                product_name_suffix = str(result_workbook["overview"]["C17"].value or "").strip()
            elif "INPUT" in result_workbook.sheetnames:
                product_name_suffix = str(result_workbook["INPUT"]["B1"].value or "").strip()
            if product_name_suffix:
                product_name_suffix = re.sub(r'[\\/:*?"<>|]+', '_', product_name_suffix)
                product_name_suffix = re.sub(r'\s+', '_', product_name_suffix).strip("._")
                product_name_suffix = product_name_suffix[:80]
            self._notify_status("保存更新後的報告，附上日期和時間...")
            print("保存更新後的報告，附上日期和時間...")
            # 保存更新後的報告，附上日期和時間
            if product_name_suffix:
                self.report_file = f'report_{product_name_suffix}_{current_time}.xlsx'
            else:
                self.report_file = f'report_{current_time}.xlsx'
            self.report_file = os.path.join(self.report_dir, self.report_file)
            print("路徑位置：", self.report_dir)
            report_workbook.save(self.report_file)
            with suppress(Exception):
                report_workbook.close()
            
            # 另存為新文件，保留原有的表格樣式，附上日期和時間
            if product_name_suffix:
                self.result_file = f'result_{product_name_suffix}_{current_time}.xlsx'
            else:
                self.result_file = f'result_{current_time}.xlsx'
            self.result_file = os.path.join(self.result_dir, self.result_file)
            self._check_cancel()
            self._sanitize_invalid_external_formulas(result_workbook)
            self._write_process_review_sheets(
                result_workbook,
                selected_stage_list,
                active_boundary,
            )
            result_workbook.save(self.result_file)
            with suppress(Exception):
                result_workbook.close()

            #    3. 用 Excel COM 自動修復並輸出最終結果
            self._check_cancel()
            try:
                path = os.path.abspath(self.result_file)
                if not os.path.exists(path):
                    raise FileNotFoundError(f"找不到結果檔：{path}")
                with ExcelComSession(
                    visible=False,
                    display_alerts=False,
                    enable_events=False,
                    screen_updating=False,
                    logger=self.logger,
                ) as session:
                    com_wb = session.open_workbook(
                        path,
                        CorruptLoad=1,
                        UpdateLinks=0,
                        ReadOnly=False,
                        IgnoreReadOnlyRecommended=True,
                        retry_count=self.runtime_config["open_retry_count"],
                        retry_delay_sec=self.runtime_config["open_retry_delay_sec"],
                        timeout_sec=self.runtime_config["open_timeout_sec"],
                    )
                    self._check_cancel()
                    session.excel.CalculateUntilAsyncQueriesDone()
                    session.save_with_retry(
                        com_wb,
                        retry_count=self.runtime_config["save_retry_count"],
                        retry_delay_sec=self.runtime_config["save_retry_delay_sec"],
                        timeout_sec=self.runtime_config["save_timeout_sec"],
                    )

            except UserCancelledError:
                raise
            except Exception as e:
                print(f"處理文件時出錯：{e}")
                err_msg = f"{e}\n{traceback.format_exc()}"
                self.last_error = err_msg
                return {"ok": False, "error": err_msg}  # 告知呼叫方：失敗

            return {
                "ok": True,
                "result_file": self.result_file,
                "report_file": self.report_file,
                "selected_stages": selected_stage_list,
                "carbon_boundary": active_boundary,
                "carbon_boundary_label": CARBON_BOUNDARY_LABELS[active_boundary],
            }
        except UserCancelledError as e:
            self.was_cancelled = True
            self.last_error = str(e)
            return {"ok": False, "cancelled": True, "error": str(e)}

        except Exception as e:
            if self._is_permission_denied_error(e):
                technical = traceback.format_exc()
                err_msg = self._file_permission_denied_message(self.file_path)
                self.last_error = err_msg
                self.last_technical_summary = self.summarize_technical_details(technical)
                return {
                    "ok": False,
                    "error_code": FILE_PERMISSION_DENIED_ERROR_CODE,
                    "message": err_msg,
                    "technical_details": technical,
                }
            err_msg = f"{e}\n{traceback.format_exc()}"
            self.last_error = err_msg
            return {"ok": False, "error": err_msg}




    def _excel_col_letter(self, col_num):
        """Return the Excel column letter for a 1-based column index."""
        result = ""
        while col_num:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _write_report_stage_totals(self, report_sheet, stage_totals):
        """Write four GWP split totals into report_temp.xlsx/general."""
        for row_offset, label in enumerate(REPORT_GENERAL_ROW_LABELS, start=2):
            report_sheet[f"A{row_offset}"] = label

        for stage_name, col_letter in REPORT_GENERAL_STAGE_COLUMNS.items():
            totals = stage_totals.get(
                stage_name,
                tuple(0 for _ in EMISSION_RESULT_COLUMNS),
            )
            for row_offset, value in enumerate(totals, start=2):
                cell = report_sheet[f"{col_letter}{row_offset}"]
                cell.value = value
                cell.number_format = "0.0000"
            report_sheet[f"{col_letter}6"] = f"=SUM({col_letter}2:{col_letter}5)"

    def _normalize_table_columns(self, header):
        """
        Make table headers safe for pandas operations.

        Excel templates may intentionally leave spacer columns blank. Pandas reads
        those blanks as NaN column names, and recent pandas versions reject merges
        when duplicate NaN columns are present.
        """
        normalized = []
        seen = set()

        for idx, col in enumerate(header, start=1):
            if pd.isna(col) or str(col).strip() == "":
                base_name = f"_blank_{self._excel_col_letter(idx)}"
            else:
                base_name = str(col).strip()

            candidate = base_name
            suffix = 1
            while candidate in seen:
                candidate = f"{base_name}_{suffix}"
                suffix += 1

            seen.add(candidate)
            normalized.append(candidate)

        return normalized

    def read_multiple_tables(self, sheet_name, file_path):
        """
        讀取工作表（如 Raw Material、Manufacturing 等）
        將每個工作表依據辨識B欄的◎符號，分為多個獨立的資料表格區段解析為 pandas DataFrame 清單​
        """
        # 讀取整個工作表，不設定標題行
        sheet = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        tables = []
        start_idx = 0

        # 遍歷所有行，辨識B欄的◎符號所在行來定位表格起始位置
        for idx, row in sheet.iterrows():
            # 檢查該行的B欄是否包含◎符號
            if '◎' in str(row[1]):
                # 如果已經找到一個表格的開始，將其保存
                if start_idx != 0:
                    # 保存表格並使用第三行作為欄位名稱
                    header = sheet.iloc[start_idx + 2]
                    table = sheet.iloc[start_idx + 3:idx].reset_index(drop=True)
                    table.columns = self._normalize_table_columns(header)
                    tables.append((start_idx, table))
                # 更新新的表格開始位置
                start_idx = idx

        # 添加最後一個表格，並使用第三行作為欄位名稱
        header = sheet.iloc[start_idx + 2]
        table = sheet.iloc[start_idx + 3:].reset_index(drop=True)
        table.columns = self._normalize_table_columns(header)
        tables.append((start_idx, table))

        # 返回表格數據
        return tables
    
    def process_tables(self, sheet_tables, sheet_name, col_start, workbook, sheet_B):
        """
        對工作表數據進行處理與加總
        根據不同表格使用不同的欄位進行計算
        並將數據進行單位換算
        """
        calculation_fields = [
            ("unspecified(kg CO2-eq)", "unspecified(kg CO2-eq)_result"),
            ("fossil(kg CO2-eq)", "fossil(kg CO2-eq)_result"),
            ("biogenic(kg CO2-eq)", "biogenic(kg CO2-eq)_result"),
            ("land transformation (kg CO2-eq)", "land transformation (kg CO2-eq)_result"),
        ]
        total_unspecified = 0
        total_fossil = 0
        total_biogenic = 0
        total_land_transformation = 0
        for i, (start_idx, sheet_data) in enumerate(sheet_tables):
            self._check_cancel()
            required_cols = ['name of database']
            missing = [c for c in required_cols if c not in sheet_data.columns]
            if missing:
                raise ValueError(
                    f"【{sheet_name} 第 {i+1} 個表格】缺少必要欄位: {missing}。\n"
                    f"實際欄位有: {list(sheet_data.columns)}"
                )
            # 使用 merge 函數來進行類似 VLOOKUP 的合併
            merged_df = sheet_data.merge(sheet_B, left_on='name of database', right_on='單位對照', how='left', suffixes=('', '_y'))

            # 根據不同表格使用不同的欄位進行計算
            if 'total' in sheet_data.columns:
                quantity_column = 'total'
            elif 'Ton‧Km' in sheet_data.columns:
                quantity_column = 'Ton‧Km'
            elif 'consumed amount allocated to single product (energy/product unit)' in sheet_data.columns:
                quantity_column = 'consumed amount allocated to single product (energy/product unit)'
            elif sheet_name in ['Recycling', 'Usage'] and 'total amount' in sheet_data.columns:
                quantity_column = 'total amount'
            else:
                raise ValueError(            
                    f"【{sheet_name} 第 {i+1} 個表格】缺少必要的數量欄位。\n"
                    f"需包含 'total'、'Ton‧Km'、'consumed amount allocated to single product (energy/product unit)' 或 'total amount'。\n"
                    f"實際欄位有: {list(sheet_data.columns)}"
                )

            quantity_col_idx = list(sheet_data.columns).index(quantity_column) + 1
            quantity_col_letter = self._excel_col_letter(quantity_col_idx)
            start_col_idx = column_index_from_string(col_start)
            output_columns = {}
            for offset, (output_name, _) in enumerate(calculation_fields):
                output_col_idx, _ = self._find_table_column(sheet_data.columns, output_name)
                if output_col_idx is not None:
                    output_columns[output_name] = self._excel_col_letter(start_col_idx + offset)
                else:
                    output_columns[output_name] = self._excel_col_letter(start_col_idx + offset)
                    self._warn(f"【{sheet_name} 第 {i+1} 個表格】找不到 {output_name} 欄位，略過該欄寫入。")
                    self._add_validation_finding(
                        "WARNING",
                        "OUTPUT_COLUMN_MISSING",
                        f"找不到 {output_name} 欄位，該分量不會寫回原表。",
                        stage=sheet_name,
                        sheet=sheet_name,
                        recommendation="確認 PLCI template 是否包含四個碳排分量欄位。",
                    )

            damage_col_idx, _ = self._find_table_column(sheet_data.columns, "Damage Assessment")
            damage_col_letter = self._excel_col_letter(start_col_idx + len(calculation_fields))
            if damage_col_idx is None:
                self._warn(f"【{sheet_name} 第 {i+1} 個表格】找不到 Damage Assessment 欄位，改寫入 {damage_col_letter} 欄。")
                self._add_validation_finding(
                    "WARNING",
                    "DAMAGE_ASSESSMENT_COLUMN_MISSING",
                    f"找不到 Damage Assessment 欄位，改寫入 {damage_col_letter} 欄。",
                    stage=sheet_name,
                    sheet=sheet_name,
                    recommendation="確認 template 欄位位置，避免客戶格式欄位偏移。",
                )

            coefficient_col_idx, _ = self._find_table_column(sheet_data.columns, "Coefficient value")
            coefficient_col_letter = (
                self._excel_col_letter(coefficient_col_idx)
                if coefficient_col_idx is not None
                else None
            )
            if coefficient_col_letter is None:
                self._warn(f"【{sheet_name} 第 {i+1} 個表格】找不到 Coefficient value 欄位，略過係數公式寫入。")
                self._add_validation_finding(
                    "WARNING",
                    "COEFFICIENT_COLUMN_MISSING",
                    "找不到 Coefficient value 欄位，略過係數公式寫入。",
                    stage=sheet_name,
                    sheet=sheet_name,
                    recommendation="若需交付計算過程，建議保留 Coefficient value 欄位或使用 Calculation_Audit。",
                )

            merged_df["_conversion_factor"] = 1.0
            merged_df["_factor_lookup_status"] = "OK"
            merged_df["_calculation_status"] = "OK"
            for _, result_col in calculation_fields:
                merged_df[f"{result_col}_coefficient"] = pd.NA
            
            # 1) 保留原本表格內既有的 result（若有）
            for output_name, result_col in calculation_fields:
                if output_name in merged_df.columns:
                    merged_df[result_col + "_manual"] = pd.to_numeric(merged_df[output_name], errors="coerce")
                elif result_col in merged_df.columns:
                    merged_df[result_col + "_manual"] = pd.to_numeric(merged_df[result_col], errors="coerce")
                else:
                    merged_df[result_col + "_manual"] = pd.NA

            # 2) 初始化計算欄位（先用手動值當預設）
            for _, result_col in calculation_fields:
                merged_df[result_col] = pd.to_numeric(
                    merged_df[result_col + "_manual"],
                    errors="coerce",
                ).astype("float64")

            # 判斷工作表和工作表B的單位是否一致
            for idx, row in merged_df.iterrows():
                self._check_cancel()
                source_unit = row.get("Unit")
                factor_unit = row.get("unit")
                if pd.notna(source_unit) and pd.notna(factor_unit):
                    if source_unit != factor_unit:
                        if source_unit in ['g', 'kg', 'ton'] and factor_unit in ['g', 'kg', 'ton']:
                            if source_unit == 'g' and factor_unit == 'kg':
                                conversion_factor = 1 / 1000
                            elif source_unit == 'g' and factor_unit == 'ton':
                                conversion_factor = 1 / 1000 / 1000 
                            elif source_unit == 'ton' and factor_unit == 'kg':
                                conversion_factor = 1 * 1000
                            elif source_unit == 'kg' and factor_unit == 'ton':
                                conversion_factor = 1 / 1000
                            elif source_unit == 'kg' and factor_unit == 'g':
                                conversion_factor = 1 * 1000
                            elif source_unit == 'ton' and factor_unit == 'g':
                                conversion_factor = 1 * 1000 * 1000
                            else:
                                conversion_factor = 1
                        else:
                            conversion_factor = 1
                            excel_row = start_idx + 4 + idx
                            merged_df.at[idx, "_calculation_status"] = "WARNING"
                            self._add_validation_finding(
                                "WARNING",
                                "UNIT_CONVERSION_UNCONFIRMED",
                                f"單位換算未定義，已暫用 1：{source_unit} -> {factor_unit}",
                                stage=sheet_name,
                                sheet=sheet_name,
                                row=str(excel_row),
                                recommendation="新增明確單位換算規則，或人工確認此列計算是否可接受。",
                            )
                    else:
                        conversion_factor = 1
                else:
                    conversion_factor = 1
                merged_df.at[idx, "_conversion_factor"] = conversion_factor
                
                # 檢查數值是否為數字類型，避免類似 TypeError 的錯誤
                quantity = self._coerce_numeric(row[quantity_column])
                if quantity is None:
                    excel_row = start_idx + 4 + idx
                    merged_df.at[idx, "_calculation_status"] = "WARNING"
                    self._add_validation_finding(
                        "WARNING",
                        "QUANTITY_MISSING",
                        f"數量欄位無法轉成數值：{quantity_column}",
                        stage=sheet_name,
                        sheet=sheet_name,
                        row=str(excel_row),
                        recommendation="確認活動數據數量欄位是否為數值，否則該列不會重新計算。",
                    )
                    continue

                # 係數查不到時不要覆蓋手動值（維持上面初始化的 manual）
                if pd.isna(row.get('單位對照')):
                    excel_row = start_idx + 4 + idx
                    merged_df.at[idx, "_factor_lookup_status"] = "MISSING_FACTOR"
                    merged_df.at[idx, "_calculation_status"] = "WARNING"
                    self._add_validation_finding(
                        "WARNING",
                        "EMISSION_FACTOR_NOT_FOUND",
                        f"name of database 無法對應 simapro10.2.0.0：{row.get('name of database')}",
                        stage=sheet_name,
                        sheet=sheet_name,
                        row=str(excel_row),
                        recommendation="確認 database 名稱、版本與單位對照表是否一致；必要時補齊係數來源。",
                    )
                    continue

                for output_name, result_col in calculation_fields:
                    coefficient_source_col = (
                        f"{output_name}_y"
                        if f"{output_name}_y" in merged_df.columns
                        else output_name
                    )
                    coefficient_value = self._coerce_numeric(row.get(coefficient_source_col))
                    if coefficient_value is None:
                        coefficient_value = 0
                    merged_df.at[idx, f"{result_col}_coefficient"] = coefficient_value
                    merged_df.at[idx, result_col] = quantity * coefficient_value * conversion_factor

            
            # 更新原始的工作表中的相關欄位，並小數點後 10 位無條件捨去
            sheet = workbook[sheet_name]
            for output_name, result_col in calculation_fields:
                output_col_letter = output_columns.get(output_name)
                if output_col_letter is None:
                    continue
                for idx, value in enumerate(merged_df[result_col], start=start_idx + 3):
                    if pd.notna(value):
                        truncated = math.trunc(float(value) * 10**10) / 10**10
                        sheet[f'{output_col_letter}{idx + 1}'] = truncated

            num_rows = len(merged_df)
            for i in range(num_rows):
                self._check_cancel()
                # Excel 的列號從 1 開始，所以 row_num 需要調整
                row_num = start_idx + 3 + i + 1  
                result_cells = [
                    f"{output_columns[output_name]}{row_num}"
                    for output_name, _ in calculation_fields
                    if output_name in output_columns
                ]
                # 將 Damage Assessment 欄位設為公式
                damage_cell = f'{damage_col_letter}{row_num}'
                sheet[damage_cell] = f"={'+'.join(result_cells)}" if result_cells else 0
                if coefficient_col_letter is not None:
                    quantity_cell = f"{quantity_col_letter}{row_num}"
                    sheet[f"{coefficient_col_letter}{row_num}"] = f'=IFERROR({damage_cell}/{quantity_cell},"")'

                audit_row = merged_df.iloc[i]
                if pd.isna(audit_row.get("name of database")) and self._coerce_numeric(audit_row.get(quantity_column)) is None:
                    continue
                self._add_calculation_audit_row(
                    stage=sheet_name,
                    source_sheet=sheet_name,
                    source_row=row_num,
                    activity_name=audit_row.get("Name", audit_row.get("name(chinese)", "")),
                    name_of_database=audit_row.get("name of database", ""),
                    quantity_column=quantity_column,
                    quantity=self._coerce_numeric(audit_row.get(quantity_column)),
                    source_unit=audit_row.get("Unit", ""),
                    factor_unit=audit_row.get("unit", ""),
                    conversion_factor=audit_row.get("_conversion_factor", 1),
                    unspecified_factor=audit_row.get("unspecified(kg CO2-eq)_result_coefficient", ""),
                    fossil_factor=audit_row.get("fossil(kg CO2-eq)_result_coefficient", ""),
                    biogenic_factor=audit_row.get("biogenic(kg CO2-eq)_result_coefficient", ""),
                    land_transformation_factor=audit_row.get("land transformation (kg CO2-eq)_result_coefficient", ""),
                    unspecified_kg_co2e=audit_row.get("unspecified(kg CO2-eq)_result", ""),
                    fossil_kg_co2e=audit_row.get("fossil(kg CO2-eq)_result", ""),
                    biogenic_kg_co2e=audit_row.get("biogenic(kg CO2-eq)_result", ""),
                    land_transformation_kg_co2e=audit_row.get("land transformation (kg CO2-eq)_result", ""),
                    damage_assessment_formula=f"={'+'.join(result_cells)}" if result_cells else "",
                    coefficient_formula=(
                        f'=IFERROR({damage_cell}/{quantity_col_letter}{row_num},"")'
                        if coefficient_col_letter is not None
                        else ""
                    ),
                    calculation_basis="component = quantity * emission_factor * conversion_factor",
                    status=audit_row.get("_calculation_status", "OK"),
                )
            
            self._notify_status("計算加總值並寫入每個表格的第一行...")
            # 計算加總值並寫入每個表格的第一行，並做小數點後 4 位四捨五入捨去
            table_totals = {
                output_name: round(merged_df[result_col].sum(), 10)
                for output_name, result_col in calculation_fields
            }

            # 計算加總值並寫入每階段的第一行
            unspecified_total = table_totals["unspecified(kg CO2-eq)"]
            fossil_total = table_totals["fossil(kg CO2-eq)"]
            biogenic_total = table_totals["biogenic(kg CO2-eq)"]
            land_transformation_total = table_totals["land transformation (kg CO2-eq)"]
            total_unspecified += unspecified_total
            total_fossil += fossil_total
            total_biogenic += biogenic_total
            total_land_transformation += land_transformation_total

            first_row_idx = start_idx + 1
            for output_name, _ in calculation_fields:
                output_col_letter = output_columns.get(output_name)
                if output_col_letter is not None:
                    sheet[f'{output_col_letter}{first_row_idx}'] = table_totals[output_name]
            first_row_damage_cell = f'{damage_col_letter}{first_row_idx}'
            first_row_result_cells = [
                f"{output_columns[output_name]}{first_row_idx}"
                for output_name, _ in calculation_fields
                if output_name in output_columns
            ]
            sheet[first_row_damage_cell] = (
                f"={'+'.join(first_row_result_cells)}"
                if first_row_result_cells
                else 0
            )
        self._notify_status("寫入所有表格的加總值...")
        # 在每個工作表的 AC/AD/AE/AF 欄位中寫入所有表格的加總值
        sheet[f'AB2'] = None
        sheet[f'AC2'] = round(total_unspecified, 10)
        sheet[f'AD2'] = round(total_fossil, 10)
        sheet[f'AE2'] = round(total_biogenic, 10)
        sheet[f'AF2'] = round(total_land_transformation, 10)
        # 在每個工作表的 AG 欄位中寫入 AC/AD/AE/AF 欄位的加總值
        sheet[f'AG2'] = round(total_unspecified + total_fossil + total_biogenic + total_land_transformation, 10)
        for summary_cell in ("AC2", "AD2", "AE2", "AF2", "AG2"):
            sheet[summary_cell].number_format = "0.0000"

        # 返回每個工作表的加總值
        return total_unspecified, total_fossil, total_biogenic, total_land_transformation
    
    def find_insert_positions(self, worksheet):
        """
        找出包含「◎」符號的行索引
        
        :param worksheet: xlsxwriter 工作表
        :return: 包含「◎」符號的行索引列表
        """
        insert_positions = []
        for row_num in range(1, worksheet.max_row + 1):
            for col_num in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value and '◎' in str(cell_value):
                    insert_positions.append(row_num - 1) # 轉換成 xlsxwriter 0起始索引
                    break
        return insert_positions
    


    def _hex6_from_openpyxl_color(self, col) -> str | None:
        """
        將 openpyxl 的 Color 物件或 rgb 值安全轉為 '#RRGGBB'，失敗回傳 None
        """
        if not col:
            return None
        # col 可能是 Color，也可能直接是字串/ RGB 物件
        raw = col.rgb if isinstance(col, Color) else col
        if raw is None:
            return None
        s = str(raw).strip()        # 有些情況是 RGB 物件，轉字串最保險
        if len(s) == 8:     # ARGB → RGB
            s = s[2:]
        # 也可能本來就 6 碼；若有 '#', 去掉
        if s.startswith('#'):
            s = s[1:]
        if re.fullmatch(r'[0-9A-Fa-f]{6}', s):
            return f'#{s}'
        return None
    
    def get_format_dict(self, cell):
        """
        讀取 openpyxl cell 的字體與填充設定，並轉換為 xlsxwriter 格式字典
        """
        fmt = {}
        # ---------- 字體 ----------
        font = cell.font
        if font.name:
            fmt['font_name'] = font.name
        if font.sz:
            fmt['font_size'] = font.sz
        if font.bold:
            fmt['bold'] = True
        if font.italic:
            fmt['italic'] = True
        if font.strike:
            fmt['font_strikeout'] = True

        # 底線：openpyxl 可能是 'single' / 'double' / 'singleAccounting' / 'doubleAccounting' / True
        if font.underline:
            ul = str(font.underline).lower()
            if ul in ('single', 'true'):
                fmt['underline'] = 1            # xlsxwriter 單底線
            elif ul == 'double':
                fmt['underline'] = 'double'
            elif ul == 'singleaccounting':
                fmt['underline'] = 'single_accounting'
            elif ul == 'doubleaccounting':
                fmt['underline'] = 'double_accounting'
            else:
                fmt['underline'] = 1
        # 字體顏色
        col = self._hex6_from_openpyxl_color(font.color)
        if col:
            fmt['font_color'] = col

        # 背景色
        # ---------- 填充（僅處理 solid） ----------
        fill = cell.fill
        if getattr(fill, 'fill_type', None) == 'solid':
            c = self._hex6_from_openpyxl_color(fill.fgColor)
            if c:
                fmt['bg_color'] = c
                fmt['pattern'] = 1  # xlsxwriter 設 bg_color 通常也要給 pattern

        # ---------- 對齊 ----------
        # 水平對齊：mapping 到 xlsxwriter
        h = cell.alignment.horizontal
        if h:
            h_map = {
                'general': None,
                'left': 'left',
                'center': 'center',
                'centercontinuous': 'center_across',  # Excel 的 centerContinuous
                'right': 'right',
                'fill': 'fill',
                'justify': 'justify',
                'distributed': 'distributed',
            }
            h_xl = h_map.get(h.lower())
            if h_xl:
                fmt['align'] = h_xl

        # 垂直對齊：xlsxwriter 使用 vcenter/vjustify/vdistributed
        v = cell.alignment.vertical
        if v:
            v_map = {
                'top': 'top',
                'center': 'vcenter',
                'bottom': 'bottom',
                'justify': 'vjustify',
                'distributed': 'vdistributed',
            }
            v_xl = v_map.get(v.lower())
            if v_xl:
                fmt['valign'] = v_xl

        # 自動換行（openpyxl 可能是 wrap_text 或 wrapText）
        wrap_val = getattr(cell.alignment, 'wrap_text', None)
        if wrap_val is None:
            wrap_val = getattr(cell.alignment, 'wrapText', None)
        if wrap_val:
            fmt['text_wrap'] = False


        # —— 邊框設定 ——
        b = cell.border
        styles = {b.left.style, b.right.style, b.top.style, b.bottom.style}
        style_map = {
        'hair': 1,      # xlsxwriter 沒有 hair，退而用最細
        'thin': 1,
        'dotted': 1,
        'dashdotdot': 1,
        'dashdot': 1,
        'dashed': 1,
        'medium': 2,
        'mediumDashed': 2,
        'mediumDashDot': 2,
        'mediumDashDotDot': 2,
        'double': 6,    # xlsxwriter 的 double 是 6（可用）
        'thick': 4,
        'slantDashDot': 2,
    }
        # 把 openpyxl 異大小寫樣式轉一致 key
        def _sty(s):
            return style_map.get(str(s), style_map.get(str(s).lower()))

        sides = {
            'top': b.top,
            'bottom': b.bottom,
            'left': b.left,
            'right': b.right,
        }

        side_styles = {side: (sides[side].style or None) for side in sides}
        style_set = {v for v in side_styles.values()}

        # 顏色：各邊採集 hex6，供後續判斷是否同色
        side_colors = {side: self._hex6_from_openpyxl_color(sides[side].color) for side in sides}
        color_set = {c for c in side_colors.values() if c}

        all_have_style = all(side_styles[s] is not None for s in sides)
        # 同邊框樣式
        if all_have_style and len(style_set) == 1:
            sty = next(iter(style_set))
            fmt['border'] = _sty(sty) or 1
            # 同色就給 border_color，不同就個別給
            if len(color_set) == 1:
                only = next(iter(color_set))
                if only:
                    fmt['border_color'] = only
            else:
                if side_colors['top']:
                    fmt['top_color'] = side_colors['top']
                if side_colors['bottom']:
                    fmt['bottom_color'] = side_colors['bottom']
                if side_colors['left']:
                    fmt['left_color'] = side_colors['left']
                if side_colors['right']:
                    fmt['right_color'] = side_colors['right']
        else:
            # 各邊分別處理
            if side_styles['top']:
                fmt['top'] = _sty(side_styles['top']) or 1
                if side_colors['top']:
                    fmt['top_color'] = side_colors['top']
            if side_styles['bottom']:
                fmt['bottom'] = _sty(side_styles['bottom']) or 1
                if side_colors['bottom']:
                    fmt['bottom_color'] = side_colors['bottom']
            if side_styles['left']:
                fmt['left'] = _sty(side_styles['left']) or 1
                if side_colors['left']:
                    fmt['left_color'] = side_colors['left']
            if side_styles['right']:
                fmt['right'] = _sty(side_styles['right']) or 1
                if side_colors['right']:
                    fmt['right_color'] = side_colors['right']

        if cell.number_format and cell.number_format != "General":
            fmt["num_format"] = cell.number_format

        return fmt




    def _get_format(self, fmt_dict, workbook):
        # 將 fmt_dict 轉成 tuple-of-tuples 作為 key（因為 dict 本身不可 hash）
        key = tuple(sorted(fmt_dict.items()))
        if key not in self._format_cache:
            self._format_cache[key] = workbook.add_format(fmt_dict)
        return self._format_cache[key]

    def _process_insert_positions(
        self,
        template_rows,
        base_insert_positions,
        source_sheet_list,
        source_data,
        target_sheet_name
    ):
        """根據插入點將來源資料插入模板列中。"""
        new_sheet_rows = template_rows.copy()
        offset = 0
        status_cb = getattr(self, "status_callback", None)

        for pos_idx, base_pos in enumerate(base_insert_positions):
            self._check_cancel()
            start_msg = f"開始處理插入點，pos_idx ={pos_idx}"
            print(start_msg)
            if status_cb:
                status_cb(start_msg)

            source_sheet_name = (
                source_sheet_list[pos_idx]
                if pos_idx < len(source_sheet_list)
                else None
            )
            source_label = source_sheet_name or f"{target_sheet_name}#{pos_idx + 1}"

            try:
                if source_sheet_name and source_sheet_name in source_data:
                    data = source_data[source_sheet_name]
                    num_data_rows = data.shape[0]
                    fixed_col_count = 27  # A-AA
                    data_rows = []
                    for i in range(num_data_rows):
                        self._check_cancel()
                        row = list(data.iloc[i])
                        if len(row) < fixed_col_count:
                            row.extend([""] * (fixed_col_count - len(row)))
                        elif len(row) > fixed_col_count:
                            row = row[:fixed_col_count]
                        data_rows.append(row)

                    # 原本預留◎符號所在列及其後兩列，共 3 列
                    insert_index = base_pos + 3 + offset

                    # 插入來源資料，將 data_rows 這個清單插入到 new_sheet_rows 的指定位置
                    new_sheet_rows[insert_index:insert_index] = data_rows
                    offset += num_data_rows
                    print(
                        f"在 {target_sheet_name} 的索引 {insert_index} 插入 {num_data_rows} 行來源資料"
                    )
                else:
                    warn_msg = (
                        f"模板中無對應來源資料，目標 {target_sheet_name} 插入點 {pos_idx + 1} 將略過（{source_label}）。"
                    )
                    print(warn_msg)
                    if status_cb:
                        status_cb(warn_msg)
            except Exception as e:
                err_msg = f"無法處理 {source_label} 工作表：{e}"
                print(err_msg)
                if status_cb:
                    status_cb(err_msg)

        return new_sheet_rows

    def _transform_sheet_impl(self):
        """
        將自動化Excel表單轉換成盤查表單格式：
        1. 用 openpyxl 讀取模板檔案（PLCI_table_format.xlsx），取得各工作表內容。
        2. 根據模板中◎符號所在行決定插入點：
           模板中原本預留◎符號所在行及後兩列（共3列）的區塊，
           若該工作表在指定清單中，則用該工作表前四列（格式定義）替換，
           並將來源資料插入於格式定義下方。
        3. 讀取來源資料（以 pandas DataFrame 形式），統計各工作表的行數。
        4. 將來源資料插入到模板內容中，同時調整後續內容位置。
        5. 利用 xlsxwriter 將調整後的所有內容寫入新檔案中。
        """
        if not self.file_path:
            self.last_error = "請選擇 Excel 文件"
            return False

        self.was_cancelled = False
        self._check_cancel()

        ok = False
        err_msg = None
        try:
            self._check_cancel()
            self._format_cache.clear()  # 清空格式快取
            self._notify_status("開始執行 Transform Sheet")
            print("開始執行 Transform Sheet")
            self.source_file_path = self.file_path
            input_values = {"product_name": "", "start_date": "", "end_date": ""}
            try:
                self._check_cancel()
                wb_input = openpyxl.load_workbook(self.source_file_path, read_only=True, data_only=True)
                if "overview" in wb_input.sheetnames:
                    ws_overview = wb_input["overview"]
                    input_values["product_name"] = ws_overview["C17"].value or ""
                if "INPUT" in wb_input.sheetnames:
                    ws_input = wb_input["INPUT"]
                    if not input_values["product_name"]:
                        input_values["product_name"] = ws_input["B1"].value or ""
                    input_values["start_date"] = ws_input["B2"].value or ""
                    input_values["end_date"] = ws_input["B3"].value or ""
            finally:
                try:
                    wb_input.close()
                except Exception:
                    pass

            target_file_path = self._get_required_resource_path(
                PLCI_TABLE_FORMAT_FILENAME,
                "PLCI 格式模板",
                os.path.join(self.base_dir, PLCI_TABLE_FORMAT_FILENAME),
            )
            # 用 openpyxl 讀取模板
            self._check_cancel()
            template_wb = openpyxl.load_workbook(target_file_path)
            print("PASS1...")
            self._notify_status("PASS1...")
            if self.update_progress_smooth:
                self.update_progress_smooth(0, 10, step=1, delay=0.02)  # 第1階段完成：10%
            # 建立格式定義字典，僅針對指定工作表
            format_definitions = {}
            for sheet_name in ['Raw Material', 'Manufacturing', 'Distribution', 'Recycling', 'Usage', 'overview']:
                if sheet_name in template_wb.sheetnames:
                    sheet = template_wb[sheet_name]
                    fd = [] # 空的串列
                    # 取工作表前五列，每個儲存格都以字典形式儲存 value 與其格式設定
                    for row in sheet.iter_rows(min_row=1, max_row=5):
                        current_row = []
                        for cell in row:
                            cell_info = {
                                "format": self.get_format_dict(cell)
                            }
                            current_row.append(cell_info)
                        fd.append(current_row)
                    format_definitions[sheet_name] = fd
                    print(f"取得 {sheet_name} 的格式定義，共 {len(fd)} 列")
                else:
                    print(f"模板中找不到工作表：{sheet_name}，無法取得格式定義")
            if self.update_progress_smooth:
                self.update_progress_smooth(1, 20, step=1, delay=0.02)  # 第2階段完成：20%
            # 設定來源資料的工作表對應關係
            self.source_sheets = {
                'Raw Material': [
                    'Raw Material(Direct Material)', 
                    'Raw Material(Indirect Material)', 
                    'Raw Material(Direct Transport)', 
                    'Raw Material(Indirect Transport'
                ],
                'Manufacturing': [
                    'Manufacturing(Manufacturing)', 
                    'Manufacturing(Gas)', 
                    'Manufacturing(Electricity)', 
                    'Manufacturing(Transport)', 
                    'Manufacturing(Waste treatment)'
                ],
                'Distribution': [
                    'Distribution(Local)', 
                    'Distribution(Air)', 
                    'Distribution(Warehouse)', 
                    'Distribution(Customer)'
                ],
                'Recycling': ['Recyling(Recyling)'],
                'Usage': ['Usage']
            }
            
            # 讀取來源資料，各工作表以 DataFrame 儲存
            source_data = {}
            for target_sheet_name, source_sheet_list in self.source_sheets.items():
                self._check_cancel()
                for sheet_name in source_sheet_list:
                    self._check_cancel()
                    try:
                        df = pd.read_excel(self.source_file_path, sheet_name=sheet_name)
                        # 將無限大值替換並填充空值，讓 xlsxwriter 能正確處理
                        df = df.replace([np.inf, -np.inf], 'Infinity')
                        df = df.fillna('')
                        source_data[sheet_name] = df
                        print(f"已讀取 {sheet_name} 工作表")
                        self._notify_status(f"已讀取 {sheet_name} 工作表")
                    except Exception as e:
                        print(f"警告: 無法讀取 {sheet_name} 工作表: {e}")
            print("讀取來源資料完成")
            if self.update_progress_smooth:
                self.update_progress_smooth(20, 30, step=1, delay=0.02)  # 第3階段完成：30%

            # 建立新的 xlsxwriter 工作簿
            current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
            product_name_suffix = str(input_values.get("product_name", "")).strip()
            if product_name_suffix:
                product_name_suffix = re.sub(r'[\\/:*?"<>|]+', '_', product_name_suffix)
                product_name_suffix = re.sub(r'\s+', '_', product_name_suffix).strip("._")
                product_name_suffix = product_name_suffix[:80]

            if product_name_suffix:
                new_file_name = f'merged_result_{product_name_suffix}_{current_datetime}.xlsx'
            else:
                new_file_name = f'merged_result_{current_datetime}.xlsx'
            new_file_path = os.path.join(self.result_dir, new_file_name)
            workbook = xlsxwriter.Workbook(new_file_path, {'nan_inf_to_errors': True})
            formula_entries = {}    

            if self.update_progress_smooth:
                self.update_progress_smooth(30, 80, step=1, delay=0.05) # 第4階段完成：80%
            # 處理每個目標工作表
            for target_sheet_name, source_sheet_list in self.source_sheets.items():
                self._check_cancel()
                print(f"處理目標工作表：{target_sheet_name}")
                worksheet = workbook.add_worksheet(target_sheet_name)
                if target_sheet_name in template_wb.sheetnames:
                    template_sheet = template_wb[target_sheet_name]
                else:
                    print(f"模板中不含 {target_sheet_name} 工作表，跳過此工作表")
                    continue
                
                template_rows = []
                for row in template_sheet.iter_rows(): 
                    current_row = []
                    for cell in row:
                        cell_info = {
                            "value": cell.value,
                            "format": self.get_format_dict(cell),  # 函式取得格式設定
                            "is_formula": cell.data_type == "f"
                        }
                        current_row.append(cell_info)
                    template_rows.append(current_row)

                # 找出模板中含有◎符號的行索引（0-based）
                base_insert_positions = []
                for idx, row in enumerate(template_rows):
                    if any(cell is not None and '◎' in str(cell) for cell in row):
                        base_insert_positions.append(idx)
                        print(f"在 {target_sheet_name} 模板中找到插入點：第 {idx} 行")

                # 複製模板內容作為最終輸出，並利用 offset 追蹤因插入或替換而產生的行偏移
                new_sheet_rows = self._process_insert_positions(
                    template_rows,
                    base_insert_positions,
                    source_sheet_list,
                    source_data,
                    target_sheet_name
                )

                default_format = [cell_info["format"] for cell_info in format_definitions[target_sheet_name][4]]
                self._fallback_fmt = workbook.add_format({
                                        'border': 1,
                                        'align': 'center',
                                        'valign': 'vcenter'
                                    })

                # 將最終結果寫入 xlsxwriter 工作表
                def _write_cell(ws, row_idx, col_idx, value, fmt):
                    if isinstance(value, str) and value.startswith("="):
                        ws.write_string(row_idx, col_idx, value, fmt)
                    else:
                        ws.write(row_idx, col_idx, value, fmt)

                for r, row in enumerate(new_sheet_rows):
                    self._check_cancel()
                    for c, cell in enumerate(row):
                        # 先取出值與格式 dict
                        if isinstance(cell, dict):
                            val = cell.get("value", "")
                            fmt_dict = cell.get("format") or {}
                            is_formula = cell.get("is_formula", False)
                        else:
                            val = cell
                            fmt_dict = {}
                            is_formula = False

                        if is_formula and isinstance(val, str):
                            formula = val if val.startswith("=") else f"={val}"
                            formula_entries.setdefault(target_sheet_name, []).append((r, c, formula))

                        if fmt_dict:
                            # 只有當 fmt_dict 裡真的有設定才建 format
                            # cell_fmt = workbook.add_format(fmt_dict, workbook)
                            cell_fmt = self._get_format(fmt_dict, workbook)
                            _write_cell(worksheet, r, c, val, cell_fmt)
                        else:
                            # 沒有自訂格式，就用 default_format
                            if isinstance(default_format, list) and c < len(default_format):
                                dfmt = default_format[c]  # 取出 column 對應的格式 dict
                                # cell_fmt = workbook.add_format(dfmt, workbook)
                                cell_fmt = self._get_format(dfmt, workbook)
                                _write_cell(worksheet, r, c, val, cell_fmt)
                            else:
                                # fallback 样式
                                cell_fmt = self._fallback_fmt
                            _write_cell(worksheet, r, c, val if val is not None else "", cell_fmt)

            if self.update_progress_smooth:
                self.update_progress_smooth(80, 95, step=1, delay=0.02) # 第5階段完成：95%
            self._check_cancel()
            workbook.close()
            print("靜態頁複製")
            self._notify_status("靜態頁複製")
            self._check_cancel()
            if not os.path.exists(target_file_path):
                raise FileNotFoundError(f"找不到範本：{target_file_path}")
            with ExcelComSession(
                visible=False,
                display_alerts=False,
                enable_events=False,
                screen_updating=False,
                logger=self.logger,
            ) as session:
                wb_tpl = session.open_workbook(
                    target_file_path,
                    CorruptLoad=1,
                    ReadOnly=True,
                    IgnoreReadOnlyRecommended=True,
                    retry_count=self.runtime_config["open_retry_count"],
                    retry_delay_sec=self.runtime_config["open_retry_delay_sec"],
                    timeout_sec=self.runtime_config["open_timeout_sec"],
                )
                self._check_cancel()
                wb_new = session.open_workbook(
                    new_file_path,
                    CorruptLoad=1,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True,
                    retry_count=self.runtime_config["open_retry_count"],
                    retry_delay_sec=self.runtime_config["open_retry_delay_sec"],
                    timeout_sec=self.runtime_config["open_timeout_sec"],
                )
                static_sheets = ['Instruction', 'overview', 'Process flow chart', 'simapro10.2.0.0']
                for sheet_name in static_sheets:
                    self._check_cancel()
                    try:
                        wb_tpl.Sheets(sheet_name).Copy(Before=wb_new.Sheets(1))
                    except Exception as e:
                        self._warn(f"複製「{sheet_name}」失敗：{e}")

                for sheet_name, formula_list in formula_entries.items():
                    self._check_cancel()
                    try:
                        ws_target = wb_new.Sheets(sheet_name)
                    except Exception as e:
                        self._warn(f"Formula writeback skipped: missing sheet {sheet_name}: {e}")
                        continue
                    for row_idx, col_idx, formula in formula_list:
                        self._check_cancel()
                        try:
                            ws_target.Cells(row_idx + 1, col_idx + 1).Formula = formula
                        except Exception as e:
                            self._warn(
                                f"Formula writeback failed: {sheet_name} {row_idx + 1},{col_idx + 1} {formula}: {e}"
                            )

                overview = wb_new.Sheets("overview")
                overview.Range("H2").Formula = "='Raw Material'!AG2+Manufacturing!AG2+Distribution!AG2+Recycling!AG2+Usage!AG2"
                overview.Range("V2").Formula = "=Usage!$K$5"
                overview.Range("C17").Value = input_values["product_name"]
                date_epoch = WINDOWS_EPOCH
                with suppress(Exception):
                    if wb_new.Date1904:
                        date_epoch = MAC_EPOCH
                overview.Range("C18").Value2 = self._to_excel_date_serial(input_values["start_date"], epoch=date_epoch)
                overview.Range("G18").Value2 = self._to_excel_date_serial(input_values["end_date"], epoch=date_epoch)
                overview.Range("C18").NumberFormat = "yyyy/m/d"
                overview.Range("G18").NumberFormat = "yyyy/m/d"

                factory_site = self.normalize_factory_site(self.factory_site)
                factory_info = FACTORY_OVERVIEW_INFO.get(factory_site)
                if not factory_site:
                    overview.Range("C3").Value = ""
                    overview.Range("C4").Value = ""
                elif factory_info:
                    overview.Range("C3").Value = factory_info["name"]
                    overview.Range("C4").Value = factory_info["address"]

                self._check_cancel()
                self._notify_status("重新整理 Excel 公式與查詢...")
                session.refresh_all_and_wait(
                    wb_new,
                    retry_count=self.runtime_config["refresh_retry_count"],
                    retry_delay_sec=self.runtime_config["refresh_retry_delay_sec"],
                    settle_sec=self.runtime_config["refresh_settle_sec"],
                    timeout_sec=self.runtime_config["refresh_timeout_sec"],
                    poll_sec=self.runtime_config["refresh_poll_sec"],
                    cancel_callback=self._check_cancel,
                    progress_callback=self._make_stage_progress_callback(95, 99),
                )
                self._notify_status("保存重新整理後的結果...")
                session.save_with_retry(
                    wb_new,
                    retry_count=self.runtime_config["save_retry_count"],
                    retry_delay_sec=self.runtime_config["save_retry_delay_sec"],
                    timeout_sec=self.runtime_config["save_timeout_sec"],
                )
                session.close_workbook(wb_tpl, save_changes=False)
                session.close_workbook(wb_new, save_changes=False)
                if self.update_progress_smooth:
                    self.update_progress_smooth(99, 100, step=1, delay=0.01)
                ok = True
            # 成功的回傳值
            self.merged_file = new_file_path
            return new_file_path
        except UserCancelledError as e:
            self.was_cancelled = True
            err_msg = str(e)
            self.last_error = err_msg
            return False
        except Exception as e:
            # 捕捉其他未預期錯誤
            tb = traceback.format_exc()
            if isinstance(e, PermissionError) or getattr(e, "errno", None) == 13:
                err_msg = "檔案開啟中無法讀取"
            else:
                err_msg = f"處理 Transform Sheet 時出錯：{e}\n{tb}"
            print(f"處理 Transform Sheet 時出錯：{e}\n{tb}")
            return False
        finally:
            # 把錯誤訊息傳回去（可用屬性或 callback）
            if not ok:
                # 方式 1：設成實例屬性，讓 GUI 執行緒讀取
                self.last_error = err_msg
                # 方式 2：有提供錯誤回呼就通知
                if getattr(self, "error_callback", None) and err_msg:
                    try:
                        self.error_callback(err_msg)
                    except Exception:
                        pass


    def _process_all_impl(self):
        """處理全部"""
        if not self.file_path:
            self.last_error = "請選擇 Excel 文件"
            return False

        try:
            self._notify_status("開始執行 Transform Sheet")
            transform_result = self.transform_sheet()
            if not transform_result.ok:
                self.last_error = transform_result.message
                return False
            merged_path = (
                transform_result.artifacts.get("path")
                or transform_result.artifacts.get("merged_file")
                or None
            )
            self._notify_status("Transform Sheet 完成，開始處理數據")
            process_result = self.process_file(file_path=merged_path)
            if not process_result.ok:
                self.last_error = process_result.message
                return False
            self._notify_status("處理全部完成")
            return True
        
        except Exception as e:
            self.last_error = f"處理全部過程中出現錯誤：{e}"
            return False

    def update_input_sheet(self, file_path, product="", start_date="", end_date="") -> TaskResult:
        run_id, started_at = self._start_task("update_input_sheet")
        if not file_path:
            result = TaskResult(
                ok=False,
                error_code="MISSING_INPUT",
                message="缺少要更新的 Excel 檔案路徑",
                artifacts={"run_id": run_id},
                elapsed_ms=self._elapsed_ms(started_at),
                warnings=list(self._warnings),
            )
            self._finish_task_log(run_id, "update_input_sheet", result)
            return result
        try:
            self._notify_status("開始更新 INPUT 工作表")
            self._emit_progress(0)
            with ExcelComSession(
                visible=False,
                display_alerts=False,
                enable_events=False,
                screen_updating=False,
                logger=self.logger,
            ) as session:
                workbook = session.open_workbook(
                    file_path,
                    ReadOnly=False,
                    retry_count=self.runtime_config["open_retry_count"],
                    retry_delay_sec=self.runtime_config["open_retry_delay_sec"],
                    timeout_sec=self.runtime_config["open_timeout_sec"],
                )
                self._emit_progress(10)
                for conn in workbook.Connections:
                    with suppress(Exception):
                        if hasattr(conn, "OLEDBConnection") and hasattr(conn.OLEDBConnection, "RefreshOnFileOpen"):
                            conn.OLEDBConnection.RefreshOnFileOpen = False
                    with suppress(Exception):
                        if hasattr(conn, "OLEDBConnection") and hasattr(conn.OLEDBConnection, "EnableRefresh"):
                            conn.OLEDBConnection.EnableRefresh = False

                ws = workbook.Worksheets("INPUT")
                input_range = ws.Range("B1:B3")
                current_values = input_range.Value
                if not current_values:
                    current_values = (("",), ("",), ("",))

                updated_values = [list(row) for row in current_values]
                if str(product).strip():
                    updated_values[0][0] = product
                if str(start_date).strip():
                    updated_values[1][0] = start_date
                if str(end_date).strip():
                    updated_values[2][0] = end_date

                input_range.Value = tuple((row[0],) for row in updated_values)
                self._emit_progress(20)

                self._notify_status("儲存 INPUT 更新內容...")
                session.save_with_retry(
                    workbook,
                    retry_count=self.runtime_config["save_retry_count"],
                    retry_delay_sec=self.runtime_config["save_retry_delay_sec"],
                    timeout_sec=self.runtime_config["save_timeout_sec"],
                )
                self._emit_progress(30)
                self._notify_status("重新整理 Excel 連線與公式...")
                session.refresh_all_and_wait(
                    workbook,
                    retry_count=self.runtime_config["refresh_retry_count"],
                    retry_delay_sec=self.runtime_config["refresh_retry_delay_sec"],
                    settle_sec=self.runtime_config["refresh_settle_sec"],
                    timeout_sec=self.runtime_config["refresh_timeout_sec"],
                    poll_sec=self.runtime_config["refresh_poll_sec"],
                    cancel_callback=self._check_cancel,
                    progress_callback=self._make_stage_progress_callback(30, 85),
                )
                self._notify_status("保存重新整理後的結果...")
                session.save_with_retry(
                    workbook,
                    retry_count=self.runtime_config["save_retry_count"],
                    retry_delay_sec=self.runtime_config["save_retry_delay_sec"],
                    timeout_sec=self.runtime_config["save_timeout_sec"],
                )
                self._emit_progress(95)
                session.close_workbook(workbook, save_changes=False)

            self._emit_progress(100)
            result = TaskResult(
                ok=True,
                message="INPUT 工作表更新完成",
                artifacts={"file_path": file_path, "run_id": run_id},
                elapsed_ms=self._elapsed_ms(started_at),
                warnings=list(self._warnings),
            )
        except UserCancelledError as exc:
            self.was_cancelled = True
            result = self._result_fail(
                error_code="USER_CANCELLED",
                user_message=str(exc),
                started_at=started_at,
                exc=exc,
            )
            result.artifacts["run_id"] = run_id
        except Exception as exc:
            result = self._result_fail(
                error_code="UPDATE_INPUT_FAILED",
                user_message="更新 INPUT 工作表失敗",
                started_at=started_at,
                exc=exc,
            )
            result.artifacts["run_id"] = run_id
        self._finish_task_log(run_id, "update_input_sheet", result)
        return result
        
    def update_excel_cache(self, result_file):
        """使用 Excel 更新公式快取值"""
        if result_file is None:
            result_file = getattr(self, "result_file", None)
        if not result_file or not os.path.exists(result_file):
            err_msg = f"找不到檔案：{result_file}"
            return False

        ok = False
        err_msg = None  

        try:
            self._check_cancel()
            with ExcelComSession(
                visible=False,
                display_alerts=False,
                enable_events=False,
                screen_updating=False,
                logger=self.logger,
            ) as session:
                wb = session.open_workbook(
                    os.path.abspath(result_file),
                    CorruptLoad=1,
                    UpdateLinks=0,
                    ReadOnly=False,
                    retry_count=self.runtime_config["open_retry_count"],
                    retry_delay_sec=self.runtime_config["open_retry_delay_sec"],
                    timeout_sec=self.runtime_config["open_timeout_sec"],
                )
                self._check_cancel()
                session.excel.CalculateUntilAsyncQueriesDone()
                session.save_with_retry(
                    wb,
                    retry_count=self.runtime_config["save_retry_count"],
                    retry_delay_sec=self.runtime_config["save_retry_delay_sec"],
                    timeout_sec=self.runtime_config["save_timeout_sec"],
                )
                session.close_workbook(wb, save_changes=False)
            ok = True
            return True
        except UserCancelledError as e:
            self.was_cancelled = True
            self.last_error = str(e)
            raise
        except Exception as e:
            err_msg = f"更新 Excel 快取值時發生錯誤：{e}"
            print(f"更新 Excel 快取值時發生錯誤：{e}")
            return False
        finally:
            # 把錯誤訊息傳回去（可用屬性或 callback）
            if not ok:
                # 方式 1：設成實例屬性，讓 GUI 執行緒讀取
                self.last_error = err_msg
                # 方式 2：有提供錯誤回呼就通知
                if getattr(self, "error_callback", None) and err_msg:
                    try:
                        self.error_callback(err_msg)
                    except Exception:
                        pass

    def _validate_report_source_workbook(self, result_file):
        required_sheets = [
            "overview",
            "Raw Material",
            "Manufacturing",
            "Distribution",
            "Usage",
            "Recycling",
        ]
        workbook = None
        try:
            workbook = openpyxl.load_workbook(result_file, read_only=True, data_only=False)
            missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
        except Exception as exc:
            raise ValueError(f"無法讀取報告來源檔案：{result_file}\n{exc}") from exc
        finally:
            if workbook is not None:
                with suppress(Exception):
                    workbook.close()

        if missing_sheets:
            missing_text = "、".join(missing_sheets)
            raise ValueError(f"報告來源檔案不是已處理盤查表單，缺少工作表：{missing_text}")

    def _split_stage_text(self, value):
        valid_stages = [stage for stage, _ in CARBON_STAGE_OPTIONS]
        if value is None:
            return []
        parts = re.split(r"[,，、;；]+", str(value))
        selected = []
        for part in parts:
            stage = part.strip()
            if stage in valid_stages and stage not in selected:
                selected.append(stage)
        return selected

    def _read_report_boundary_metadata(self, result_file):
        boundary = DEFAULT_CARBON_BOUNDARY
        boundary_label = CARBON_BOUNDARY_LABELS[boundary]
        included_stages = self.stages_for_carbon_boundary(boundary)
        boundary_total = None
        workbook = None
        try:
            workbook = openpyxl.load_workbook(result_file, read_only=True, data_only=True)
            if "ISO14067_Checklist" in workbook.sheetnames:
                sheet = workbook["ISO14067_Checklist"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    item, _status, evidence, _recommendation = (list(row) + [None, None, None, None])[:4]
                    if item == "產品碳足跡邊界" and evidence:
                        with suppress(ValueError):
                            boundary = self.normalize_carbon_boundary(evidence)
                            boundary_label = CARBON_BOUNDARY_LABELS[boundary]
                    elif item == "納入階段" and evidence:
                        parsed = self._split_stage_text(evidence)
                        if parsed:
                            included_stages = parsed
            else:
                inferred_stages = []
                for stage, _ in CARBON_STAGE_OPTIONS:
                    if stage not in workbook.sheetnames:
                        continue
                    marker = workbook[stage]["AB2"].value
                    if isinstance(marker, str) and marker.startswith("Excluded by boundary:"):
                        continue
                    inferred_stages.append(stage)
                if inferred_stages:
                    included_stages = inferred_stages
                if included_stages == self.stages_for_carbon_boundary("cradle_to_gate"):
                    boundary = "cradle_to_gate"
                    boundary_label = CARBON_BOUNDARY_LABELS[boundary]
            if "overview" in workbook.sheetnames:
                boundary_total = self._coerce_numeric(workbook["overview"]["H2"].value)
            if boundary_total is None:
                total = 0
                found_total = False
                for stage in included_stages:
                    if stage not in workbook.sheetnames:
                        continue
                    value = self._coerce_numeric(workbook[stage]["AG2"].value)
                    if value is None:
                        continue
                    total += value
                    found_total = True
                if found_total:
                    boundary_total = total
        except Exception as exc:
            self._warn(f"無法讀取報告邊界 metadata，改用預設搖籃到墳墓：{exc}")
            boundary = DEFAULT_CARBON_BOUNDARY
            boundary_label = CARBON_BOUNDARY_LABELS[boundary]
            included_stages = self.stages_for_carbon_boundary(boundary)
        finally:
            if workbook is not None:
                with suppress(Exception):
                    workbook.close()

        valid_stage_set = {stage for stage, _ in CARBON_STAGE_OPTIONS}
        included_stages = [stage for stage in included_stages if stage in valid_stage_set]
        if not included_stages:
            included_stages = self.stages_for_carbon_boundary(boundary)
        excluded_stages = [stage for stage, _ in CARBON_STAGE_OPTIONS if stage not in set(included_stages)]
        return {
            "boundary": boundary,
            "boundary_label": boundary_label,
            "included_stages": included_stages,
            "excluded_stages": excluded_stages,
            "boundary_total": boundary_total,
        }

    def _report_boundary_context(self, boundary_metadata):
        included_stages = boundary_metadata["included_stages"]
        excluded_stages = boundary_metadata["excluded_stages"]
        context = {
            "carbon_boundary": boundary_metadata["boundary_label"],
            "carbon_boundary_key": boundary_metadata["boundary"],
            "carbon_boundary_label": boundary_metadata["boundary_label"],
            "carbon_boundary_type": boundary_metadata["boundary"],
            "included_stages_text": "、".join(included_stages),
            "excluded_stages_text": "、".join(excluded_stages) if excluded_stages else "無",
            "included_lifecycle_stages": "、".join(included_stages),
            "excluded_lifecycle_stages": "、".join(excluded_stages) if excluded_stages else "無",
            "boundary_scope_statement": (
                f"本報告產品碳足跡邊界為{boundary_metadata['boundary_label']}，"
                f"納入階段：{'、'.join(included_stages)}；"
                f"排除階段：{('、'.join(excluded_stages) if excluded_stages else '無')}。"
            ),
            "iso14067_review_statement": "本報告為 ISO 14067 合理性檢核輔助，不代表第三方查證或 ISO 14067 認證。",
            "boundary_total": (
                f"{round(boundary_metadata['boundary_total'], 4)} kg CO2e"
                if boundary_metadata.get("boundary_total") is not None
                else ""
            ),
            "is_cradle_to_gate": boundary_metadata["boundary"] == "cradle_to_gate",
            "is_cradle_to_grave": boundary_metadata["boundary"] == "cradle_to_grave",
            "include_raw_material": "Raw Material" in included_stages,
            "include_manufacturing": "Manufacturing" in included_stages,
            "include_distribution": "Distribution" in included_stages,
            "include_usage": "Usage" in included_stages,
            "include_recycling": "Recycling" in included_stages,
        }
        included_set = set(included_stages)
        for stage, _label in CARBON_STAGE_OPTIONS:
            stage_key = re.sub(r'\W+', '_', stage).strip('_')
            included = stage in included_set
            context[f"{stage_key}_included_in_boundary"] = included
            context[f"{stage_key}_boundary_status"] = "納入" if included else "排除"
            context[f"{stage_key}_excluded_reason"] = "" if included else f"Excluded by boundary: {boundary_metadata['boundary_label']}"
        return context

    def _generate_report_impl(self, template_choice, result_file=None):
        """
        數據處理完後產生完整報告書流程：
        1. 根據 template_choice 選擇 Word 模板
        2. 使用指定的 result_file 或 self.result_file 作為數據來源，依序執行盤查表單各項函式：
            - 統整各工作表數據 (process_all_worksheets)
            - 將數據插入 Word (insert_data_to_word)
            - 生成圖表 (generate_bar_chart)
            - 針對 Raw Material、Manufacturing 等工作表進行細部處理與圖表生成
            - 前十大統整及運輸相關數據處理
        3. 最後將完整報告書存檔，檔名格式為 "智邦-產品碳足跡盤查總報告書_{today_date}.docx"
        
        """
        self.was_cancelled = False
        self._check_cancel()
        # 檢查是否已有數據處理過的檔案，才能進行
        selected_result_file = result_file or getattr(self, "result_file", "")
        if not selected_result_file:
            self.last_error = "請先處理檔案或選擇已處理盤查表單，再產生報告。"
            return False
        # === 1. 讀取 Excel 盤查表單，並開啟 Word 模板 ===
        # 使用先前數據處理後產生的檔案名稱
        result_file = os.path.abspath(selected_result_file)
        if not os.path.exists(result_file):
            self.last_error = f"找不到已處理盤查表單：{result_file}"
            return False
        self.result_file = result_file
        print(result_file)

        try:
            self._validate_report_source_workbook(result_file)
        except ValueError as e:
            self.last_error = str(e)
            return False

        # # test code
        # result_file = r'D:\OneDrive - Accton Technology Corporation\Python\code\Excel_Vlookup_Python\結果\result_20250519_174550.xlsx'
        # template_file= r'D:\OneDrive - Accton Technology Corporation\Python\code\Excel_Vlookup_Python\智邦-產品碳足跡盤查總報告書_竹南_temp.docx'

        if result_file:
            self._check_cancel()
            try:
                # 在讀取前先更新公式快取值，確保公式計算後的值有被存入檔案中
                cache_ok = self.update_excel_cache(result_file)
                if cache_ok is False:
                    return False
            except Exception as e:
                messagebox.showerror("錯誤", f"{e}")
                return  False

        self._check_cancel()
        # 依據 template_choice 選擇不同模板
        template_filename = REPORT_TEMPLATE_FILENAMES.get(template_choice)
        if not template_filename:
            messagebox.showerror("錯誤", "未知的報告模板選項")
            return
        template_file = self._get_required_resource_path(
            template_filename,
            "Word 報告模板",
            os.path.join(self.base_dir, template_filename),
        )
        if self.update_progress_smooth:
            self.update_progress_smooth(0, 10, step=1, delay=0.02)  # 第一階段完成：10%
        
        # 開啟選定的 Word 模板
        try:
            self._check_cancel()
            doc = Document(template_file)
        except Exception as e:
            print("錯誤", f"開啟 Word template 失敗：{e}")
            return  False
        
        if self.update_progress_smooth:
            self.update_progress_smooth(10, 20, step=1, delay=0.02)  # 第二階段完成：20%

        # === 2. 定義工作表名稱，讀取盤查表單存放資料至 context 清單 ===
        boundary_metadata = self._read_report_boundary_metadata(result_file)
        sheet_names = boundary_metadata["included_stages"]
        transport_sheets = [
            stage
            for stage in ['Raw Material', 'Manufacturing', 'Distribution']
            if stage in set(sheet_names)
        ]
        boundary_context = self._report_boundary_context(boundary_metadata)

        self._notify_status("讀取數據處理後產生的檔案...")
        self._check_cancel()
        df = pd.read_excel(result_file, sheet_name="overview") # 讀取盤查表單'overview'所需的欄位數值​

        today_date = datetime.today().strftime("%Y-%m-%d_%H%M%S")
        common_context = {'today_date': today_date,
                        'year': datetime.today().strftime("%Y"),
                        'month': datetime.today().strftime("%m")}
        common_context.update(boundary_context)
        if self.update_progress_smooth:
            self.update_progress_smooth(20, 30, step=1, delay=0.02)  # 第三階段完成：30%
        # 建立存放各筆資料的 context 清單
        all_contexts = []
        for _, row in df.iterrows():
            self._check_cancel()
            if pd.isna(row['start_date']) or pd.isna(row['end_date']):
                continue

            self.context = {
                'product_name': row['product_name'],
                'product_module': row['product_module'],
                'product_size': row['product_size'],
                'Gross_weight': row['product_weight'],
                'Net_weight': row['product_net_weight'],
                'Power': row['product_on_mode_Power'],
                'start_date': row['start_date'].strftime('%Y年%m月%d日'),
                'end_date': row['end_date'].strftime('%Y年%m月%d日'),
                # 'warranty': row['warranty'],
                'report_year': row['start_date'].strftime('%Y年'),
            }
            # 將共用參數加入每筆資料中
            self.context.update(common_context)
            all_contexts.append(self.context)
        if self.update_progress_smooth:
            self.update_progress_smooth(30, 40, step=1, delay=0.02)  # 第四階段完成：40%

        if all_contexts:
            try:
                self._check_cancel()
                # 建立 DocxTemplate 物件
                doc = DocxTemplate(template_file)
                # 模板中可使用 {% for item in all_contexts %} ... {% endfor %} 來逐筆列印資料
                doc.render(self.context) #使用 docxtpl 模組來套用這些資料到 Word 模板中
                full_output_name = f"智邦-產品碳足跡盤查總報告書_{today_date}.docx"   #命名output_doc
                full_output_path = os.path.join(self.tmp_dir, full_output_name)
                doc.save(full_output_path)
            except Exception as e:
                messagebox.showerror("錯誤", f"生成報告時發生錯誤：{e}")
                return
        else:
            messagebox.showwarning("警告", "匯入為空值，未生成 Word 文件")
            return False
        if self.update_progress_smooth:
            self.update_progress_smooth(40, 50, step=1, delay=0.02)  # 第五階段完成：50%

        # === 3. 以盤查表單作為基底，繼續處理其數據與圖表，生成完整報告書 ===
        # 呼叫各個盤查表單統整計算函式，將數據與圖表插入報告中
        self._notify_status("呼叫各個盤查表單統整計算函式，將數據與圖表生成...")
        self._check_cancel()
        all_results = self.process_all_worksheets(result_file, sheet_names)
        self._check_cancel()
        self.insert_data_to_word(all_results, sheet_names)
        self._check_cancel()
        self.generate_bar_chart(doc, all_results, sheet_names)
        if self.update_progress_smooth:
            self.update_progress_smooth(50, 60, step=1, delay=0.02)  # 第六階段完成：60%
        
        # Raw Material 處理與圖表生成
        self._notify_status("Raw Material 處理與圖表生成...")
        self._check_cancel()
        resulall_data_1, Raw_data = self.process_worksheet(result_file, 'Raw Material')
        self._check_cancel()
        self.process_insert_raw_data(result_file)
        self._check_cancel()
        self.generate_insert_raw_charts(doc, Raw_data)
        if self.update_progress_smooth:
            self.update_progress_smooth(60, 70, step=1, delay=0.02)  # 第七階段完成：70%
        
        # Manufacturing 處理與圖表生成
        print("Manufacturing 處理與圖表生成...")
        self._check_cancel()
        resulall_data_2, Manu_data = self.process_worksheet(result_file, 'Manufacturing')
        self._check_cancel()
        self.process_insert_manufacturing_data(result_file)
        self._check_cancel()
        self.generate_insert_manufacturing_charts(doc, Manu_data)
        self._check_cancel()
        self.generate_and_insert_electric_chart(doc, resulall_data_2)
        if self.update_progress_smooth:
            self.update_progress_smooth(70, 80, step=1, delay=0.02)  # 第八階段完成：80%
        
        # 前十大統整處理
        self._notify_status("前十大統整處理與圖表生成...")
        self._check_cancel()
        self.process_top10_data(sheet_names, result_file, doc)
        if self.update_progress_smooth:
            self.update_progress_smooth(80, 95, step=1, delay=0.02)  # 第十階段處理完畢前：95%
        
        # 運輸相關數據處理
        self._notify_status("運輸相關數據處理與圖表生成...")
        self._check_cancel()
        Air_all_data = self.process_transport_data(result_file, transport_sheets)
        self._check_cancel()
        self.analyze_and_chart_generate(Air_all_data, doc)

        # 將儲存在 self.context  的數據 & 圖表匯入
        self._notify_status("所有數據與圖表匯入報告書...")
        self._check_cancel()
        doc.render(self.context)    

        # === 4. 存檔完整報告書 ===
        self._notify_status("保存文件...")
        self._check_cancel()
        full_report_file = os.path.join(
            self.report_dir, f"智邦-產品碳足跡盤查總報告書_{today_date}.docx")
        doc.save(full_report_file)
        if self.update_progress_smooth:
            self.update_progress_smooth(95, 100, step=1, delay=0.02)  # 完全完成：100%
        print(f"【Finished】報告書匯入已完成_產品碳足跡盤查總報告書_{today_date}")

        return full_report_file

    def process_worksheet(self, file_name, sheet_name):
        """處理單個表單的數據，返回結果字典和整合數據框。"""
        self._check_cancel()
        df = pd.read_excel(file_name, sheet_name=sheet_name)  
        group_starts = df.index[df.iloc[:, 1].str.contains('^◎', na=False)].tolist()
        # 初始化一个空的字典，用于存储每个数据组的结果
        resulall_data = {}
        all_data = pd.DataFrame()    
        # 循环处理每个数据群组
        for j in range(len(group_starts)):
            self._check_cancel()
            start_idx = group_starts[j]
            end_idx = group_starts[j + 1] if j < len(group_starts) - 1 else df.shape[0]

            # 使用切片选择每个数据群组的数据
            group_data = df.iloc[start_idx:end_idx, :]

            # 删除第一列和第二列的无效数据，并将第三列作为列标题
            group_data = group_data.iloc[2:, 1:].copy()
            group_data.columns = group_data.iloc[0, :]
            group_data = group_data.iloc[1:, :].copy()

            num_cols = [
                'unspecified(kg CO2-eq)',
                'fossil(kg CO2-eq)',
                'biogenic(kg CO2-eq)',
                'land transformation (kg CO2-eq)',
                'Damage Assessment',
            ]
            for c in num_cols:
                if c not in group_data.columns:
                    group_data[c] = 0
            # 1) 型別轉換與空值補 0
            for c in num_cols:
                self._check_cancel()
                group_data[c] = pd.to_numeric(group_data[c], errors='coerce').fillna(0)
            # 2) 過濾：只保留「至少一個數值欄位非 0」的列
            mask = group_data[num_cols].sum(axis=1) != 0
            group_data = group_data.loc[mask]
            # 3) 把 Name 欄原本的空值 (NaN) 補成一個自訂標籤
            group_data['Name'] = group_data['Name'].fillna('空白群組')

            # 處理 name of database 欄位，将不同的值合并为一个字符串，使用分号分隔
            grouped_c = group_data.groupby('Name')['name of database'].apply(
                lambda x: ';'.join(sorted(set(x.dropna())))).reset_index()
            # 處理 unspecified(kg CO2-eq) 欄位，将它们加总
            unspecified_values = group_data.groupby('Name')['unspecified(kg CO2-eq)'].sum().reset_index()
            # 處理 fossil(kg CO2-eq) 欄位，将它们加总
            fossil_values = group_data.groupby('Name')['fossil(kg CO2-eq)'].sum().reset_index()
            # 處理 biogenic(kg CO2-eq) 欄位，将它们加总
            biogenic_values = group_data.groupby('Name')['biogenic(kg CO2-eq)'].sum().reset_index()
            # 處理 land transformation (kg CO2-eq) 欄位，将它们加总
            land_values = group_data.groupby('Name')['land transformation (kg CO2-eq)'].sum().reset_index()
            # 處理 Damage Assessment 欄位，将它们加总
            summed_values = group_data.groupby('Name')['Damage Assessment'].sum().reset_index()
            
            # 合并 grouped_c, fossil_values, biogenic_values, land_values, summed_values，以 'Name' 为键
            data_frames = [grouped_c, unspecified_values, fossil_values, biogenic_values, land_values, summed_values]
            merged_data = reduce(lambda left,right: pd.merge(left, right, on='Name', how='outer'), data_frames)
            merged_data = merged_data.sort_values(by='Damage Assessment', ascending=False)
            print(f"{sheet_name} group {j + 1}: summarized {len(merged_data)} rows")
            # 依據 'Damage Assessment' 列的數值大小降序排序
            
            # 将每个数据群组的结果添加到字典中
            resulall_data[f'G{j + 1}'] = merged_data
            # 將每個資料群組的整合數據添加到 all_data 中
            all_data = pd.concat([all_data, merged_data], axis=0)
            
        all_data = all_data.sort_values(by='Damage Assessment', ascending=False)    
        return resulall_data, all_data

    def process_all_worksheets(self, file_name, sheet_names):
        """處理多個表單的數據，返回所有結果。"""
        self._check_cancel()
        all_results = {}
        for sheet in sheet_names:
            self._check_cancel()
            resulall_data, all_data = self.process_worksheet(file_name, sheet)
            all_results[sheet] = {'resulall_data': resulall_data, 'all_data': all_data}
            # print(all_results)
        return all_results
        
    def insert_data_to_word(self, all_results, sheet_names):
        """
        將數據插入 Word 文件中的指定標籤位置。

        Parameters:
        - doc: Document，Word 文件對象。
        - data_mapping: dict，標籤與數據的對應字典，例如 {'[TAG_1]': 'value1', '[TAG_2]': 'value2'}。
        """
        self._check_cancel()
        print("【Process_2】開始將數據匯入 Word 文件")
        # 遍歷文檔中的所有段落，尋找標籤
        total_damage_assessment = 0
        emission_totals = {col: 0 for col in EMISSION_RESULT_COLUMNS}
        for sheet in all_results.keys():
            self._check_cancel()
            df = all_results[sheet]['all_data']
            total_damage_assessment += pd.to_numeric(
                df.get('Damage Assessment', pd.Series(dtype=float)),
                errors='coerce',
            ).fillna(0).sum()
            for emission_col in EMISSION_RESULT_COLUMNS:
                emission_totals[emission_col] += pd.to_numeric(
                    df.get(emission_col, pd.Series(dtype=float)),
                    errors='coerce',
                ).fillna(0).sum()
         # 將碳排五階段統整的數值儲存至self.context
        selected_stage_set = set(sheet_names)
        for stage_name, _label in CARBON_STAGE_OPTIONS:
            if stage_name in selected_stage_set:
                continue
            stage_key = re.sub(r'\W+', '_', stage_name).strip('_')
            self.context[f'{stage_key}_unspecified'] = "0.0000"
            self.context[f'{stage_key}_fossil'] = "0.0000"
            self.context[f'{stage_key}_biogenic'] = "0.0000"
            self.context[f'{stage_key}_land'] = "0.0000"
            self.context[f'{stage_key}_sum'] = "0.0000"
            self.context[f'{stage_key}_Total_percentage'] = "0%"
        sum_list = [] 
        for sheet in sheet_names:
            self._check_cancel()
            sheet_key = re.sub(r'\W+', '_', sheet).strip('_')
            df = all_results[sheet]['all_data']
            stage_emission_totals = {
                emission_col: pd.to_numeric(
                    df.get(emission_col, pd.Series(dtype=float)),
                    errors='coerce',
                ).fillna(0).sum()
                for emission_col in EMISSION_RESULT_COLUMNS
            }
            sum    = pd.to_numeric(
                df.get('Damage Assessment', pd.Series(dtype=float)),
                errors='coerce',
            ).fillna(0).sum()
            percentage  = sum / total_damage_assessment * 100 if total_damage_assessment else 0
            names = (
                df['name of database']
                  .dropna()
                  .astype(str)
                  .unique()
                  .tolist()
            )
            sum_list.append((sheet_key, sum, names))

            for emission_col, suffix in EMISSION_CONTEXT_SUFFIXES.items():
                self.context[f'{sheet_key}_{suffix}'] = self._format_report_number(
                    stage_emission_totals[emission_col]
                )
            self.context[f'{sheet_key}_sum']              = self._format_report_number(sum)
            self.context[f'{sheet_key}_Total_percentage'] = f"{round(percentage, 2)}%"

        sorted_sums = sorted(sum_list, key=lambda x: x[1], reverse=True)[:5]

        for idx, (sheet_key, val, names) in enumerate(sorted_sums, start=1):
            self._check_cancel()
            pct = val / total_damage_assessment * 100 if total_damage_assessment else 0
            # 存到 self.context
            self.context[f'Carbon_percentage_{idx}'] = f"{pct:.2f}%"
            self.context[f'Carbon_stage_{idx}'] = f"{sheet_key}階段"
            self.context[f'Carbon_name_{idx}'] = ";".join(names)
        for idx in range(len(sorted_sums) + 1, 6):
            self.context[f'Carbon_percentage_{idx}'] = ""
            self.context[f'Carbon_stage_{idx}'] = ""
            self.context[f'Carbon_name_{idx}'] = ""
        self.context['sum_percentage_unspecified'] = (
            f"{round(emission_totals[UNSPECIFIED_EMISSION_COLUMN] / total_damage_assessment * 100, 2)}%"
            if total_damage_assessment else "0%"
        )
        self.context['sum_percentage_1'] = (
            f"{round(emission_totals['fossil(kg CO2-eq)'] / total_damage_assessment * 100, 2)}%"
            if total_damage_assessment else "0%"
        )
        self.context['sum_percentage_2'] = (
            f"{round(emission_totals['biogenic(kg CO2-eq)'] / total_damage_assessment * 100, 2)}%"
            if total_damage_assessment else "0%"
        )
        self.context['sum_percentage_3'] = (
            f"{round(emission_totals['land transformation (kg CO2-eq)'] / total_damage_assessment * 100, 2)}%"
            if total_damage_assessment else "0%"
        )
        self.context['Total'] = f"{self._format_report_number(total_damage_assessment)} kg CO2e"

        print("【Process_2】已匯入全階段統計表格數值") 

        # 初始化用於儲存 [Total_percentage_i] 的DataFrame
        total_percentage_df = pd.DataFrame(columns=['Sheet', 'Total_Percentage'])
        # 假設 all_results 已經被填充了數據
        for sheet in sheet_names:
            self._check_cancel()
            # 假設你已經有了每個工作表的 total_damage_assessment 值
            total_percentage = (
                all_results[sheet]['all_data']['Damage Assessment'].sum() / total_damage_assessment * 100
                if total_damage_assessment
                else 0
            )
            # 將數據添加到DataFrame中
            total_percentage_df = pd.concat([total_percentage_df, pd.DataFrame({'Sheet': [sheet], 'Total_Percentage': [total_percentage]})], ignore_index=True)
        # 根據 Total_Percentage 降序排序
        total_percentage_df.sort_values(by='Total_Percentage', ascending=False, inplace=True)
        print("total_percentage_df:", total_percentage_df)

        for j, row in total_percentage_df.iterrows():
            self._check_cancel()
            self.context[f'Sheet_name_{j+1}']       = row['Sheet']
            self.context[f'Total_percentage_{j+1}'] = f"{round(row['Total_Percentage'],2)}%"
        for idx in range(len(total_percentage_df) + 1, 6):
            self.context[f'Sheet_name_{idx}'] = ""
            self.context[f'Total_percentage_{idx}'] = ""

    def generate_bar_chart(self, doc, all_results, sheet_names):
        """
        生成全階段的長條圖、堆疊長條圖，並將圖表插入 Word 文件對應標籤的位置。

        Parameters:
        - doc: Document
            Word 文件的 Document 物件。
        - all_results: dict
            每個表單處理結果的字典，例如：
            {
                'Raw Material': {
                    'all_data': <DataFrame>,
                    'resulall_data': <dict_of_dataframes>
                },
                'Manufacturing': {...},
                ...
            }
        - sheet_names: list
            記錄各個工作表名稱的清單，如 ['Raw Material', 'Manufacturing', 'Distribution', ...]
        """
        self._check_cancel()
        print("【Process_3】開始生成長條圖")
        # ------------------- 1. 計算各 Sheet 的 Damage Assessment 百分比長條圖 (bar_chart_1) -------------------
        # 先計算 total_damage_assessment
        total_damage_assessment = sum(
            all_results[sheet]['all_data']['Damage Assessment'].sum() 
            for sheet in all_results
        )
        # 创建一个颜色列表，包含前十项的颜色和一个总和项的颜色
        colors = ['#FF9D47', '#F03535', '#027671', '#0033AA', '#04DCCE', 'grey']
        # 计算每个工作表的百分比
        percentages = []
        sheet_labels = []
        for sheet in all_results:
            self._check_cancel()
            sheet_sum = all_results[sheet]['all_data']['Damage Assessment'].sum()
            percentage = (sheet_sum / total_damage_assessment) * 100 if total_damage_assessment else 0
            percentages.append(percentage)
            sheet_labels.append(sheet)
        # 繪製 bar_chart_1
        plt.figure(figsize=(10, 6))  # 設定圖表大小
        bars = plt.bar(sheet_labels, percentages, color=colors, width=0.2)  # 創建條形圖
        # 添加 X/Y 軸標籤與標題
        plt.xlabel('Sheet Name')
        plt.ylabel('Percentage of Total Damage Assessment')
        plt.title('Percentage of Damage Assessment by Sheet')
        # 在每個 bar 上方添加數值標籤
        for i, bar in enumerate(bars):
            self._check_cancel()
            bar.set_label(sheet_names[i])  # 如果您想在 legend 中顯示 sheet_names[i]
            yval = bar.get_height()
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                yval,
                f'{round(yval, 2)}%',
                va='bottom',
                ha='center'
            )
        plt.xticks(rotation=0)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(labels=list(sheet_names), loc='upper right')
        plt.tight_layout()  # 確保標籤、標題不重疊
        bar_chart_1_path = self._chart_path("bar_chart_1.png")
        plt.savefig(bar_chart_1_path, bbox_inches='tight')
        # plt.show()
        # ------------------- 2. 產生各 Sheet 在四個 GWP 分項的佔比 (bar_chart_2) -------------------
        categories = list(EMISSION_RESULT_COLUMNS)
        category_data = {category: [] for category in categories}
        sheet_labels = list(all_results.keys())  # 重新整理 labels

        # 計算 percentage
        for category in categories:
            self._check_cancel()
            for sheet in sheet_labels:
                self._check_cancel()
                category_value = pd.to_numeric(
                    all_results[sheet]['all_data'].get(category, pd.Series(dtype=float)),
                    errors='coerce',
                ).fillna(0).sum()
                percentage = (category_value / total_damage_assessment) * 100 if total_damage_assessment > 0 else 0
                category_data[category].append(percentage)

        bar_width = 0.15  # 每個 bar 的寬度
        #category_spacing = 0.8  # 类别间的额外空间
        index = np.arange(len(categories))  # X 軸位置

        plt.figure(figsize=(10, 6))
        bars_all = []  # 用於存放所有條形的物件引用

        for i, sheet in enumerate(sheet_labels):
            self._check_cancel()
            bar_positions = index + i * bar_width
            bar = plt.bar(
                bar_positions,
                [category_data[cat][i] for cat in categories],
                bar_width,
                label=sheet,
                color=colors[i % len(colors)]
            )
            bars_all.append(bar)

        # 為每個 bar 添加數值標籤
        for bar_group in bars_all:
            self._check_cancel()
            for bar in bar_group:
                self._check_cancel()
                height = bar.get_height()
                plt.text(
                    bar.get_x() + bar.get_width() / 2,
                    height,
                    f'{height:.2f}%',
                    ha='center',
                    va='bottom'
                )
        # 添加图表元素
        plt.xlabel('Category')
        plt.ylabel('Percentage')
        plt.title('Values by Category and Sheet')
        plt.xticks(index + bar_width * len(sheet_labels) / 2, categories)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(title='Sheet Name')
        plt.tight_layout()
        bar_chart_2_path = self._chart_path("bar_chart_2.png")
        plt.savefig(bar_chart_2_path, bbox_inches='tight')
        # plt.show()

        # ------------------- 3. 將繪製好的圖儲存至self.context -------------------
        chart_1 = InlineImage(doc,
                        bar_chart_1_path,
                        width=Inches(5.83),
                        height=Inches(3.81))
        chart_2 = InlineImage(doc,
                        bar_chart_2_path,
                        width=Inches(5.83),
                        height=Inches(3.81))

        self.context['Chart_1'] = chart_1
        self.context['Chart_2'] = chart_2

    def process_insert_raw_data(self, file_name):
        """
        讀取、統整 'Raw Material' 工作表，並將統整結果插入至 Word 的對應標籤位置。
        
        Parameters
        ----------
        file_name : str
            Excel 檔案名稱 (如 'P_lci表單_tset.xlsx')。
        doc : docx.document.Document
            已讀取的 Word 檔案 Document 物件。
        
        Returns
        -------
        Raw_data : pandas.DataFrame
            統整後的 Raw Material 資料表。(即 all_data)
        """
        self._check_cancel()
        print("【Process_4】開始處理原材料數據")
        # (A) 改用通用的 process_worksheet
        # resulall_data_1 可以保留在需要的話使用，但主要我們只需要 all_data
        resulall_data_1, Raw_data = self.process_worksheet(file_name, 'Raw Material')

        # (B) 開始將 Raw_data 插入 Word
        raw_sum = Raw_data['Damage Assessment'].sum()
        self.context['Raw_total'] = self._format_report_number(raw_sum)

        for idx, row in Raw_data.head(10).reset_index(drop=True).iterrows():
            self._check_cancel()
            i = idx + 1  # 1-based index
            self.context[f'Raw_Name_{i}']              = row['Name']
            self.context[f'Raw_name_of_database_{i}']  = row['name of database']
            self.context[f'Raw_Damage_Assessment_{i}'] = self._format_report_number(row['Damage Assessment'])
            # 百分比
            pct = row['Damage Assessment'] / raw_sum * 100
            self.context[f'Raw_percentage_{i}']        = f"{round(pct, 2)}%"

        # （如果少於十筆，也可選擇把沒有資料的 key 先設成空字串）
        for i in range(len(Raw_data)+1, 11):
            self._check_cancel()
            self.context[f'Raw_Name_{i}']              = ""
            self.context[f'Raw_name_of_database_{i}']  = ""
            self.context[f'Raw_Damage_Assessment_{i}'] = ""
            self.context[f'Raw_percentage_{i}']        = ""

        # 將統整好的前十大Raw Material數值儲存至self.context
        remaining_val = Raw_data['Damage Assessment'][10:].sum()
        self.context['Remaining_processes_1'] = self._format_report_number(remaining_val)
        total_dmg = Raw_data['Damage Assessment'].sum()
        if total_dmg > 0:
            pct = remaining_val / total_dmg * 100
        else:
            pct = 0
        self.context['Remaining_percentage_1'] = f"{round(pct, 2)}%"
        
        
        return Raw_data

    def generate_insert_raw_charts(self, doc, Raw_data):
        """
        繪製 Raw Material 的前十大 Damage Assessment 長條圖與圓餅圖，
        並將產生的圖片插入 Word 中指定的標籤位置。

        Parameters
        ----------
        doc : docx.document.Document
            Word 文件的 Document 物件。
        Raw_data : pandas.DataFrame
            包含 'Name' 與 'Damage Assessment' 欄位的資料表。

        Returns
        -------
        None
            直接在函式內完成繪圖、儲存圖片與插入 Word 不返回任何值。
        """
        self._check_cancel()
        print("【Process_5】開始生成並插入原材料圖表")
        # ------------------ 1. 準備繪圖資料 ------------------
        name_values = Raw_data['Name'].head(10).fillna(0)
        damage_values = Raw_data['Damage Assessment'].head(10)

        remaining_name = 'Remaining processes'
        remaining_value = Raw_data['Damage Assessment'][10:].sum()

        # 如果剩餘值是 NaN，則改成 0
        if pd.isna(remaining_value):
            remaining_value = 0
        # ------------------ 2. 繪製長條圖 (bar_chart_3.png) ------------------
        colors = [
            '#e0e462', '#d9ed92', '#b5e48c', '#99d98c', '#76c893', 
            '#52b69a', '#34a0a4', '#168aad', '#1a759f', '#184e77', 'grey'
        ]

        plt.figure(figsize=(10, 6))
        bars = plt.bar(name_values, damage_values, color=colors[:-1])
        plt.bar(remaining_name, remaining_value, color=colors[-1])  # 顯示剩餘部分

        # 添加圖表標籤/標題
        plt.xlabel('Name')
        plt.ylabel('Damage Assessment')
        plt.title('Damage Assessment by Name')

        # 在每個 bar 上方顯示對應數值
        for i, bar in enumerate(bars):
            self._check_cancel()
            bar.set_label(name_values.iloc[i])
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval, self._format_report_number(yval),
                    ha='center', va='bottom')

        # 美化與保存
        plt.xticks(rotation=90)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(labels=list(name_values) + [remaining_name], loc='upper right')
        plt.tight_layout()
        bar_chart_3_path = self._chart_path("bar_chart_3.png")
        plt.savefig(bar_chart_3_path, bbox_inches='tight')
        # plt.show()

        # ------------------ 3. 繪製圓餅圖 (pie_chart_4.png) ------------------
        if len(name_values) < 10:
            labels = list(name_values)
            sizes = list(damage_values)
            # explode 陣列根據資料數量設定（第一塊稍微突起）
            explode = [0.01] + [0] * (len(name_values) - 1)
        else:
            labels = list(name_values) + [remaining_name]
            sizes = list(damage_values) + [remaining_value]
            explode = (0.01, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)  # 突出第一塊

        sizes = [0 if np.isnan(x) else x for x in sizes]

        # 過濾掉大小為 0 的項目，同時移除對應的 labels 與 explode
        filtered = [(lab, size, exp) for lab, size, exp in zip(labels, sizes, explode) if size != 0]
        if filtered:
            labels, sizes, explode = zip(*filtered)
            labels, sizes, explode = list(labels), list(sizes), list(explode)
        else:
            # 如果全部資料都為 0 或 NaN，可依需求處理，例如設定預設值
            labels = ['No Data']
            sizes = [1]
            explode = [0]
            
        # 檢查總和是否為 0，避免後續除法錯誤
        if sum(sizes) == 0:
            labels = ['No Data']
            sizes = [1]
            explode = [0]

        # 當沒有有效數據時，避免 `annotate()` 出錯
        only_no_data = (labels == ['No Data'])

        #繪製圓餅圖
        plt.figure(figsize=(5.83, 3.81))
        if len(sizes) == 1:
            # 只有一筆有效資料，直接用 autopct 標示圓心百分比
            wedges, texts, autotexts = plt.pie(
                sizes,
                explode=explode,
                labels=labels,
                colors=colors[:len(labels)],
                autopct=lambda pct: f"{pct:1.1f}%",
                startangle=180,
                wedgeprops={'width': 0.3, 'edgecolor': 'w', 'linewidth': 2}
            )
        else:
            wedges, texts, autotexts = plt.pie(
                sizes, 
                explode=explode, 
                colors=colors, 
                autopct='',  # 不在此使用autotext，由我們手動加上
                startangle=180,
                wedgeprops={'width': 0.3, 'edgecolor': 'w', 'linewidth': 2}
            )

        # 在每個 wedge 上加百分比標籤（帶箭頭）
        if not only_no_data:
            for i, wedge in enumerate(wedges):
                self._check_cancel()
                ang = (wedge.theta2 - wedge.theta1) / 2 + wedge.theta1
                x = wedge.r * 0.85 * np.cos(np.deg2rad(ang))
                y = wedge.r * 0.85 * np.sin(np.deg2rad(ang))

                percentage = f"{100 * sizes[i] / sum(sizes):1.1f}%"
                connectionstyle = f"angle,angleA=0,angleB={ang}"
                kw = dict(
                    arrowprops=dict(arrowstyle="->", connectionstyle=connectionstyle),
                    zorder=0, va="center"
                )
                plt.annotate(
                    percentage,
                    xy=(x, y),
                    xytext=(1.35 * np.sign(x), 1.4 * y),
                    textcoords='data',
                    horizontalalignment='center',
                    **kw
                )

        plt.axis('equal')  # 使圓餅圖保持為圓形
        plt.subplots_adjust(left=0.3, right=0.7)
        plt.title('Damage Assessment by Name (Pie Chart)')

        legend = plt.legend(labels, loc='upper right', bbox_to_anchor=(1.5, 1))
        if labels == ['No Data']:
            plt.title('No Data Available')  # 設定標題，避免 `tight_layout()` 崩潰
        else:
            plt.tight_layout()
        pie_chart_4_path = self._chart_path("pie_chart_4.png")
        plt.savefig(pie_chart_4_path, bbox_inches='tight')
        # plt.show()


        # ------------------ 4. 將繪製好的圖儲存至self.context ------------------

        chart_3 = InlineImage(doc,
                        bar_chart_3_path,
                        width=Inches(5.83),
                        height=Inches(3.81))
        chart_4 = InlineImage(doc,
                        pie_chart_4_path,
                        width=Inches(5.83),
                        height=Inches(3.81))

        self.context['Chart_3'] = chart_3
        self.context['Chart_4'] = chart_4

        print("【Process_5】Raw Material已匯入至報告書")

    def process_insert_manufacturing_data(self, file_name):
        """
        使用通用的 process_worksheet 函式處理 'Manufacturing' 表單，
        並將處理結果插入 Word 文件(doc)中的指定標籤。

        Parameters
        ----------
        file_name : str
            Excel 檔案名稱 (如 'P_lci表單_tset.xlsx')。
        doc : docx.document.Document
            已讀取的 Word 文件 Document 物件。

        Returns
        -------
        resulall_data_2 : dict
            以 {'G1': <DataFrame>, 'G2': <DataFrame>, ...} 形式存放的群組資料。
        Manu_data : pandas.DataFrame
            綜合所有群組的彙整資料 (Damage Assessment 降冪排序)。
        """
        self._check_cancel()
        print("【Process_6】開始處理製造數據")
        # 1. 呼叫通用函式 process_worksheet
        resulall_data_2, Manu_data = self.process_worksheet(file_name, 'Manufacturing')

        # 2. 用 Manu_data 插入 Word (表格標籤)

        Manu_sum = Manu_data['Damage Assessment'].sum()
        self.context['Manufacturing_total'] = self._format_report_number(Manu_sum)

        for idx, row in Manu_data.head(10).reset_index(drop=True).iterrows():
            self._check_cancel()
            i = idx + 1  # 1-based index
            self.context[f'Manufacturing_Name_{i}']              = row['Name']
            self.context[f'Manufacturing_name_of_database_{i}']  = row['name of database']
            self.context[f'Manufacturing_Damage_Assessment_{i}'] = self._format_report_number(row['Damage Assessment'])
            # 百分比
            pct = row['Damage Assessment'] / Manu_sum * 100
            self.context[f'Manufacturing_percentage_{i}']        = f"{round(pct, 2)}%"

        # （如果少於十筆，也可選擇把沒有資料的 key 先設成空字串）
        for i in range(len(Manu_data)+1, 11):
            self._check_cancel()
            self.context[f'Manufacturing_Name_{i}']              = ""
            self.context[f'Manufacturing_name_of_database_{i}']  = ""
            self.context[f'Manufacturing_Damage_Assessment_{i}'] = ""
            self.context[f'Manufacturing_percentage_{i}']        = ""


        remaining_val = Manu_data['Damage Assessment'][10:].sum()
        self.context['Remaining_processes_2'] = self._format_report_number(remaining_val)
        total_dmg = Manu_data['Damage Assessment'].sum()
        if total_dmg > 0:
            pct = remaining_val / total_dmg * 100
        else:
            pct = 0
        self.context['Remaining_percentage_2'] = f"{round(pct, 2)}%"

        print("【Process_6】已匯入Manufacturing表格資料")


        # 4. 回傳結果，若外部還需使用
        return resulall_data_2, Manu_data

    def generate_insert_manufacturing_charts(self, doc, Manu_data):
        """將Manufacturing的Manu_data數據繪製長條圖並匯入至Word"""
        self._check_cancel()
        print("【Process_7】開始生成並插入製造圖表")
        # 取得要繪製的資料，若缺值就以預設值替代
        name_values = Manu_data['Name'].head(10).fillna(0)
        damage_values = Manu_data['Damage Assessment'].head(10)

        remaining_name = 'Remaining processes'
        remaining_value = Manu_data['Damage Assessment'][10:].sum()
        # 如果剩餘值是 NaN，則改成 0
        if pd.isna(remaining_value):
            remaining_value = 0

        # 创建一个颜色列表，包含前十项的颜色和一个总和项的颜色
        colors = ['#e0e462', '#d9ed92', '#b5e48c', '#99d98c', '#76c893', '#52b69a', '#34a0a4', '#168aad', '#1a759f', '#184e77', 'grey']

        # 创建一个条形图
        plt.figure(figsize=(10, 6))  # 设置图表的大小
        bars = plt.bar(name_values, damage_values, color=colors)  # 创建条形图
        plt.bar(remaining_name, remaining_value, color='grey')  # 创建条形图
        # 添加标签和标题
        plt.xlabel('Name')  # x轴标签
        plt.ylabel('Damage Assessment')  # y轴标签
        plt.title('Damage Assessment by Name')  # 图表标题
        for i, bar in enumerate(bars):
            self._check_cancel()
            bar.set_label(name_values.iloc[i])
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2, yval, self._format_report_number(yval), ha='center', va='bottom')
        # 旋转x轴标签，以避免重叠
        plt.xticks(rotation=90)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(labels=list(name_values) + ['Remaining processes'], loc='upper right')
        # 顯示圖表
        plt.tight_layout()# 調整佈局，確保標籤和標題不重疊
        bar_chart_5_path = self._chart_path("bar_chart_5.png")
        plt.savefig(bar_chart_5_path, bbox_inches='tight')
        # plt.show()



        # 繪製一個餅圖
        if len(name_values) < 10:
            labels = list(name_values)
            sizes = list(damage_values)
            # explode 陣列根據資料數量設定（第一塊稍微突起）
            explode = [0.01] + [0] * (len(name_values) - 1)
        else:
            labels = list(name_values) + ['Remaining processes']
            sizes = list(damage_values) + [remaining_value]
            explode = [0] * len(name_values) + [0.2]  # 如果你想要突出某个块，可以设置它的值大于0
        
        sizes = [0 if np.isnan(x) else x for x in sizes]

        # 過濾掉大小為 0 的項目，同時移除對應的 labels 與 explode
        filtered = [(lab, size, exp) for lab, size, exp in zip(labels, sizes, explode) if size != 0]
        if filtered:
            labels, sizes, explode = zip(*filtered)
            labels, sizes, explode = list(labels), list(sizes), list(explode)
        else:
            # 如果全部資料都為 0 或 NaN，可依需求處理，例如設定預設值
            labels = ['No Data']
            sizes = [1]
            explode = [0]

        # 檢查總和是否為 0，避免後續除法錯誤
        if sum(sizes) == 0:
            # 如果所有值都是 0，可以給一個預設值，或跳出錯誤處理
            labels = ['No Data']
            sizes = [1]
            explode = [0]
            
        only_no_data = (labels == ['No Data']) # 當沒有有效數據時，避免 `annotate()` 出錯
        
        #繪製圓餅圖
        plt.figure(figsize=(8, 6))
        
        if len(sizes) == 1:
            # 只有一筆有效資料，直接用 autopct 標示圓心百分比
            wedges, texts, autotexts = plt.pie(
                sizes,
                explode=explode,
                labels=labels,
                colors=colors[:len(labels)],
                autopct=lambda pct: f"{pct:1.1f}%",
                startangle=180,
                wedgeprops={'width': 0.3, 'edgecolor': 'w', 'linewidth': 2}
            )
        else:
            wedges, texts, autotexts = plt.pie(
                sizes, 
                explode=explode, 
                colors=colors, 
                autopct='', 
                startangle=180, 
                wedgeprops={'width': 0.3, 'edgecolor': 'w', 'linewidth': 2}
            )
        if not only_no_data:
        # 為每個區塊添加注釋（包裝在 try/except 中以防個別失敗）
            for i, wedge in enumerate(wedges):
                self._check_cancel()
                ang = (wedge.theta2 - wedge.theta1) / 2 + wedge.theta1
                x = wedge.r * 0.85 * np.cos(np.deg2rad(ang))
                y = wedge.r * 0.85 * np.sin(np.deg2rad(ang))
                percentage = f"{100 * sizes[i] / sum(sizes):1.1f}%"
                connectionstyle = f"angle,angleA=0,angleB={ang}"# 設定指針樣式
                kw = dict(
                    arrowprops=dict(arrowstyle="->", connectionstyle=connectionstyle),
                    zorder=0, va="center"
                )
                
                # 新增註釋
                plt.annotate(
                    percentage,
                    xy=(x, y),
                    xytext=(1.35*np.sign(x), 1.4*y),
                    textcoords='data',
                    horizontalalignment='center',  # 水平居中对齐
                    **kw
                )

        plt.axis('equal')  # 使得圓餅圖是正圓的
        plt.subplots_adjust(left=0.3, right=0.7)
        plt.title('Damage Assessment by Name (Pie Chart)')# 新增標題
        legend = plt.legend(labels, loc='upper right', bbox_to_anchor=(1.5, 1))# 新增圖例

        # 顯示圖表
        if labels == ['No Data']:
            plt.title('No Data Available')  # 設定標題，避免 `tight_layout()` 崩潰
        else:
            plt.tight_layout()
        pie_chart_6_path = self._chart_path("pie_chart_6.png")
        plt.savefig(pie_chart_6_path) 
        # plt.show()
        print("【Process_7】已完成製造圖表生成與插入")  
        #--------------------------6. 將繪製好的圖儲存至self.context---------------------------

        chart_5 = InlineImage(doc,
                        bar_chart_5_path,
                        width=Inches(5.83),
                        height=Inches(3.81))
        chart_6 = InlineImage(doc,
                        pie_chart_6_path,
                        width=Inches(5.83),
                        height=Inches(3.81))

        self.context['Chart_5'] = chart_5
        self.context['Chart_6'] = chart_6

        print("【Process_7】Manufacturing已匯入至報告書")

    def generate_and_insert_electric_chart(self, doc, resulall_data_2):
        """
        從 resulall_data_2['G3'] 取得電力數據，繪製水平方向的長條圖，並將圖片插入 Word 的 [Chart_8] 標籤處。

        Parameters
        ----------
        doc : docx.document.Document
            Word 文件的 Document 物件。
        resulall_data_2 : dict
            包含多個群組資料的字典，必須存在 'G3' 這個鍵。 
            例如: resulall_data_2['G3'] => 需要包含 'Name' 與 'Damage Assessment' 欄位的 DataFrame。
        
        Returns
        -------
        None
            直接在函式內完成繪圖並插入圖片，不回傳任何值。
        """
        self._check_cancel()
        print("【Process_8】開始生成並插入電力圖表")
        # 1. 取得電力資料 (G3 群組)
        if 'G3' not in resulall_data_2:
            raise KeyError("resulall_data_2 中沒有 'G3' 群組，無法繪製電力數據圖表。")

        elec_data = resulall_data_2['G3'].copy()
        elec_data["Damage Assessment"] = pd.to_numeric(
            elec_data["Damage Assessment"],
            errors="coerce",
        ).fillna(0)
        elec_data = elec_data.sort_values(by='Damage Assessment', ascending=False)
        self.insert_electric_data_to_word(elec_data)

        # 若需要檢查 grouped_d，可視需求加上
        grouped_d = elec_data.groupby('name of database')['Name'].apply(' ; '.join).reset_index()
        
        print("電力群組資料 (G3) 結構：")
        print(elec_data.head())  # 可自行檢查前幾筆
        
        # 2. 繪製長條圖 (bar_chart_8.png)
        name_values = elec_data['Name'].head(10).fillna(0)
        damage_values = elec_data['Damage Assessment'].head(10)
        
        colors = [
            '#e0e462', '#d9ed92', '#b5e48c', '#99d98c', '#76c893',
            '#52b69a', '#34a0a4', '#168aad', '#1a759f', '#184e77', 'grey'
        ]
        
        plt.figure(figsize=(10, 6))
        bars = plt.barh(name_values, damage_values, color=colors[:len(name_values)])
        plt.xlabel('Name')
        plt.ylabel('Damage Assessment')
        plt.title('電力碳排')

        # 在每個長條顯示數值
        for i, bar in enumerate(bars):
            self._check_cancel()
            val = bar.get_width()  # bar.get_width() 對應 x 軸長度(因為是 barh)
            plt.text(val, bar.get_y() + bar.get_height() / 2,
                    self._format_report_number(val),
                    va='center')

        plt.xticks(rotation=90)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(labels=list(name_values), loc='upper right')
        plt.tight_layout()
        bar_chart_8_path = self._chart_path("bar_chart_8.png")
        plt.savefig(bar_chart_8_path, bbox_inches='tight')
        # plt.show()

        
        chart_8 = InlineImage(doc,
            bar_chart_8_path,
            width=Inches(5.83),
            height=Inches(3.81))

        self.context['Chart_8'] = chart_8

        print("【Process_8】已完成電力圖表生成與插入")

    def insert_electric_data_to_word(self, elec_data):
        """
        將電力單筆數值匯入 Word 文檔中的表格欄位。

        Parameters
        ----------
        elec_data : pandas.DataFrame
            Manufacturing(Electricity) 統整後的電力資料。
        """
        self._check_cancel()
        print("【Process_8】開始將電力數據匯入 Word 文件")

        elec_data = elec_data.copy()
        if "Damage Assessment" not in elec_data.columns:
            elec_data["Damage Assessment"] = 0
        if "Coefficient value" not in elec_data.columns:
            elec_data["Coefficient value"] = ""
        elec_data["Damage Assessment"] = pd.to_numeric(
            elec_data["Damage Assessment"],
            errors="coerce",
        ).fillna(0)
        elec_data = elec_data.sort_values(by="Damage Assessment", ascending=False)

        # 清空以前的 Electric_* keys，避免同一個物件重複產報告時殘留舊值。
        for k in list(self.context):
            self._check_cancel()
            if k.startswith("Electric_"):
                del self.context[k]

        self.context["Electric_Name"] = ""
        self.context["Electric_name_of_database"] = ""
        self.context["Electric_Damage_Assessment"] = ""
        self.context["Electric_Coefficient_value"] = ""
        self.context["Electric_Coefficient value"] = ""
        self.context["Electric_percentage"] = ""
        self.context["Electric_Name_1"] = ""
        self.context["Electric_name_of_database_1"] = ""
        self.context["Electric_Damage_Assessment_1"] = ""
        self.context["Electric_percentage_1"] = ""

        total = elec_data["Damage Assessment"].sum()
        if not elec_data.empty:
            self._check_cancel()
            row = elec_data.iloc[0]
            damage_value = row.get("Damage Assessment", 0)
            coefficient_value = row.get("Coefficient value", "")
            pct = (damage_value / total * 100) if total else 0
            self.context["Electric_Name"] = row.get("Name", "")
            self.context["Electric_name_of_database"] = row.get("name of database", "")
            self.context["Electric_Damage_Assessment"] = self._format_report_number(damage_value)
            self.context["Electric_Coefficient_value"] = self._format_report_number(coefficient_value)
            self.context["Electric_Coefficient value"] = self.context["Electric_Coefficient_value"]
            self.context["Electric_percentage"] = f"{pct:.2f}%"

            # Keep the first-index placeholders working for older templates.
            self.context["Electric_Name_1"] = self.context["Electric_Name"]
            self.context["Electric_name_of_database_1"] = self.context["Electric_name_of_database"]
            self.context["Electric_Damage_Assessment_1"] = self.context["Electric_Damage_Assessment"]
            self.context["Electric_percentage_1"] = self.context["Electric_percentage"]

        remaining_sum = elec_data["Damage Assessment"].iloc[1:].sum()
        remaining_pct = remaining_sum / total * 100 if total else 0
        self.context["Remaining_processes_5"] = self._format_report_number(remaining_sum)
        self.context["Remaining_percentage_5"] = f"{remaining_pct:.2f}%"

        print("【Process_8】已匯入電力統計表格數值")

    def insert_electric_top10_to_word(self, elec_data):
        """Backward compatible wrapper for older callers."""
        return self.insert_electric_data_to_word(elec_data)

    def process_top10_data(self, sheet_names, input_file, doc):
        """
        前十大數值統整並匯入 Word 文檔的函數。

        Parameters:
        - sheet_names: list, 所有工作表的名稱。
        - input_file: str, Excel 文件名稱。
        - doc: Document, Word 文件對象。

        Returns:
        - combined_all_data: DataFrame, 統整的前十大數值數據。
        """
        self._check_cancel()
        print("【Process_9】開始生成前十大數據長條圖")
        combined_all_data = pd.DataFrame()
        all_results = {}

        # 處理每個工作表數據
        for sheet in sheet_names:
            self._check_cancel()
            resulall_data, all_data = self.process_worksheet(input_file, sheet)
            all_results[sheet] = {'resulall_data': resulall_data, 'all_data': all_data}

        # 合併所有工作表的數據
        for sheet, data in all_results.items():
            self._check_cancel()
            combined_all_data = pd.concat([combined_all_data, data['all_data']], axis=0)

        # 按照 'Damage Assessment' 列進行排序
        combined_all_data = combined_all_data.sort_values(by='Damage Assessment', ascending=False)
        print(combined_all_data.head())

        # 匯入前十大數據到 Word 表格
        self.insert_top10_to_word(combined_all_data)

        # 繪製長條圖並匯入 Word
        self.top10_bar_chart(combined_all_data, doc)

        return combined_all_data

    def insert_top10_to_word(self, combined_all_data):
        """
        將前十大數值匯入到 Word 文檔中的表格和段落。

        Parameters:
        - doc: Document, Word 文件對象。
        - combined_all_data: DataFrame, 統整的前十大數據。
        - all_results: dict, 全階段的處理數據。
        """
        self._check_cancel()
        print("【Process_10】開始將前十大數據匯入 Word 文件")
        
        # 1. 前十大 Name, name_of_database, Damage_Assessment 與 percentage
        total_damage = combined_all_data["Damage Assessment"].sum() if "Damage Assessment" in combined_all_data.columns else 0
        for j in range(1, 11):
            self._check_cancel()
            idx = j - 1
            if idx >= len(combined_all_data):
                self.context[f"Top10_Name_{j}"] = ""
                self.context[f"Top10_name_of_database_{j}"] = ""
                self.context[f"Top10_Damage_Assessment_{j}"] = ""
                self.context[f"Top10_percentage_{j}"] = ""
                continue
            row = combined_all_data.iloc[idx]
            # 名稱
            self.context[f"Top10_Name_{j}"] = row["Name"]
            # 對應的 database 字串
            self.context[f"Top10_name_of_database_{j}"] = row["name of database"]
            # Damage Assessment 四位小數
            self.context[f"Top10_Damage_Assessment_{j}"] = f"{row['Damage Assessment']:.4f}"
            # 百分比：該筆 / 總和 *100，保留兩位小數
            pct = row["Damage Assessment"] / total_damage * 100 if total_damage else 0
            self.context[f"Top10_percentage_{j}"] = f"{pct:.2f}%"


        # 2. 剩餘製程合計與百分比（從第 11 筆開始到最後）
        remaining_sum = combined_all_data["Damage Assessment"].iloc[10:].sum()
        remaining_pct = remaining_sum / total_damage * 100 if total_damage else 0
        self.context["Remaining_processes_3"]   = f"{remaining_sum:.4f}"
        self.context["Remaining_percentage_3"] = f"{remaining_pct:.2f}%"

        print("【Process_10】已匯入前十大統計表格數值")

    def top10_bar_chart(self, combined_all_data, doc):
        """
        繪製前十大數據的長條圖並匯入 Word 文檔。

        Parameters:
        - combined_all_data: DataFrame, 統整的前十大數據。
        - doc: Document, Word 文件對象。
        """
        self._check_cancel()
        print("【Process_11】開始生成前十大數據長條圖")
        name_values = combined_all_data['Name'].head(10)
        damage_values = combined_all_data['Damage Assessment'].head(10)

        remaining_name = 'Remaining processes'
        remaining_value = combined_all_data['Damage Assessment'][10:].sum()

        colors = ['#e0e462', '#d9ed92', '#b5e48c', '#99d98c', '#76c893', '#52b69a', '#34a0a4', '#168aad', '#1a759f', '#184e77', 'grey']

        plt.figure(figsize=(10, 6))
        bars = plt.bar(name_values, damage_values, color=colors[:len(name_values)])
        plt.bar(remaining_name, remaining_value, color='grey')

        plt.xlabel('Name')
        plt.ylabel('Damage Assessment')
        plt.title('Damage Assessment by Name')
        for i, bar in enumerate(bars):
            self._check_cancel()
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval, round(yval, 4), ha='center', va='bottom')

        plt.xticks(rotation=90)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(labels=list(name_values) + ['Remaining processes'], loc='upper right')
        plt.tight_layout()

        bar_chart_path = self._chart_path("bar_chart_7.png")
        plt.savefig(bar_chart_path, bbox_inches='tight')
        plt.close()
        # 將繪製好的圖儲存至self.context
        chart_7 = InlineImage(doc,
                        bar_chart_path,
                        width=Inches(5.83),
                        height=Inches(3.81))

        self.context['Chart_7'] = chart_7

        print("【Process_11】已完成前十大數據長條圖生成")

    def process_transport_data(self, file_name, transport_sheets):
        """
        處理運輸相關的數據，整合多個工作表並進行分析。

        Parameters:
            file_name (str): Excel 檔案名稱。
            transport_sheets (list): 包含工作表名稱的列表。

        Returns:
            dict: 每個工作表的分組結果。
            DataFrame: 合併後的所有數據。
        """
        self._check_cancel()
        print("【Process_12】開始處理運輸數據")
        # transport_all_results = {}
        if not hasattr(self, "context") or self.context is None:
            self.context = {}
        stage_air_totals = {stage_name: 0.0 for stage_name in AIR_STAGE_CONTEXT_KEYS}
        Air_all_data = pd.DataFrame(
            columns=["type of transport", "Name", "Damage Assessment", "name of database"]
        )

        for sheet_name in transport_sheets:
            self._check_cancel()
            sheet_df = pd.read_excel(file_name, sheet_name=sheet_name)
            group_starts = sheet_df.index[sheet_df.iloc[:, 1].str.contains('^◎', na=False)].tolist()
            # resulall_data_3 = {}

            for j in range(len(group_starts)):
                self._check_cancel()
                start_idx = group_starts[j]
                end_idx = group_starts[j + 1] if j < len(group_starts) - 1 else sheet_df.shape[0]
                sub_df = sheet_df.iloc[start_idx:end_idx, :]

                # 清理數據
                sub_df = sub_df.iloc[2:, 1:]
                sub_df.columns = sub_df.iloc[0, :]
                sub_df = sub_df.iloc[1:, :]
                
                required_columns = {'type of transport', 'Name', 'Damage Assessment'}
                if not required_columns.issubset(set(sub_df.columns)):
                    continue
                if 'name of database' not in sub_df.columns:
                    sub_df['name of database'] = ""
                transport_type = sub_df['type of transport'].astype(str).str.strip().str.casefold()
                df_air = sub_df[transport_type.isin(AIR_TRANSPORT_TYPES)].copy()
                if df_air.empty:
                    continue    
                df_air['type of transport'] = 'Air'
                df_air['Name'] = df_air['Name'].fillna('空白群組')
                df_air['name of database'] = df_air['name of database'].fillna("")
                df_air['Damage Assessment'] = pd.to_numeric(
                    df_air['Damage Assessment'],
                    errors='coerce',
                ).fillna(0)
                if sheet_name in stage_air_totals:
                    stage_air_totals[sheet_name] += df_air['Damage Assessment'].sum()

                # 分組和統計
                transport_grouped = df_air.groupby(['type of transport', 'Name'])
                summed_values = transport_grouped['Damage Assessment'].sum().reset_index(name='Damage Assessment')
                dbnames = transport_grouped['name of database'] \
                        .agg(lambda x: ';'.join(sorted(set(x.dropna())))) \
                        .reset_index(name='name of database')
                # data_frames = [grouped_c, fossil_values, biogenic_values, land_values, damage_values, database_names]
                merged = pd.merge(summed_values, dbnames,
                            on=['type of transport','Name'],
                            how='outer')
                merged= merged.sort_values(by='Damage Assessment', ascending=False)
                # resulall_data_3[f'G{j + 1}'] = merged
                Air_all_data = pd.concat([Air_all_data, merged], axis=0)

        Air_all_data["Damage Assessment"] = pd.to_numeric(
            Air_all_data["Damage Assessment"],
            errors="coerce",
        ).fillna(0)
        Air_all_data = Air_all_data.sort_values(by='Damage Assessment', ascending=False)

        for stage_name, context_key in AIR_STAGE_CONTEXT_KEYS.items():
            self.context[context_key] = self._format_report_number(stage_air_totals[stage_name])

        # ---- 新增：清空以前的 Air_* keys（若有的話） ----
        for k in list(self.context):
            self._check_cancel()
            if k.startswith('Air_'):
                del self.context[k]
        for idx in range(1, 11):
            self.context[f'Air_Name_{idx}'] = ""
            self.context[f'Air_name_of_database_{idx}'] = ""
            self.context[f'Air_Damage_Assessment_{idx}'] = ""
            self.context[f'Air_percentage_{idx}'] = ""
        # ---- 新增：把 merged 的每一列放到 self.context  ----
        total = Air_all_data['Damage Assessment'].sum()
        for idx, (_, row) in enumerate(Air_all_data.head(10).iterrows(), start=1):
            self._check_cancel()
            damage_value = pd.to_numeric(row.get('Damage Assessment', 0), errors='coerce')
            damage_value = 0 if pd.isna(damage_value) else damage_value
            self.context[f'Air_Name_{idx}']              = row.get('Name', "")
            self.context[f'Air_name_of_database_{idx}']  = row.get('name of database', "")
            self.context[f'Air_Damage_Assessment_{idx}'] = self._format_report_number(damage_value)
            # 百分比四捨五入到小數點 2 位
            pct = (damage_value / total * 100) if total else 0
            self.context[f'Air_percentage_{idx}']        = f"{pct:.2f}%"

        # 2. 剩餘製程合計與百分比（從第 11 筆開始到最後）
        remaining_sum = Air_all_data["Damage Assessment"].iloc[10:].sum()
        # remaining_pct = remaining_sum / Air_all_data["Damage Assessment"].sum() * 100

        if total == 0:
            remaining_pct = 0.0
        else:
            remaining_pct = remaining_sum / total * 100

        self.context["Remaining_processes_4"]   = self._format_report_number(remaining_sum)
        self.context["Remaining_percentage_4"] = f"{remaining_pct:.2f}%"

        print("【Process_12】已完成運輸數據處理")
        return Air_all_data

    def analyze_and_chart_generate(self, Air_all_data, doc):
        """
        分析合併後的運輸數據，生成報告並插入 Word 文件。

        Parameters:
            transport_all_results (dict): 每個工作表的數據結果。
            doc (Document): Word 文件對象。
            output_image (str): 長條圖保存的文件名。
            output_doc (str): Word 文件保存的文件名。

        Returns:
            None
        """
        self._check_cancel()
        print("【Process_13】開始生成運輸相關圖表並插入 Word 文件")
        # 分析運輸數據
        name_values = Air_all_data['Name'].head(10)
        damage_values = Air_all_data['Damage Assessment'].head(10)

        remaining_name = 'Remaining processes'
        remaining_value = Air_all_data['Damage Assessment'][10:].sum()

        if Air_all_data.empty:
            print("No air transport data available.")
            return

        # 生成長條圖
        name_values = Air_all_data['Name'].head(10).fillna(0)
        damage_values = Air_all_data['Damage Assessment'].head(10)

        remaining_name = 'Remaining processes'
        remaining_value = Air_all_data['Damage Assessment'][10:].sum()

        # 如果剩餘值是 NaN，則改成 0
        if pd.isna(remaining_value):
            remaining_value = 0

        # 建立一個顏色列表，包含前十項的顏色和一個總和項的顏色
        colors = [
            '#e0e462', '#d9ed92', '#b5e48c', '#99d98c', '#76c893', 
            '#52b69a', '#34a0a4', '#168aad', '#1a759f', '#184e77', 'grey'
        ]

        # 建立一個長條圖
        plt.figure(figsize=(10, 6))  # 設定圖表的大小
        bars = plt.bar(name_values, damage_values, color=colors)  # 建立長條圖
        plt.bar(remaining_name, remaining_value, color='grey')  # 建立長條圖

        # 添加標籤和標題
        plt.xlabel('Name')  # x軸標籤
        plt.ylabel('Damage Assessment')  # y軸標籤
        plt.title('運輸碳排')  # 圖表標題

        for i, bar in enumerate(bars):
            self._check_cancel()
            bar.set_label(name_values.iloc[i])
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval, round(yval, 4), ha='center', va='bottom')

        # 旋轉x軸標籤，以避免重叠
        plt.xticks(rotation=45)
        plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
        plt.legend(labels=list(name_values) + ['Remaining processes'], loc='upper right')

        # 顯示圖表
        plt.tight_layout()  # 調整佈局，確保標籤和標題不重疊
        bar_chart_9_path = self._chart_path("bar_chart_9.png")
        plt.savefig(bar_chart_9_path, bbox_inches='tight')
        # plt.show()
        # 將繪製好的圖儲存至self.context
        chart_9 = InlineImage(doc,
                        bar_chart_9_path,
                        width=Inches(5.83),
                        height=Inches(3.81))

        self.context['Chart_9'] = chart_9

        print("【Process_13】已完成運輸相關圖表生成與插入")

    def update_progress_smooth(self, start, end, step=1, delay=0.05):
        """
        從 start 到 end 平滑更新進度，
        每次增加 step，延遲 delay 秒（單位秒）。
        """
        self._check_cancel()
        if self.progress_callback:
            # 確保整數更新
            for value in range(start, end + 1, step):
                self._check_cancel()
                self.progress_callback(value)
                time.sleep(delay)
















